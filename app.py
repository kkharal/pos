from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_cors import CORS
import json
import os
import re
from datetime import datetime, timedelta
from functools import wraps
from database import init_db, get_db_connection, hash_password
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import smtplib
from html import unescape
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import atexit
import fcntl
import logging
import time

app = Flask(__name__)
_secret = os.environ.get('SECRET_KEY')
if not _secret:
    raise RuntimeError("SECRET_KEY environment variable is required. Generate one with: python3 -c \"import secrets; print(secrets.token_hex(32))\"")
app.secret_key = _secret
CORS(app, origins=os.environ.get('ALLOWED_ORIGINS', 'http://localhost:5000').split(','))

# Build/version token that can be used in templates for cache-busting query params.
ASSET_VERSION = os.environ.get('ASSET_VERSION') or datetime.utcnow().strftime('%Y%m%d%H%M%S')

# Product image upload settings
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'products')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'webp'}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5 MB


def normalize_category_label(value):
    """Normalize common category variants to a single display label."""
    if value is None:
        return ''

    text = str(value).strip()
    if not text:
        return ''

    normalized = re.sub(r'[-_\s]+', ' ', text).strip().lower()
    if normalized in {'co ord set', 'co ord sets', 'co ord', 'coord set', 'coord sets', 'co ord set'}:
        return 'Co-ord Sets'

    return text


@app.context_processor
def inject_asset_version():
    return {'asset_version': ASSET_VERSION}


@app.after_request
def apply_cache_headers(response):
    """
    Force clients to revalidate HTML and static assets so stale mobile caches refresh.
    This is intentionally strict to support emergency cache-reset scenarios.
    """
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response

    if response.mimetype == 'text/html':
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

    return response


@app.before_request
def enforce_shop_selection_for_mutations():
    """Block write operations when super_admin/shop_owner is in all-shops mode."""
    if request.method not in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        return None

    role = session.get('role')
    if role not in ('super_admin', 'shop_owner'):
        return None

    if get_current_shop_id() is not None:
        return None

    allowed_paths = {
        '/api/shops',
        '/api/shops/switch',
        '/api/change-password',
    }
    if request.path in allowed_paths:
        return None

    return jsonify({'success': False, 'error': 'Please select a shop before performing this action'}), 403


# Initialize database on startup
init_db()

# Configure logging for scheduler
logging.basicConfig()
logging.getLogger('apscheduler').setLevel(logging.INFO)

# Automatic Low Stock Check Function
def scheduled_low_stock_check():
    """
    Scheduled job that checks for low stock and sends email alerts.
    Loops over each shop and uses that shop's own email settings.
    Falls back to global (shop_id=NULL) settings if no per-shop config is set.
    Runs at configured times (default 09:00, 12:00, 18:00).
    """
    try:
        print(f"[SCHEDULER {datetime.now()}] Running scheduled low stock check...")

        conn = get_db_connection()
        cursor = conn.cursor()

        shops = cursor.execute('SELECT id, name FROM shops ORDER BY id').fetchall()
        conn.close()

        for shop in shops:
            _run_low_stock_check_for_shop(shop['id'], shop['name'])

        print(f"[SCHEDULER] ✓ Completed per-shop low stock checks for {len(shops)} shop(s)")

    except Exception as e:
        print(f"[SCHEDULER] ✗ Error during scheduled check: {str(e)}")


def _run_low_stock_check_for_shop(shop_id, shop_name):
    """Run low stock check and send email alert for a single shop."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        email_settings = get_shop_email_settings(cursor, shop_id)
        alert_email = email_settings.get('alert_email', '')
        threshold = int(email_settings.get('low_stock_threshold') or 5)
        
        # Fetch actual scheduled times for this shop
        times_row = cursor.execute(
            'SELECT value FROM settings WHERE `key` = ? AND shop_id = ?', ('scheduler_times', shop_id)
        ).fetchone()
        scheduled_times = json.loads(times_row['value']) if times_row else []

        # Fetch timezone for this shop (fallback to global, then UTC)
        tz_row = cursor.execute(
            'SELECT value FROM settings WHERE `key` = ? AND shop_id = ?', ('scheduler_timezone', shop_id)
        ).fetchone()
        timezone = tz_row['value'] if tz_row and tz_row['value'] else None
        
        # If no shop-specific times, try global fallback
        if not scheduled_times:
            g_times = cursor.execute(
                'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', ('scheduler_times',)
            ).fetchone()
            scheduled_times = json.loads(g_times['value']) if g_times else ["09:00", "12:00", "18:00"]

        if not timezone:
            g_tz = cursor.execute(
                'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', ('scheduler_timezone',)
            ).fetchone()
            timezone = g_tz['value'] if g_tz and g_tz['value'] else 'UTC'

        if not alert_email:
            print(f"[SCHEDULER] Shop '{shop_name}' (id={shop_id}): no alert email configured, skipping")
            conn.close()
            return

        low_stock_products = cursor.execute('''
                        SELECT id, name, sku, category, size, color, stock_quantity,
                   COALESCE(low_stock_threshold, ?) as threshold
            FROM products
            WHERE shop_id = ?
                            AND is_active = 1
              AND stock_quantity <= COALESCE(low_stock_threshold, ?)
            ORDER BY stock_quantity ASC
        ''', (threshold, shop_id, threshold)).fetchall()

        if not low_stock_products:
            print(f"[SCHEDULER] Shop '{shop_name}': no low stock items found")
            conn.close()
            return

        # Dedupe guard: only one send per shop per alert slot (date+time in shop timezone).
        slot_date, slot_time = _get_current_slot_for_timezone(timezone)
        claimed = _claim_alert_slot(
            cursor,
            shop_id,
            'scheduled_low_stock',
            slot_date,
            slot_time,
            timezone,
        )
        conn.commit()
        conn.close()

        if not claimed:
            print(
                f"[SCHEDULER] Duplicate blocked for shop '{shop_name}' "
                f"at slot {slot_date} {slot_time} ({timezone})"
            )
            return

        # Build email body
        products_html = '<table class="low-stock-table" style="width: 100%; border-collapse: collapse; margin-top: 20px; table-layout: fixed;">'
        products_html += '<tr class="low-stock-head" style="background-color: #f8f9fa; border-bottom: 2px solid #dee2e6;">'
        products_html += '<th style="padding: 12px; text-align: left; width: 40%;">Product</th>'
        products_html += '<th style="padding: 12px; text-align: left; width: 20%;">SKU</th>'
        products_html += '<th style="padding: 12px; text-align: center; width: 14%;">Size</th>'
        products_html += '<th style="padding: 12px; text-align: left; width: 16%;">Color</th>'
        products_html += '<th style="padding: 12px; text-align: center; width: 10%;">Stock</th>'
        products_html += '</tr>'

        plain_lines = [
            f'Scheduled Low Stock Alert - {shop_name}',
            '',
            f'Total low stock items: {len(low_stock_products)}',
            '',
            'Product | SKU | Size | Color | Stock',
            '-' * 90,
        ]

        for product in low_stock_products:
            products_html += '<tr class="low-stock-row" style="border-bottom: 1px solid #dee2e6;">'
            products_html += f'<td class="low-stock-cell" data-label="Product" style="padding: 10px; word-break: break-word; overflow-wrap: anywhere;">{product["name"]}</td>'
            products_html += f'<td class="low-stock-cell" data-label="SKU" style="padding: 10px; word-break: break-word; overflow-wrap: anywhere;">{product["sku"] or "-"}</td>'
            products_html += f'<td class="low-stock-cell" data-label="Size" style="padding: 10px; text-align: center; word-break: break-word;">{product["size"] or "-"}</td>'
            products_html += f'<td class="low-stock-cell" data-label="Color" style="padding: 10px; word-break: break-word; overflow-wrap: anywhere;">{product["color"] or "-"}</td>'
            products_html += f'<td class="low-stock-cell low-stock-value" data-label="Stock" style="padding: 10px; text-align: center; color: #e74c3c; font-weight: bold;">{product["stock_quantity"]}</td>'
            products_html += '</tr>'
            plain_lines.append(
                f'{product["name"]} | {product["sku"] or "-"} | {product["size"] or "-"} | '
                f'{product["color"] or "-"} | {product["stock_quantity"]}'
            )

        products_html += '</table>'
        plain_text = '\n'.join(plain_lines)
        
        # Build schedule message based on actual configured times
        if len(scheduled_times) == 1:
            schedule_msg = f"Scheduled checks run at {scheduled_times[0]} daily."
        elif len(scheduled_times) == 2:
            schedule_msg = f"Scheduled checks run at {scheduled_times[0]} and {scheduled_times[1]} daily."
        else:
            schedule_msg = f"Scheduled checks run at {', '.join(scheduled_times[:-1])} and {scheduled_times[-1]} daily."

        email_body = f'''
        <html>
        <head>
            <style>
                @media only screen and (max-width: 600px) {{
                    .email-wrap {{ padding: 12px !important; }}
                    .low-stock-table {{ table-layout: auto !important; }}
                    .low-stock-head {{ display: none !important; }}
                    .low-stock-row {{
                        display: block !important;
                        border: 1px solid #e5e7eb !important;
                        border-radius: 8px !important;
                        margin-bottom: 10px !important;
                    }}
                    .low-stock-cell {{
                        display: block !important;
                        width: 100% !important;
                        box-sizing: border-box !important;
                        text-align: left !important;
                        padding: 8px 10px !important;
                        border-bottom: 1px solid #f1f5f9 !important;
                    }}
                    .low-stock-cell:last-child {{ border-bottom: none !important; }}
                    .low-stock-cell::before {{
                        content: attr(data-label) ': ';
                        font-weight: 700;
                        color: #374151;
                    }}
                    .low-stock-value {{ color: #e74c3c !important; font-weight: 700 !important; }}
                }}
            </style>
        </head>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div class="email-wrap" style="max-width: 800px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #e74c3c; border-bottom: 3px solid #e74c3c; padding-bottom: 10px;">
                    ⚠️ Scheduled Low Stock Alert
                </h2>
                <p>The following products in <strong>{shop_name}</strong> are running low on stock and need to be reordered:</p>
                <p><strong>Total low stock items: {len(low_stock_products)}</strong></p>
                {products_html}
                <div style="margin-top: 30px; padding: 15px; background-color: #fff3cd; border-left: 4px solid #ffc107;">
                    <p style="margin: 0;"><strong>Note:</strong> Please restock these items as soon as possible to avoid running out of inventory.</p>
                </div>
                <p style="margin-top: 20px; color: #6c757d; font-size: 0.9em;">
                    This is an automated alert from {shop_name}.<br>
                    {schedule_msg}
                </p>
            </div>
        </body>
        </html>
        '''

        try:
            send_email(
                email_settings,
                alert_email,
                f'{shop_name} - Scheduled Low Stock Alert ({len(low_stock_products)} items)',
                email_body,
                from_name=shop_name,
                text_body=plain_text
            )
            _mark_alert_slot_sent(
                shop_id,
                'scheduled_low_stock',
                slot_date,
                slot_time,
                timezone,
                len(low_stock_products),
            )
        except Exception:
            # Release the claim if send fails so a retry can still send.
            _release_alert_slot(
                shop_id,
                'scheduled_low_stock',
                slot_date,
                slot_time,
                timezone,
            )
            raise

        print(f"[SCHEDULER] ✓ Alert sent for shop '{shop_name}': {len(low_stock_products)} low stock item(s)")

    except Exception as e:
        print(f"[SCHEDULER] ✗ Error for shop '{shop_name}' (id={shop_id}): {str(e)}")


def _get_current_slot_for_timezone(timezone_name):
    """Return (YYYY-MM-DD, HH:MM) in the provided timezone."""
    import pytz

    try:
        tz = pytz.timezone(timezone_name or 'UTC')
    except Exception:
        tz = pytz.timezone('UTC')
        timezone_name = 'UTC'

    now_local = datetime.now(tz)
    return now_local.strftime('%Y-%m-%d'), now_local.strftime('%H:%M')


def _claim_alert_slot(cursor, shop_id, alert_type, slot_date, slot_time, timezone_name):
    """Try to claim a scheduled alert slot. Returns True only for first claimant."""
    cursor.execute(
        '''
        INSERT IGNORE INTO alert_send_log (
            shop_id, alert_type, slot_date, slot_time, timezone_name
        ) VALUES (?, ?, ?, ?, ?)
        ''',
        (shop_id, alert_type, slot_date, slot_time, timezone_name),
    )
    return cursor.rowcount == 1


def _mark_alert_slot_sent(shop_id, alert_type, slot_date, slot_time, timezone_name, item_count):
    """Mark claimed slot as sent with item count."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            '''
            UPDATE alert_send_log
            SET sent_at = NOW(), item_count = ?
            WHERE shop_id = ?
              AND alert_type = ?
              AND slot_date = ?
              AND slot_time = ?
              AND timezone_name = ?
            ''',
            (item_count, shop_id, alert_type, slot_date, slot_time, timezone_name),
        )
        conn.commit()
    finally:
        conn.close()


def _release_alert_slot(shop_id, alert_type, slot_date, slot_time, timezone_name):
    """Release claimed slot when send fails, allowing retry."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            '''
            DELETE FROM alert_send_log
            WHERE shop_id = ?
              AND alert_type = ?
              AND slot_date = ?
              AND slot_time = ?
              AND timezone_name = ?
              AND sent_at IS NULL
            ''',
            (shop_id, alert_type, slot_date, slot_time, timezone_name),
        )
        conn.commit()
    finally:
        conn.close()


# ── Recurring Expense Auto-Creation ──────────────────────────────────────────

def scheduled_recurring_expenses():
    """
    Scheduled job that processes recurring expenses.
    Checks all active recurring expenses across all shops.
    If next_due_date <= today, creates an expense record and advances the due date.
    Runs once daily at 00:05.
    """
    try:
        print(f"[SCHEDULER {datetime.now()}] Processing recurring expenses...")

        conn = get_db_connection()
        cursor = conn.cursor()
        today = datetime.now().strftime('%Y-%m-%d')

        # Find all active recurring expenses that are due
        due_items = cursor.execute('''
            SELECT re.*, ec.name as category_name
            FROM recurring_expenses re
            LEFT JOIN expense_categories ec ON re.category_id = ec.id
            WHERE re.is_active = 1 AND re.next_due_date <= ?
        ''', (today,)).fetchall()

        created_count = 0
        for item in due_items:
            try:
                # Create the expense record
                cursor.execute('''
                    INSERT INTO expenses (category_id, amount, description, expense_date, payment_method, receipt_ref, created_by, shop_id)
                    VALUES (?, ?, ?, ?, 'cash', ?, NULL, ?)
                ''', (
                    item['category_id'],
                    item['amount'],
                    f"[Auto] {item['description'] or item['category_name'] or 'Recurring expense'}",
                    item['next_due_date'],
                    f"REC-{item['id']}",
                    item['shop_id']
                ))
                expense_id = cursor.lastrowid

                _record_finance_transaction(
                    cursor,
                    shop_id=item['shop_id'],
                    direction='OUT',
                    amount=item['amount'],
                    transaction_type='expense_payment',
                    source_table='expenses',
                    source_id=expense_id,
                    reference=f'EXP-{expense_id}',
                    notes='Auto recurring expense',
                    transaction_at=item['next_due_date'],
                )

                # Advance next_due_date based on frequency
                next_date = _advance_date(item['next_due_date'], item['frequency'])
                cursor.execute(
                    'UPDATE recurring_expenses SET next_due_date = ? WHERE id = ?',
                    (next_date, item['id'])
                )
                created_count += 1

            except Exception as e:
                print(f"[SCHEDULER] ✗ Error processing recurring expense id={item['id']}: {str(e)}")

        conn.commit()
        conn.close()
        print(f"[SCHEDULER] ✓ Created {created_count} expense(s) from {len(due_items)} due recurring item(s)")

    except Exception as e:
        print(f"[SCHEDULER] ✗ Error during recurring expense processing: {str(e)}")


def _advance_date(date_str, frequency):
    """Advance a date string by the given frequency. Returns new date string."""
    from dateutil.relativedelta import relativedelta

    if isinstance(date_str, str):
        current = datetime.strptime(date_str, '%Y-%m-%d')
    else:
        current = datetime.combine(date_str, datetime.min.time())

    if frequency == 'daily':
        next_dt = current + relativedelta(days=1)
    elif frequency == 'weekly':
        next_dt = current + relativedelta(weeks=1)
    elif frequency == 'monthly':
        next_dt = current + relativedelta(months=1)
    elif frequency == 'quarterly':
        next_dt = current + relativedelta(months=3)
    elif frequency == 'yearly':
        next_dt = current + relativedelta(years=1)
    else:
        next_dt = current + relativedelta(months=1)

    return next_dt.strftime('%Y-%m-%d')


# Initialize the background scheduler
scheduler = BackgroundScheduler()

# Tracks last known per-shop config: {shop_id: {'times': [...], 'timezone': '...'}}
_last_scheduler_config = {}

def _read_all_shop_configs():
    """Read scheduler times and timezone for every shop from DB.
    Falls back to global (shop_id IS NULL) settings if a shop has none set.
    Returns {shop_id: {'times': [...], 'timezone': '...'}}"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        shops = cursor.execute('SELECT id, name FROM shops ORDER BY id').fetchall()

        # Global fallback values
        g_tz = cursor.execute(
            'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', ('scheduler_timezone',)
        ).fetchone()
        g_times = cursor.execute(
            'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', ('scheduler_times',)
        ).fetchone()
        global_tz = g_tz['value'] if g_tz and g_tz['value'] else 'UTC'
        global_times = json.loads(g_times['value']) if g_times else ["09:00", "12:00", "18:00"]

        result = {}
        for shop in shops:
            sid = shop['id']
            tz_row = cursor.execute(
                'SELECT value FROM settings WHERE `key` = ? AND shop_id = ?', ('scheduler_timezone', sid)
            ).fetchone()
            times_row = cursor.execute(
                'SELECT value FROM settings WHERE `key` = ? AND shop_id = ?', ('scheduler_times', sid)
            ).fetchone()
            result[sid] = {
                'name': shop['name'],
                'timezone': tz_row['value'] if tz_row and tz_row['value'] else global_tz,
                'times': json.loads(times_row['value']) if times_row else global_times,
            }
        conn.close()
        return result
    except Exception as e:
        print(f"[SCHEDULER] Error reading configs: {e}")
        return {}

def load_scheduler_jobs():
    """Rebuild all low-stock-check jobs from per-shop DB config.
    Only modifies jobs if the scheduler is running in this process."""
    global _last_scheduler_config
    if not scheduler.running:
        print("[SCHEDULER] load_scheduler_jobs called in non-scheduler worker, skipping")
        return

    configs = _read_all_shop_configs()

    try:
        # Remove all existing low stock check jobs
        for job in scheduler.get_jobs():
            if job.id.startswith('low_stock_check_'):
                scheduler.remove_job(job.id)

        # Add one job per (shop, time) combination
        for shop_id, cfg in configs.items():
            timezone = cfg['timezone']
            for idx, time_str in enumerate(cfg['times']):
                try:
                    hour, minute = map(int, time_str.split(':'))
                    job_id = f'low_stock_check_{shop_id}_{idx}'
                    scheduler.add_job(
                        func=_run_low_stock_check_for_shop,
                        args=[shop_id, cfg['name']],
                        trigger=CronTrigger(hour=hour, minute=minute, timezone=timezone),
                        id=job_id,
                        name=f'Low Stock Check - {cfg["name"]} {time_str}',
                        replace_existing=True
                    )
                    print(f"[SCHEDULER] Shop '{cfg['name']}': job {time_str} ({timezone})")
                except Exception as e:
                    print(f"[SCHEDULER] Error adding job {time_str} for shop {shop_id}: {e}")

        _last_scheduler_config = {sid: {'times': c['times'], 'timezone': c['timezone']} for sid, c in configs.items()}
        print(f"[SCHEDULER] Loaded jobs for {len(configs)} shop(s)")

    except Exception as e:
        print(f"[SCHEDULER] Error loading jobs: {e}")

def _scheduler_config_watcher():
    """Runs every minute. Reloads jobs if any shop's DB config changed."""
    global _last_scheduler_config
    configs = _read_all_shop_configs()
    changed = False
    for sid, cfg in configs.items():
        last = _last_scheduler_config.get(sid, {})
        if cfg['times'] != last.get('times') or cfg['timezone'] != last.get('timezone'):
            changed = True
            break
    if changed or set(configs.keys()) != set(_last_scheduler_config.keys()):
        print("[SCHEDULER] Config change detected — reloading jobs")
        load_scheduler_jobs()

# Load scheduler jobs and start the scheduler (only in one worker process)
_scheduler_lock_file = None

def _start_scheduler_once():
    """Start scheduler in only one process using a file lock to prevent duplicates in multi-worker setups."""
    global _scheduler_lock_file
    try:
        _scheduler_lock_file = open('/tmp/pos_scheduler.lock', 'w')
        fcntl.flock(_scheduler_lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except (IOError, OSError):
        if _scheduler_lock_file:
            _scheduler_lock_file.close()
            _scheduler_lock_file = None
        print("[SCHEDULER] Scheduler already running in another worker, skipping")
        return

    configs = _read_all_shop_configs()

    # Add one job per (shop, time) combination
    for shop_id, cfg in configs.items():
        timezone = cfg['timezone']
        for idx, time_str in enumerate(cfg['times']):
            try:
                hour, minute = map(int, time_str.split(':'))
                job_id = f'low_stock_check_{shop_id}_{idx}'
                scheduler.add_job(
                    func=_run_low_stock_check_for_shop,
                    args=[shop_id, cfg['name']],
                    trigger=CronTrigger(hour=hour, minute=minute, timezone=timezone),
                    id=job_id,
                    name=f'Low Stock Check - {cfg["name"]} {time_str}',
                    replace_existing=True
                )
                print(f"[SCHEDULER] Shop '{cfg['name']}': job {time_str} ({timezone})")
            except Exception as e:
                print(f"[SCHEDULER] Error adding job {time_str} for shop {shop_id}: {e}")

    _last_scheduler_config.update({sid: {'times': c['times'], 'timezone': c['timezone']} for sid, c in configs.items()})

    # Add recurring expense processor - runs daily at 00:05
    scheduler.add_job(
        func=scheduled_recurring_expenses,
        trigger=CronTrigger(hour=0, minute=5),
        id='recurring_expenses_daily',
        name='Recurring Expenses - Daily 00:05',
        replace_existing=True
    )
    print("[SCHEDULER] Added recurring expense job: daily at 00:05")

    # Config watcher — picks up DB changes within 1 minute automatically
    scheduler.add_job(
        func=_scheduler_config_watcher,
        trigger=CronTrigger(minute='*'),
        id='config_watcher',
        name='Scheduler Config Watcher',
        replace_existing=True
    )
    print("[SCHEDULER] Added config watcher job")

    scheduler.start()
    print("[SCHEDULER] Background scheduler started")

    atexit.register(lambda: scheduler.shutdown())

_start_scheduler_once()


def _touch_user_activity():
    """Best-effort heartbeat for currently authenticated user activity."""
    user_id = session.get('user_id')
    if not user_id:
        return
    conn = None
    try:
        conn = get_db_connection()
        conn.execute(
            '''
            UPDATE users
            SET last_activity = NOW()
            WHERE id = ?
              AND (last_activity IS NULL OR last_activity < DATE_SUB(NOW(), INTERVAL 1 MINUTE))
            ''',
            (user_id,),
        )
        conn.commit()
    except Exception:
        # Never block request flow for heartbeat updates.
        pass
    finally:
        if conn:
            conn.close()

# Decorator for login required
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        _touch_user_activity()
        return f(*args, **kwargs)
    return decorated_function

# Decorator for admin only
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') not in ('admin', 'super_admin', 'shop_owner'):
            return jsonify({'success': False, 'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

# Decorator for super_admin only
def super_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'super_admin':
            return jsonify({'success': False, 'error': 'Super admin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

# ── Shop-scope helpers ────────────────────────────────────────────────────────

def get_current_shop_id():
    """Return the shop_id to scope queries to.
    Returns None for super_admin in all-shops mode (no filter applied).
    For shop_owner without active selection, returns None (filtered via shop_filter)."""
    role = session.get('role')
    if role == 'super_admin':
        return session.get('active_shop_id')  # None → all-shops mode
    if role == 'shop_owner':
        return session.get('active_shop_id')  # None → all assigned shops mode
    return session.get('shop_id')

def _get_user_shop_ids():
    """Return list of shop IDs assigned to current shop_owner user."""
    user_id = session.get('user_id')
    if not user_id:
        return []
    conn = get_db_connection()
    rows = conn.execute(
        'SELECT shop_id FROM user_shops WHERE user_id = ?', (user_id,)
    ).fetchall()
    conn.close()
    return [r['shop_id'] for r in rows]

def shop_filter(alias=''):
    """Return (sql_snippet, params_list) to append to WHERE clauses.
    alias: optional table alias, e.g. 's', 'p', 'sh' (dot added automatically).
    Returns ('', []) when super_admin is viewing all shops."""
    role = session.get('role')
    shop_id = get_current_shop_id()

    # shop_owner with no active shop selected → filter to assigned shops
    if role == 'shop_owner' and shop_id is None:
        shop_ids = _get_user_shop_ids()
        if not shop_ids:
            # No shops assigned — show nothing
            col = f'{alias}.shop_id' if alias else 'shop_id'
            return f' AND {col} = ?', [-1]
        col = f'{alias}.shop_id' if alias else 'shop_id'
        placeholders = ','.join(['?'] * len(shop_ids))
        return f' AND {col} IN ({placeholders})', shop_ids

    if shop_id is None:
        return '', []
    col = f'{alias}.shop_id' if alias else 'shop_id'
    return f' AND {col} = ?', [shop_id]


def _as_float(value, default=0.0):
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _get_opening_balance(cursor):
    """Return configured opening balance across current shop scope."""
    flt_sql, flt_params = shop_filter('st')
    row = cursor.execute(
        f'''
            SELECT COALESCE(SUM(CAST(st.value AS DECIMAL(14,2))), 0) as opening_balance
            FROM settings st
            WHERE st.`key` = 'finance_opening_balance' {flt_sql}
        ''',
        flt_params,
    ).fetchone()
    return _as_float(row['opening_balance'] if row else 0)


def _record_finance_transaction(
    cursor,
    *,
    shop_id,
    direction,
    amount,
    transaction_type,
    source_table=None,
    source_id=None,
    reference=None,
    notes=None,
    created_by=None,
    transaction_at=None,
):
    """Insert a cash movement row into finance ledger with idempotency guard."""
    direction = (direction or '').upper().strip()
    if direction not in ('IN', 'OUT'):
        raise ValueError('direction must be IN or OUT')

    amount_value = round(_as_float(amount), 2)
    if amount_value <= 0:
        return

    cursor.execute(
        '''
            INSERT INTO finance_transactions
                (shop_id, direction, amount, transaction_type, source_table, source_id, reference, notes, created_by, transaction_at)
            VALUES
                (?, ?, ?, ?, ?, ?, ?, ?, ?, COALESCE(?, NOW()))
            ON DUPLICATE KEY UPDATE id = id
        ''',
        (
            shop_id,
            direction,
            amount_value,
            (transaction_type or '').strip(),
            (source_table or '').strip() or None,
            source_id,
            (reference or '').strip() or None,
            (notes or '').strip() or None,
            created_by,
            transaction_at,
        ),
    )

    audit_payload = {
        'shop_id': shop_id,
        'direction': direction,
        'amount': amount_value,
        'transaction_type': (transaction_type or '').strip(),
        'source_table': (source_table or '').strip() or None,
        'source_id': source_id,
        'reference': (reference or '').strip() or None,
        'notes': (notes or '').strip() or None,
    }
    _record_finance_audit(
        cursor,
        shop_id=shop_id,
        action_type='finance_transaction_created',
        entity_type='finance_transaction',
        entity_id=cursor.lastrowid or None,
        amount=amount_value,
        direction=direction,
        reference=(reference or '').strip() or None,
        notes=(notes or '').strip() or None,
        details=audit_payload,
        created_by=created_by,
    )


def _record_finance_audit(
    cursor,
    *,
    shop_id,
    action_type,
    entity_type,
    entity_id=None,
    amount=None,
    direction=None,
    reference=None,
    notes=None,
    details=None,
    created_by=None,
):
    """Write an audit row for finance actions without affecting balances."""
    cursor.execute(
        '''
            INSERT INTO finance_audit_log
                (shop_id, action_type, entity_type, entity_id, amount, direction, reference, notes, details_json, created_by)
            VALUES
                (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (
            shop_id,
            (action_type or '').strip() or 'finance_action',
            (entity_type or '').strip() or 'finance',
            entity_id,
            round(_as_float(amount), 2) if amount is not None else None,
            (direction or '').upper().strip() or None,
            (reference or '').strip() or None,
            (notes or '').strip() or None,
            json.dumps(details or {}, default=str),
            created_by,
        ),
    )


def _get_available_funds(cursor, date_from=None, date_to=None):
    flt_sql, flt_params = shop_filter('ft')
    range_sql = ''
    params = list(flt_params)

    if date_from:
        range_sql += ' AND ft.transaction_at >= ?'
        params.append(f'{date_from} 00:00:00')
    if date_to:
        range_sql += ' AND ft.transaction_at <= ?'
        params.append(f'{date_to} 23:59:59')

    totals = cursor.execute(
        f'''
            SELECT
                COALESCE(SUM(CASE WHEN ft.direction = 'IN' THEN ft.amount ELSE 0 END), 0) as money_in,
                COALESCE(SUM(CASE WHEN ft.direction = 'OUT' THEN ft.amount ELSE 0 END), 0) as money_out
            FROM finance_transactions ft
            WHERE 1=1 {flt_sql} {range_sql}
        ''',
        params,
    ).fetchone()

    money_in = _as_float(totals['money_in'])
    money_out = _as_float(totals['money_out'])
    opening_balance = _get_opening_balance(cursor)
    return {
        'opening_balance': opening_balance,
        'money_in': money_in,
        'money_out': money_out,
        'available_funds': opening_balance + money_in - money_out,
    }


def normalize_color_value(value):
    """Normalize color text to a consistent display format."""
    if value is None:
        return None

    color = str(value).strip()
    if not color:
        return None

    # Normalize common separators so values are stored consistently.
    color = re.sub(r'[\-_/&]+', ' ', color)
    color = re.sub(r'\s+', ' ', color).strip().lower()

    phrase_map = {
        'b w': 'black white',
        'b and w': 'black white',
    }
    color = phrase_map.get(color, color)

    word_map = {
        'oragne': 'orange',
    }
    color = ' '.join(word_map.get(part, part) for part in color.split(' '))

    # Canonicalize token order so equivalent permutations map together
    # (e.g. "golden brown" and "brown golden").
    tokens = [t for t in color.split(' ') if t]
    # collapse duplicate tokens while preserving order ("blue blue dark" -> "blue dark")
    seen = set()
    uniq = []
    for t in tokens:
        if t not in seen:
            seen.add(t)
            uniq.append(t)
    tokens = uniq

    # Prefer modifier-first phrasing for common modifiers like 'dark'/'light'
    modifiers = {'dark', 'light', 'medium', 'pale'}
    if len(tokens) > 1:
        mod_tokens = [t for t in tokens if t in modifiers]
        non_mod = [t for t in tokens if t not in modifiers]
        if mod_tokens and non_mod:
            # e.g. ['blue','dark'] -> 'dark blue'
            color = ' '.join(mod_tokens + non_mod)
        else:
            color = ' '.join(sorted(tokens))

    if not color:
        return None

    return ' '.join(part.capitalize() for part in color.split(' '))


def color_normalized_key(value):
    """Return case-insensitive key used for duplicate detection."""
    normalized = normalize_color_value(value)
    return normalized.lower() if normalized else None


def size_normalized_key(value):
    """Return case-insensitive, whitespace-normalized size key."""
    if value is None:
        return None
    size = str(value).strip()
    return size.lower() if size else None


def normalize_product_name(name):
    """Normalize product name by collapsing multiple spaces and trimming."""
    if not name:
        return name
    return ' '.join(str(name).split())


def find_existing_variant_by_normalized_color(cursor, shop_id, name, size, color):
    """Find an existing active variant matching shop+name+size+normalized-color.
    Normalizes both incoming and DB product names to handle whitespace variance."""
    name_normalized = normalize_product_name(name)
    target_size_key = size_normalized_key(size)

    # Query active candidates in shop and compare normalized fields in Python.
    # This avoids misses from case/whitespace differences in DB values.
    candidates = cursor.execute(
        '''
        SELECT id, color, name, size
        FROM products
        WHERE shop_id = ?
          AND is_active = 1
        ''',
        (shop_id,)
    ).fetchall()

    target_key = color_normalized_key(color)
    for candidate in candidates:
        # Compare normalized product names
        if normalize_product_name(candidate['name']) == name_normalized:
            if size_normalized_key(candidate['size']) == target_size_key and color_normalized_key(candidate['color']) == target_key:
                return candidate

    return None

def get_settings_shop_id():
    """Return shop_id for settings operations.
    super_admin with no active shop → None (global settings).
    shop_owner with no active shop → first assigned shop.
    admin → their shop_id."""
    role = session.get('role')
    if role == 'super_admin' and not session.get('active_shop_id'):
        return None  # global settings
    if role == 'shop_owner' and not session.get('active_shop_id'):
        ids = _get_user_shop_ids()
        return ids[0] if ids else None
    return get_current_shop_id() or session.get('shop_id')


def _get_initial_shop_name_for_template():
    """Resolve a best-effort shop name for first paint in server-rendered templates."""
    if 'user_id' not in session:
        return 'POS System'

    role = session.get('role')
    active_shop_id = session.get('active_shop_id')
    shop_id = active_shop_id if role in ('super_admin', 'shop_owner') and active_shop_id else session.get('shop_id')

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if shop_id:
            shop = cursor.execute('SELECT name FROM shops WHERE id = ?', (shop_id,)).fetchone()
        else:
            # all-shops mode: pick first assigned/available shop for display
            if role == 'shop_owner':
                shop = cursor.execute(
                    'SELECT s.name FROM shops s '
                    'INNER JOIN user_shops us ON us.shop_id = s.id '
                    'WHERE us.user_id = ? ORDER BY s.id LIMIT 1',
                    (session.get('user_id'),)
                ).fetchone()
            else:
                shop = cursor.execute('SELECT name FROM shops ORDER BY id LIMIT 1').fetchone()
    finally:
        conn.close()

    return shop['name'] if shop and shop['name'] else 'POS System'


@app.context_processor
def inject_template_globals():
    return {
        'initial_shop_name': _get_initial_shop_name_for_template()
    }

# Authentication routes
@app.route('/api/public/shop-name', methods=['GET'])
def get_public_shop_name():
    """Public endpoint for brand name (no auth required for login page)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    # Return the global brand_name setting for the login page
    brand = cursor.execute(
        "SELECT value FROM settings WHERE `key` = 'brand_name' AND shop_id IS NULL"
    ).fetchone()
    conn.close()
    brand_name = brand['value'] if brand and brand['value'] else 'POS System'
    return jsonify({'shop_name': brand_name, 'shop_icon': '🛒'})

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')

    data = request.json
    username = data.get('username')
    password = data.get('password')

    conn = get_db_connection()
    user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()

    if user and user['password'] == hash_password(password):
        forwarded_for = request.headers.get('X-Forwarded-For', '')
        access_ip = (forwarded_for.split(',')[0].strip() if forwarded_for else None) or request.headers.get('X-Real-IP') or request.remote_addr
        user_agent = (request.headers.get('User-Agent') or '')[:500]

        conn.execute(
            '''
            UPDATE users
            SET last_login = NOW(),
                last_activity = NOW(),
                portal_access_count = COALESCE(portal_access_count, 0) + 1,
                last_portal_access_at = NOW(),
                last_access_ip = ?
            WHERE id = ?
            ''',
            (access_ip, user['id'])
        )
        conn.execute(
            '''
            INSERT INTO user_access_log (user_id, username, role, shop_id, access_ip, user_agent)
            VALUES (?, ?, ?, ?, ?, ?)
            ''',
            (user['id'], user['username'], user['role'], user['shop_id'], access_ip, user_agent)
        )
        conn.commit()
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = user['role']
        session['full_name'] = user['full_name']
        session['shop_id'] = user['shop_id']         # None for super_admin/shop_owner
        session['active_shop_id'] = None              # super_admin/shop_owner starts in all-shops mode
        conn.close()
        return jsonify({'success': True, 'role': user['role']})

    conn.close()
    return jsonify({'success': False, 'error': 'Invalid username or password'}), 401

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/current-user')
@login_required
def get_current_user():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Session timeout (global setting)
    timeout_row = cursor.execute(
        'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', ('session_timeout',)
    ).fetchone()
    session_timeout = int(timeout_row['value']) if timeout_row else 1800

    # Shop-specific settings
    role = session.get('role')
    active_shop_id = session.get('active_shop_id')
    shop_id = active_shop_id if role in ('super_admin', 'shop_owner') and active_shop_id else session.get('shop_id')

    if shop_id:
        shop = cursor.execute(
            'SELECT name, icon, currency_symbol FROM shops WHERE id = ?', (shop_id,)
        ).fetchone()
    else:
        # super_admin/shop_owner in all-shops mode – use first assigned/available shop as display default
        if role == 'shop_owner':
            shop = cursor.execute(
                'SELECT s.name, s.icon, s.currency_symbol FROM shops s '
                'INNER JOIN user_shops us ON us.shop_id = s.id WHERE us.user_id = ? ORDER BY s.id LIMIT 1',
                (session.get('user_id'),)
            ).fetchone()
        else:
            shop = cursor.execute(
                'SELECT name, icon, currency_symbol FROM shops ORDER BY id LIMIT 1'
            ).fetchone()

    conn.close()

    shop_name = shop['name'] if shop else 'POS System'
    shop_icon = shop['icon'] if shop else '🛒'
    currency = shop['currency_symbol'] if shop else '$'

    return jsonify({
        'username': session.get('username'),
        'role': role,
        'full_name': session.get('full_name'),
        'currency': currency,
        'shop_name': shop_name,
        'shop_icon': shop_icon,
        'shop_id': shop_id,
        'active_shop_id': active_shop_id,
        'session_timeout': session_timeout
    })

# Routes for pages
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/products')
@login_required
def products_page():
    return render_template('products.html')

@app.route('/pos')
@login_required
def pos_page():
    return render_template('pos.html')

@app.route('/sales')
@login_required
def sales_page():
    return render_template('sales.html')

@app.route('/reports')
@login_required
def reports_page():
    return render_template('reports.html')

@app.route('/settings')
@login_required
@admin_required
def settings_page():
    return render_template('settings.html')

@app.route('/users')
@login_required
@admin_required
def users_page():
    return render_template('users.html')

@app.route('/shops')
@login_required
def shops_page():
    if session.get('role') != 'super_admin':
        return redirect(url_for('index'))
    return render_template('shops.html')

# ── Shop management API (super_admin only) ────────────────────────────────────

@app.route('/api/shops', methods=['GET'])
@login_required
def get_shops():
    conn = get_db_connection()
    role = session.get('role')
    if role == 'super_admin':
        shops = conn.execute('''
            SELECT s.*, COUNT(u.id) as user_count
            FROM shops s
            LEFT JOIN users u ON u.shop_id = s.id
            GROUP BY s.id ORDER BY s.name
        ''').fetchall()
    elif role == 'shop_owner':
        user_id = session.get('user_id')
        shops = conn.execute('''
            SELECT s.*, COUNT(u.id) as user_count
            FROM shops s
            INNER JOIN user_shops us ON us.shop_id = s.id AND us.user_id = ?
            LEFT JOIN users u ON u.shop_id = s.id
            GROUP BY s.id ORDER BY s.name
        ''', (user_id,)).fetchall()
    else:
        shop_id = session.get('shop_id')
        shops = conn.execute('''
            SELECT s.*, COUNT(u.id) as user_count
            FROM shops s
            LEFT JOIN users u ON u.shop_id = s.id
            WHERE s.id = ?
            GROUP BY s.id
        ''', (shop_id,)).fetchall()

    result = [dict(s) for s in shops]

    # If stats requested, enrich with product_count, customer_count, month_sales
    if request.args.get('stats') == '1':
        first_of_month = datetime.now().strftime('%Y-%m-01')
        for shop in result:
            sid = shop['id']
            shop['product_count'] = conn.execute(
                'SELECT COUNT(*) FROM products WHERE shop_id = ? AND is_active = 1', (sid,)
            ).fetchone()[0]
            shop['customer_count'] = conn.execute(
                'SELECT COUNT(*) FROM customers WHERE shop_id = ?', (sid,)
            ).fetchone()[0]
            month_sales_row = conn.execute(
                'SELECT COALESCE(SUM(total_amount), 0) FROM sales WHERE shop_id = ? AND sale_date >= ?',
                (sid, first_of_month)
            ).fetchone()
            shop['month_sales'] = month_sales_row[0] if month_sales_row else 0

    conn.close()
    return jsonify(result)

@app.route('/api/shops', methods=['POST'])
@login_required
@super_admin_required
def create_shop():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Shop name is required'}), 400
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO shops (name, address, phone, icon, currency_symbol, low_stock_threshold) VALUES (?, ?, ?, ?, ?, ?)',
            (name, data.get('address', ''), data.get('phone', ''),
             data.get('icon', '🛒'), data.get('currency_symbol', '$'),
             int(data.get('low_stock_threshold', 5)))
        )
        new_shop_id = cursor.lastrowid

        # Seed default expense categories for the new shop
        default_categories = [
            ('Employee Salary', '👤', 'Staff wages and salaries'),
            ('Shop Rent', '🏠', 'Monthly shop rental'),
            ('Electricity / Utilities', '⚡', 'Electric, water, gas bills'),
            ('Transportation', '🚗', 'Delivery and travel costs'),
            ('Maintenance & Repairs', '🔧', 'Shop repairs and maintenance'),
            ('Packaging & Supplies', '📦', 'Bags, boxes, wrapping'),
            ('Marketing', '📣', 'Advertising and promotions'),
            ('Miscellaneous', '📝', 'Other expenses'),
        ]
        for cat_name, cat_icon, cat_desc in default_categories:
            cursor.execute(
                'INSERT INTO expense_categories (name, icon, description, shop_id) VALUES (?, ?, ?, ?)',
                (cat_name, cat_icon, cat_desc, new_shop_id)
            )

        conn.commit()
        return jsonify({'success': True, 'id': new_shop_id}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/shops/<int:shop_id>', methods=['PUT'])
@login_required
@admin_required
def update_shop(shop_id):
    # Admins can only update their own shop; shop_owner can update assigned shops
    role = session.get('role')
    if role == 'shop_owner':
        allowed = _get_user_shop_ids()
        if shop_id not in allowed:
            return jsonify({'success': False, 'error': 'Access denied'}), 403
    elif role != 'super_admin' and session.get('shop_id') != shop_id:
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Shop name is required'}), 400
    conn = get_db_connection()
    try:
        conn.execute(
            'UPDATE shops SET name=?, address=?, phone=?, icon=?, currency_symbol=?, low_stock_threshold=? WHERE id=?',
            (name, data.get('address', ''), data.get('phone', ''),
             data.get('icon', '🛒'), data.get('currency_symbol', '$'),
             int(data.get('low_stock_threshold', 5)), shop_id)
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/shops/<int:shop_id>', methods=['DELETE'])
@login_required
@super_admin_required
def delete_shop(shop_id):
    conn = get_db_connection()
    # Prevent deleting the last shop
    count = conn.execute('SELECT COUNT(*) FROM shops').fetchone()[0]
    if count <= 1:
        conn.close()
        return jsonify({'success': False, 'error': 'Cannot delete the last shop'}), 400
    # Check for users in this shop
    user_count = conn.execute('SELECT COUNT(*) FROM users WHERE shop_id = ?', (shop_id,)).fetchone()[0]
    if user_count > 0:
        conn.close()
        return jsonify({'success': False, 'error': f'Cannot delete shop with {user_count} user(s). Reassign users first.'}), 400
    try:
        conn.execute('DELETE FROM shops WHERE id = ?', (shop_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/shops/switch', methods=['POST'])
@login_required
def switch_shop():
    """Super_admin or shop_owner switches their active shop context."""
    role = session.get('role')
    if role not in ('super_admin', 'shop_owner'):
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    data = request.json
    shop_id = data.get('shop_id')  # None or int
    if shop_id is not None:
        shop_id = int(shop_id)
        # shop_owner can only switch to their assigned shops
        if role == 'shop_owner':
            allowed = _get_user_shop_ids()
            if shop_id not in allowed:
                return jsonify({'success': False, 'error': 'Access denied to this shop'}), 403
    session['active_shop_id'] = shop_id
    return jsonify({'success': True, 'active_shop_id': shop_id})

# API Routes - Products
@app.route('/api/products', methods=['GET'])
@login_required
def get_products():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    products = conn.execute(
        f'SELECT * FROM products WHERE is_active = 1 {flt_sql} ORDER BY name', flt_params
    ).fetchall()
    conn.close()

    products_list = [dict(row) for row in products]

    # Hide cost_price from non-admin users
    if session.get('role') not in ('admin', 'super_admin', 'shop_owner'):
        for product in products_list:
            product.pop('cost_price', None)

    # Attach image_url with cache-busting timestamp
    for product in products_list:
        ip = product.get('image_path')
        if ip:
            iu = product.get('image_updated_at')
            ts_qs = ''
            if iu:
                try:
                    ts_qs = '?v=' + str(int(datetime.strptime(str(iu), '%Y-%m-%d %H:%M:%S').timestamp()))
                except Exception:
                    pass
            product['image_url'] = f'/static/uploads/products/{ip}{ts_qs}'
        else:
            product['image_url'] = None

    return jsonify(products_list)

@app.route('/api/products', methods=['POST'])
@login_required
@admin_required
def add_product():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()

    # Auto-generate SKU if not provided
    sku = data.get('sku', '').strip()
    shop_id = get_current_shop_id()
    if not sku:
        prefix = data.get('category', 'GEN')[:3].upper()
        existing = cursor.execute(
            "SELECT COUNT(*) FROM products WHERE sku LIKE ? AND shop_id = ?", (f'{prefix}-%', shop_id)
        ).fetchone()[0]
        sku = f'{prefix}-{existing + 1:03d}'
        # Ensure uniqueness within this shop
        while cursor.execute("SELECT COUNT(*) FROM products WHERE sku=? AND shop_id=?", (sku, shop_id)).fetchone()[0] > 0:
            existing += 1
            sku = f'{prefix}-{existing + 1:03d}'

    try:
        name = normalize_product_name(data['name'].strip())
        category = data['category'].strip()
        size = data.get('size', '').strip() or None
        color = normalize_color_value(data.get('color', ''))

        sibling_group = cursor.execute(
            '''
            SELECT name, variant_group, category
            FROM products
            WHERE shop_id = ? AND is_active = 1 AND variant_group IS NOT NULL
            ORDER BY id
            LIMIT 100
            ''',
            (shop_id,)
        ).fetchall()

        # Find sibling by normalized name comparison
        existing_product = None
        variant_group = None
        for row in sibling_group:
            if normalize_product_name(row['name']) == name:
                variant_group = row['variant_group']
                if row['category']:
                    category = row['category']
                # Fetch existing product ID
                existing_product = cursor.execute(
                    'SELECT id FROM products WHERE shop_id = ? AND is_active = 1 AND name = ?',
                    (shop_id, row['name'])
                ).fetchone()
                break

        if not variant_group and (size or color):
            import uuid
            variant_group = f"{name[:20].lower().replace(' ', '-')}-{uuid.uuid4().hex[:8]}"

        # If product exists AND size/color provided, auto-add as variant
        if existing_product and (size or color):
            existing_variant = find_existing_variant_by_normalized_color(
                cursor,
                shop_id,
                name,
                size,
                color,
            )
            if existing_variant:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': f'This size/color combination already exists. Update its stock instead.'
                }), 409
            
            # Add variant to existing product group
            cursor.execute('''
                INSERT INTO products (shop_id, name, category, cost_price, price, stock_quantity, sku, description, size, color, variant_group)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (shop_id, existing_product['name'], category, data.get('cost_price', 0), data['price'],
                  data.get('stock_quantity', 0), sku, data.get('description', ''),
                  size, color, variant_group))

            product_id = cursor.lastrowid

            if data.get('stock_quantity', 0) > 0:
                cost_price = data.get('cost_price', 0)
                note = f'Initial stock | Cost: {cost_price:.2f}'
                cursor.execute('''
                    INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                    VALUES (?, ?, ?, ?, ?)
                ''', (shop_id, product_id, data['stock_quantity'], 'initial', note))

            conn.commit()
            return jsonify({'success': True, 'id': product_id, 'variant': True}), 201

        # Reject only if product exists but no size/color provided
        if existing_product and not size and not color:
            conn.close()
            return jsonify({
                'success': False,
                'error': f'A product named "{name}" already exists. Add a variant (with size/color) instead or update the existing product.'
            }), 409

        # Block duplicate variants
        if size or color:
            existing_variant = find_existing_variant_by_normalized_color(
                cursor,
                shop_id,
                name,
                size,
                color,
            )
            if existing_variant:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': f'A variant with color "{color or "-"}" and size "{size or "-"}" already exists for "{name}". Update its stock instead.'
                }), 409

        cursor.execute('''
            INSERT INTO products (shop_id, name, category, cost_price, price, stock_quantity, sku, description, size, color, variant_group)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (shop_id, name, category, data.get('cost_price', 0), data['price'],
              data.get('stock_quantity', 0), sku, data.get('description', ''),
              size, color, variant_group))

        product_id = cursor.lastrowid

        # Log stock history if initial stock > 0
        if data.get('stock_quantity', 0) > 0:
            cost_price = data.get('cost_price', 0)
            note = f'Initial stock | Cost: {cost_price:.2f}'
            cursor.execute('''
                INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                VALUES (?, ?, ?, ?, ?)
            ''', (shop_id, product_id, data['stock_quantity'], 'initial', note))

        conn.commit()
        return jsonify({'success': True, 'id': product_id}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/products/<int:product_id>', methods=['PUT'])
@login_required
@admin_required
def update_product(product_id):
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Fetch old product for audit log
        current_shop = get_current_shop_id()
        if current_shop:
            old = cursor.execute('SELECT name, category, cost_price, price, sku, description, low_stock_threshold, size, color, shop_id FROM products WHERE id=? AND shop_id=?', (product_id, current_shop)).fetchone()
        else:
            old = cursor.execute('SELECT name, category, cost_price, price, sku, description, low_stock_threshold, size, color, shop_id FROM products WHERE id=?', (product_id,)).fetchone()
        if not old:
            return jsonify({'success': False, 'error': 'Product not found'}), 404

        product_shop_id = old['shop_id']
        changes = []
        field_map = {
            'name': ('name', str),
            'category': ('category', str),
            'cost_price': ('cost_price', float),
            'price': ('price', float),
            'sku': ('sku', str),
            'description': ('description', str),
            'low_stock_threshold': ('low_stock_threshold', lambda v: int(v) if v is not None and v != '' else None),
            'size': ('size', lambda v: v.strip() if v else None),
            'color': ('color', lambda v: v.strip() if v else None),
        }
        for field, (db_col, conv) in field_map.items():
            if field in data:
                old_val = old[db_col] if db_col in old.keys() else None
                new_val = conv(data[field]) if data[field] is not None and data[field] != '' else (None if field == 'low_stock_threshold' else data[field])
                if str(old_val) != str(new_val):
                    changes.append((field, str(old_val), str(new_val)))

        cursor.execute('''
            UPDATE products
            SET name=?, category=?, cost_price=?, price=?, description=?, sku=?, low_stock_threshold=?, size=?, color=?
            WHERE id=?
        ''', (data['name'], data['category'], data.get('cost_price', 0), data['price'],
              data.get('description', ''), data.get('sku', ''),
              int(data['low_stock_threshold']) if data.get('low_stock_threshold') not in (None, '', 'null') else None,
              data.get('size', '').strip() or None, data.get('color', '').strip() or None,
              product_id))

        # Write audit log entries
        for field, old_val, new_val in changes:
            cursor.execute('''
                INSERT INTO product_audit_log (shop_id, product_id, user_id, username, action, field_name, old_value, new_value)
                VALUES (?, ?, ?, ?, 'edit', ?, ?, ?)
            ''', (product_shop_id, product_id, session.get('user_id'), session.get('username'), field, old_val, new_val))

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/products/group-update', methods=['PUT'])
@login_required
@admin_required
def update_product_group():
    """Update all variants of a product group (same name) at once.
    Updates: name, category, cost_price, price, description."""
    data = request.json
    old_name = data.get('old_name')
    new_name = data.get('name', '').strip()
    category = data.get('category', '').strip()
    cost_price = float(data.get('cost_price', 0))
    price = float(data.get('price', 0))
    description = data.get('description', '').strip()

    if not old_name or not new_name or not category:
        return jsonify({'success': False, 'error': 'Name and Category are required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        flt_sql, flt_params = shop_filter()
        # Get all variants in this group
        variants = cursor.execute(
            f'SELECT id, shop_id, name, category, cost_price, price, description FROM products WHERE name=? {flt_sql}',
            [old_name] + flt_params
        ).fetchall()

        if not variants:
            return jsonify({'success': False, 'error': 'Product group not found'}), 404

        # Determine old group price (use first variant as reference)
        old_group_price = float(variants[0]['price'] or 0)
        old_group_cost = float(variants[0]['cost_price'] or 0)

        # Update name, category, description for ALL variants
        cursor.execute(
            f'UPDATE products SET name=?, category=?, description=? WHERE name=? {flt_sql}',
            [new_name, category, description, old_name] + flt_params
        )

        # Update price/cost ONLY on non-overridden variants
        # A variant is "overridden" if its price differs from the old group price
        for v in variants:
            v_price = float(v['price'] or 0)
            v_cost = float(v['cost_price'] or 0)
            price_matches = abs(v_price - old_group_price) < 0.01
            cost_matches = abs(v_cost - old_group_cost) < 0.01

            updates = []
            params = []
            if price_matches and abs(v_price - price) > 0.001:
                updates.append('price=?')
                params.append(price)
            if cost_matches and abs(v_cost - cost_price) > 0.001:
                updates.append('cost_price=?')
                params.append(cost_price)
            if updates:
                params.append(v['id'])
                cursor.execute(f'UPDATE products SET {", ".join(updates)} WHERE id=?', params)

        # Audit log - record changes for first variant (represents the group)
        first = variants[0]
        changes = []
        if first['name'] != new_name:
            changes.append(('name', first['name'], new_name))
        if first['category'] != category:
            changes.append(('category', first['category'], category))
        if old_group_cost != cost_price:
            changes.append(('cost_price', str(first['cost_price']), str(cost_price)))
        if old_group_price != price:
            changes.append(('price', str(first['price']), str(price)))
        if (first['description'] or '') != description:
            changes.append(('description', first['description'] or '', description))

        for field, old_val, new_val in changes:
            for v in variants:
                cursor.execute('''
                    INSERT INTO product_audit_log (shop_id, product_id, user_id, username, action, field_name, old_value, new_value)
                    VALUES (?, ?, ?, ?, 'group_edit', ?, ?, ?)
                ''', (v['shop_id'], v['id'], session.get('user_id'), session.get('username'), field, old_val, new_val))

        conn.commit()
        return jsonify({'success': True, 'updated': len(variants)})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/products/<int:product_id>/variant-update', methods=['PUT'])
@login_required
@admin_required
def update_variant_detail(product_id):
    """Update variant-level fields: sku, low_stock_threshold, price/cost override."""
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        flt_sql, flt_params = shop_filter()
        product = cursor.execute(
            f'SELECT id, shop_id, sku, low_stock_threshold, cost_price, price FROM products WHERE id=? {flt_sql}',
            [product_id] + flt_params
        ).fetchone()
        if not product:
            return jsonify({'success': False, 'error': 'Variant not found'}), 404

        updates = []
        params = []

        if 'sku' in data:
            updates.append('sku=?')
            params.append(data['sku'].strip() if data['sku'] else None)
        if 'low_stock_threshold' in data:
            updates.append('low_stock_threshold=?')
            params.append(int(data['low_stock_threshold']) if data['low_stock_threshold'] not in (None, '', 'null') else None)
        if 'price_override' in data and data['price_override'] is not None:
            updates.append('price=?')
            params.append(float(data['price_override']))
        if 'cost_override' in data and data['cost_override'] is not None:
            updates.append('cost_price=?')
            params.append(float(data['cost_override']))

        if updates:
            params.append(product_id)
            cursor.execute(f'UPDATE products SET {", ".join(updates)} WHERE id=?', params)

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/products/<int:product_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_product(product_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql, flt_params = shop_filter()
    try:
        # Ensure variant exists in current shop scope.
        product = cursor.execute(
            f'SELECT id FROM products WHERE id=? AND is_active = 1 {flt_sql}',
            [product_id] + flt_params
        ).fetchone()
        if not product:
            conn.close()
            return jsonify({'success': False, 'error': 'Variant not found'}), 404

        # Soft delete to preserve historical reporting and transaction history.
        cursor.execute(
            f'UPDATE products SET is_active = 0 WHERE id=? AND is_active = 1 {flt_sql}',
            [product_id] + flt_params
        )

        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

# API Routes - Product Image
@app.route('/api/products/<int:product_id>/image', methods=['POST', 'DELETE'])
@login_required
@admin_required
def product_image(product_id):
    from PIL import Image as PILImage, ImageOps
    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql, flt_params = shop_filter()
    try:
        product = cursor.execute(
            f'SELECT id, name, shop_id, image_path FROM products WHERE id=? AND is_active=1 {flt_sql}',
            [product_id] + flt_params
        ).fetchone()
        if not product:
            return jsonify({'success': False, 'error': 'Product not found'}), 404

        shop_id = product['shop_id']

        if request.method == 'DELETE':
            if product['image_path']:
                file_path = os.path.join(UPLOAD_FOLDER, product['image_path'])
                if os.path.exists(file_path):
                    os.remove(file_path)
            cursor.execute(
                'UPDATE products SET image_path=NULL, image_updated_at=NULL WHERE name=? AND shop_id=?',
                (product['name'], shop_id)
            )
            conn.commit()
            return jsonify({'success': True})

        # POST — upload
        if 'image' not in request.files:
            return jsonify({'success': False, 'error': 'No image file provided'}), 400
        f = request.files['image']
        if not f or not f.filename:
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ALLOWED_IMAGE_EXTENSIONS:
            return jsonify({'success': False, 'error': 'Invalid file type. Allowed: jpg, jpeg, png, webp'}), 400
        allowed_mime = {'image/jpeg', 'image/png', 'image/webp'}
        if f.content_type and f.content_type.split(';')[0].strip() not in allowed_mime:
            return jsonify({'success': False, 'error': 'Invalid file content type'}), 400

        data = f.read(MAX_IMAGE_SIZE + 1)
        if len(data) > MAX_IMAGE_SIZE:
            return jsonify({'success': False, 'error': 'Image too large. Maximum size is 5 MB'}), 400

        # Use min variant ID in the group as stable filename anchor
        row = cursor.execute(
            'SELECT MIN(id) as min_id FROM products WHERE name=? AND shop_id=? AND is_active=1',
            (product['name'], shop_id)
        ).fetchone()
        anchor_id = row['min_id'] if row and row['min_id'] else product_id
        filename = f'{anchor_id}.webp'
        dest = os.path.join(UPLOAD_FOLDER, filename)

        # Convert, center-crop to square, resize to 800×800, save as WebP
        img = PILImage.open(BytesIO(data))
        img = ImageOps.exif_transpose(img)
        img = img.convert('RGB')
        w, h = img.size
        side = min(w, h)
        img = img.crop(((w - side) // 2, (h - side) // 2, (w + side) // 2, (h + side) // 2))
        img = img.resize((800, 800), PILImage.LANCZOS)
        img.save(dest, 'WEBP', quality=82, method=4)

        now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute(
            'UPDATE products SET image_path=?, image_updated_at=? WHERE name=? AND shop_id=?',
            (filename, now, product['name'], shop_id)
        )
        conn.commit()
        ts = str(int(time.time()))
        return jsonify({'success': True, 'image_url': f'/static/uploads/products/{filename}?v={ts}'})

    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


# API Routes - Stock Management
@app.route('/api/products/<int:product_id>/stock', methods=['POST'])
@login_required
@admin_required
def update_stock(product_id):
    data = request.json
    quantity_change = data['quantity_change']
    action_type = data['action_type']  # 'import' or 'export'
    note = data.get('note', '')

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Get current stock (scoped to shop)
        flt_sql, flt_params = shop_filter()
        product = cursor.execute(
            f'SELECT stock_quantity, shop_id FROM products WHERE id=? AND is_active = 1 {flt_sql}',
            [product_id] + flt_params
        ).fetchone()
        if not product:
            return jsonify({'success': False, 'error': 'Product not found'}), 404

        current_stock = product['stock_quantity']
        product_shop_id = product['shop_id']
        new_stock = current_stock + quantity_change

        if new_stock < 0:
            return jsonify({'success': False, 'error': 'Insufficient stock'}), 400

        # Update stock
        cursor.execute(
            f'UPDATE products SET stock_quantity=? WHERE id=? {flt_sql}',
            [new_stock, product_id] + flt_params
        )

        # Log history
        cursor.execute('''
            INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
            VALUES (?, ?, ?, ?, ?)
        ''', (product_shop_id, product_id, quantity_change, action_type, note))

        conn.commit()
        return jsonify({'success': True, 'new_stock': new_stock})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/stock-history', methods=['GET'])
@login_required
@admin_required
def get_stock_history():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('sh')
    history = conn.execute(f'''
        SELECT sh.*, p.name as product_name, p.sku
        FROM stock_history sh
        JOIN products p ON sh.product_id = p.id
        WHERE 1=1 {flt_sql}
        ORDER BY sh.created_at DESC
        LIMIT 100
    ''', flt_params).fetchall()
    conn.close()
    return jsonify([dict(row) for row in history])

@app.route('/api/products/<int:product_id>/stock-history', methods=['GET'])
@login_required
def get_product_stock_history(product_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Get product info (scoped to shop)
    flt_sql, flt_params = shop_filter()
    product = cursor.execute(
        f'SELECT name, sku, stock_quantity, cost_price FROM products WHERE id = ? {flt_sql}',
        [product_id] + flt_params
    ).fetchone()
    if not product:
        conn.close()
        return jsonify({'success': False, 'error': 'Product not found'}), 404

    # Get stock history
    history = cursor.execute('''
        SELECT id, quantity_change, action_type, note, created_at
        FROM stock_history
        WHERE product_id = ?
        ORDER BY created_at DESC
    ''', (product_id,)).fetchall()

    # Enrich entries with cost info
    result = []
    for row in history:
        entry = dict(row)
        entry['unit_cost'] = None
        if entry['action_type'] == 'purchase' and entry['note']:
            # Extract PO id from note (format: "PO #<id> ...")
            po_match = re.search(r'PO #(\d+)', entry['note'])
            if po_match:
                po_id = int(po_match.group(1))
                po_item = cursor.execute('''
                    SELECT unit_cost FROM po_items
                    WHERE po_id = ? AND product_id = ?
                ''', (po_id, product_id)).fetchone()
                if po_item:
                    entry['unit_cost'] = po_item['unit_cost']
        # For old initial stock entries without cost in note, use product cost_price
        if entry['action_type'] == 'initial' and '| Cost:' not in (entry['note'] or ''):
            entry['unit_cost'] = product['cost_price']
        result.append(entry)

    conn.close()
    return jsonify({
        'product': dict(product),
        'history': result
    })

# API Routes - Sales/POS
@app.route('/api/sales', methods=['POST'])
@login_required
def create_sale():
    data = request.json
    items = data['items']
    total_amount = data['total_amount']
    discount_amount = data.get('discount_amount', 0)
    payment_method = data.get('payment_method', 'cash')
    cash_tendered = data.get('cash_tendered', 0)
    customer_name = data.get('customer_name', '').strip()
    customer_id = data.get('customer_id')
    amount_paid = data.get('amount_paid')  # For partial/credit payments
    due_date = data.get('due_date')  # For credit sales
    current_shop_id = get_current_shop_id()

    # POS sales must always be recorded against a specific shop.
    if not current_shop_id:
        return jsonify({'success': False, 'error': 'Select a specific shop before creating a sale'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # If customer_id provided, validate credit limit for credit/partial sales
        if customer_id and payment_method == 'credit':
            flt_sql, flt_params = shop_filter()
            customer = cursor.execute(
                f'SELECT * FROM customers WHERE id=? {flt_sql}', [customer_id] + flt_params
            ).fetchone()
            if not customer:
                return jsonify({'success': False, 'error': 'Customer not found'}), 404
            if customer['customer_type'] != 'wholesale':
                return jsonify({'success': False, 'error': 'Only wholesale customers can use credit'}), 400
            credit_remaining = customer['credit_limit'] - customer['balance']
            credit_needed = total_amount - (amount_paid or 0)
            if credit_needed > credit_remaining:
                return jsonify({'success': False, 'error': f'Exceeds credit limit. Available credit: {credit_remaining:.2f}'}), 400
            # Use customer name from account
            customer_name = customer['name']

        # Calculate total cost from products in the active shop only.
        total_cost = 0
        for item in items:
            product = cursor.execute(
                'SELECT cost_price FROM products WHERE id=? AND shop_id=? AND is_active = 1',
                (item['product_id'], current_shop_id),
            ).fetchone()
            if not product:
                return jsonify({'success': False, 'error': f'Product ID {item["product_id"]} not found in selected shop'}), 404
            if product:
                total_cost += product['cost_price'] * item['quantity']

        # Create sale record
        cursor.execute('''
            INSERT INTO sales (shop_id, user_id, total_amount, total_cost, discount_amount, items_json, payment_method, cash_tendered, customer_name, customer_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (current_shop_id, session.get('user_id'), total_amount, total_cost, discount_amount, json.dumps(items), payment_method, cash_tendered, customer_name, customer_id))
        sale_id = cursor.lastrowid

        # Validate stock availability before decrementing
        for item in items:
            product_id = item['product_id']
            quantity = item['quantity']
            product = cursor.execute(
                'SELECT stock_quantity, name FROM products WHERE id=? AND shop_id=? AND is_active = 1',
                (product_id, current_shop_id),
            ).fetchone()
            if not product:
                conn.rollback()
                return jsonify({'success': False, 'error': f'Product ID {product_id} not found in selected shop'}), 404
            if product['stock_quantity'] < quantity:
                conn.rollback()
                return jsonify({'success': False, 'error': f'Insufficient stock for {product["name"]}. Available: {product["stock_quantity"]}, Requested: {quantity}'}), 400

        # Update stock for each item
        for item in items:
            product_id = item['product_id']
            quantity = item['quantity']

            cursor.execute('''
                UPDATE products
                SET stock_quantity = stock_quantity - ?
                WHERE id=? AND shop_id=?
            ''', (quantity, product_id, current_shop_id))

            # Log stock history
            cursor.execute('''
                INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                VALUES (?, ?, ?, ?, ?)
            ''', (current_shop_id, product_id, -quantity, 'sale', f'Sale by {session.get("username")}'))

        # Handle credit/partial payment - create invoice and update customer balance
        if customer_id and payment_method in ('credit', 'partial'):
            paid = float(amount_paid or 0)
            outstanding = total_amount - paid

            # Determine invoice status
            if paid >= total_amount:
                inv_status = 'paid'
            elif paid > 0:
                inv_status = 'partial'
            else:
                inv_status = 'unpaid'

            cursor.execute('''
                INSERT INTO invoices (shop_id, sale_id, customer_id, total_amount, paid_amount, status, due_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (current_shop_id, sale_id, customer_id, total_amount, paid, inv_status, due_date))
            invoice_id = cursor.lastrowid

            # Record the initial payment if any
            if paid > 0:
                cursor.execute('''
                    INSERT INTO payments (shop_id, customer_id, invoice_id, amount, payment_method, received_by, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (current_shop_id, customer_id, invoice_id, paid, 'cash', session.get('user_id'), f'Initial payment for sale #{sale_id}'))

            # Update customer balance (what they owe)
            cursor.execute('''
                UPDATE customers SET balance = balance + ? WHERE id = ?
            ''', (outstanding, customer_id))

        # Available funds include only cash actually received now.
        paid_now = 0.0
        if payment_method in ('partial', 'credit'):
            paid_now = _as_float(amount_paid)
        else:
            paid_now = _as_float(total_amount)

        _record_finance_transaction(
            cursor,
            shop_id=current_shop_id,
            direction='IN',
            amount=paid_now,
            transaction_type='sale_payment',
            source_table='sales',
            source_id=sale_id,
            reference=f'SALE-{sale_id}',
            notes=f'Sale payment via {payment_method}',
            created_by=session.get('user_id'),
        )

        conn.commit()
        return jsonify({'success': True, 'sale_id': sale_id}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/sales', methods=['GET'])
@login_required
def get_sales():
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    user_id = request.args.get('user_id')

    conn = get_db_connection()

    # Build query with filters
    flt_sql, flt_params = shop_filter('s')
    query = f'''
        SELECT s.*, u.username, u.full_name,
               i.id AS invoice_id, i.status AS invoice_status,
               i.paid_amount AS invoice_paid_amount, i.total_amount AS invoice_total_amount
        FROM sales s
        LEFT JOIN users u ON s.user_id = u.id
        LEFT JOIN invoices i ON i.sale_id = s.id
        WHERE 1=1 {flt_sql}
    '''
    params = list(flt_params)

    # Date range filter
    if start_date and end_date:
        query += ' AND DATE(s.sale_date) BETWEEN ? AND ?'
        params.extend([start_date, end_date])
    elif start_date:
        query += ' AND DATE(s.sale_date) >= ?'
        params.append(start_date)
    elif end_date:
        query += ' AND DATE(s.sale_date) <= ?'
        params.append(end_date)

    # User filter
    if user_id and user_id != 'all':
        query += ' AND s.user_id = ?'
        params.append(int(user_id))

    query += ' ORDER BY sale_date DESC LIMIT 500'

    sales = conn.execute(query, params).fetchall()

    # Fetch all refund data for these sales
    sale_ids = [s['id'] for s in sales]
    refund_map = {}
    if sale_ids:
        placeholders = ','.join('?' * len(sale_ids))
        refunds = conn.execute(f'''
            SELECT sale_id, SUM(refund_amount) as total_refunded,
                   COUNT(*) as refund_count, GROUP_CONCAT(items_json, '|||') as all_items
            FROM sale_returns WHERE sale_id IN ({placeholders})
            GROUP BY sale_id
        ''', sale_ids).fetchall()
        for r in refunds:
            refund_map[r['sale_id']] = {
                'total_refunded': r['total_refunded'],
                'refund_count': r['refund_count'],
                'all_items': r['all_items']
            }

    conn.close()

    result = []
    for sale in sales:
        sale_dict = dict(sale)
        sale_dict['items'] = json.loads(sale_dict['items_json'])

        # Add refund info
        ref = refund_map.get(sale_dict['id'])
        if ref:
            sale_dict['refund_total'] = ref['total_refunded']
            sale_dict['refund_count'] = ref['refund_count']
            # Determine if fully or partially refunded
            sale_dict['refund_status'] = 'refunded' if ref['total_refunded'] >= sale_dict['total_amount'] else 'partial'
        else:
            sale_dict['refund_total'] = 0
            sale_dict['refund_count'] = 0
            sale_dict['refund_status'] = 'none'

        # Only show profit to admins
        if session.get('role') in ('admin', 'super_admin', 'shop_owner'):
            sale_dict['profit'] = sale_dict['total_amount'] - (sale_dict['total_cost'] or 0)
        else:
            # Hide cost and profit from regular users
            sale_dict.pop('total_cost', None)

        result.append(sale_dict)

    return jsonify(result)

@app.route('/api/sales/users', methods=['GET'])
@login_required
def get_sales_users():
    """Get list of users who have made sales (for filter dropdown)"""
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('s')
    users = conn.execute(f'''
        SELECT DISTINCT u.id, u.username, u.full_name
        FROM users u
        INNER JOIN sales s ON u.id = s.user_id
        WHERE 1=1 {flt_sql}
        ORDER BY u.username
    ''', flt_params).fetchall()
    conn.close()
    return jsonify([dict(row) for row in users])

@app.route('/api/sales/<int:sale_id>', methods=['GET'])
@login_required
def get_sale(sale_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('s')
    sale = conn.execute(
        f'SELECT s.*, u.username FROM sales s LEFT JOIN users u ON s.user_id = u.id WHERE s.id = ? {flt_sql}',
        [sale_id] + flt_params
    ).fetchone()
    if not sale:
        conn.close()
        return jsonify({'error': 'Sale not found'}), 404
    sale_dict = dict(sale)
    sale_dict['items'] = json.loads(sale_dict['items_json'])

    # Subtract already-refunded quantities
    refunds = conn.execute(
        'SELECT items_json, refund_amount FROM sale_returns WHERE sale_id = ?', (sale_id,)
    ).fetchall()
    conn.close()

    refunded_qty = {}
    refund_total = 0.0
    for refund in refunds:
        refund_total += float(refund.get('refund_amount') or 0)
        for ri in json.loads(refund['items_json']):
            pid = ri['product_id']
            refunded_qty[pid] = refunded_qty.get(pid, 0) + ri['quantity']

    for item in sale_dict['items']:
        original = item['quantity']
        already = refunded_qty.get(item['product_id'], 0)
        item['original_quantity'] = original
        item['refunded_quantity'] = already
        item['quantity'] = max(0, original - already)

    sale_dict['refund_total'] = round(refund_total, 2)
    sale_dict['remaining_refundable_amount'] = round(max(0, float(sale_dict.get('total_amount') or 0) - refund_total), 2)

    return jsonify(sale_dict)

@app.route('/api/refunds', methods=['POST'])
@login_required
def create_refund():
    data = request.json
    sale_id = data.get('sale_id')
    items = data.get('items', [])
    reason = data.get('reason', '')

    if not sale_id or not items:
        return jsonify({'error': 'sale_id and items required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        flt_sql, flt_params = shop_filter()
        sale = cursor.execute(
            f'SELECT id, total_amount, discount_amount, items_json FROM sales WHERE id=? {flt_sql}',
            [sale_id] + flt_params
        ).fetchone()
        if not sale:
            return jsonify({'error': 'Sale not found'}), 404

        sale_items = json.loads(sale['items_json'])
        sale_items_by_id = {}
        gross_sale_total = 0.0
        for si in sale_items:
            pid = int(si['product_id'])
            qty = int(si.get('quantity', 0) or 0)
            price = float(si.get('price', 0) or 0)
            sale_items_by_id[pid] = {
                'quantity': qty,
                'price': price,
                'name': si.get('name', f'Product #{pid}')
            }
            gross_sale_total += price * qty

        refunded_rows = cursor.execute(
            'SELECT items_json, refund_amount FROM sale_returns WHERE sale_id = ?',
            (sale_id,)
        ).fetchall()

        already_refunded_qty = {}
        total_refunded_so_far = 0.0
        for rr in refunded_rows:
            total_refunded_so_far += float(rr.get('refund_amount') or 0)
            for ri in json.loads(rr['items_json']):
                pid = int(ri['product_id'])
                already_refunded_qty[pid] = already_refunded_qty.get(pid, 0) + int(ri.get('quantity', 0) or 0)

        normalized_items = []
        selected_gross_total = 0.0
        for item in items:
            product_id = int(item['product_id'])
            quantity = int(item.get('quantity', 0) or 0)
            if quantity <= 0:
                continue

            if product_id not in sale_items_by_id:
                return jsonify({'error': f'Product #{product_id} is not in sale #{sale_id}'}), 400

            sold_qty = sale_items_by_id[product_id]['quantity']
            already_qty = already_refunded_qty.get(product_id, 0)
            remaining_qty = sold_qty - already_qty
            if quantity > remaining_qty:
                return jsonify({'error': f'Refund quantity for {sale_items_by_id[product_id]["name"]} exceeds remaining refundable quantity ({remaining_qty})'}), 400

            line_price = float(sale_items_by_id[product_id]['price'])
            selected_gross_total += line_price * quantity

            normalized_items.append({
                'product_id': product_id,
                'name': item.get('name') or sale_items_by_id[product_id]['name'],
                'price': line_price,
                'quantity': quantity
            })

        if not normalized_items:
            return jsonify({'error': 'No refundable items selected'}), 400

        sale_total_after_discount = float(sale['total_amount'] or 0)
        remaining_refundable_amount = max(0.0, sale_total_after_discount - total_refunded_so_far)
        if remaining_refundable_amount <= 0:
            return jsonify({'error': 'This sale is already fully refunded'}), 400

        if gross_sale_total > 0:
            discount_factor = sale_total_after_discount / gross_sale_total
        else:
            discount_factor = 1.0
        discount_factor = max(0.0, min(1.0, discount_factor))

        computed_refund_amount = round(selected_gross_total * discount_factor, 2)
        refund_amount = min(computed_refund_amount, round(remaining_refundable_amount, 2))
        if refund_amount <= 0:
            return jsonify({'error': 'Refund amount is zero after discount allocation'}), 400

        for item in normalized_items:
            product_id = item['product_id']
            quantity = item['quantity']
            cursor.execute(
                f'UPDATE products SET stock_quantity = stock_quantity + ? WHERE id=? {flt_sql}',
                [quantity, product_id] + flt_params
            )
            cursor.execute('''
                INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                VALUES (?, ?, ?, ?, ?)
            ''', (get_current_shop_id(), product_id, quantity, 'return', f'Refund for sale #{sale_id} by {session.get("username")}'))

        cursor.execute('''
            INSERT INTO sale_returns (shop_id, sale_id, user_id, items_json, refund_amount, reason)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (get_current_shop_id(), sale_id, session.get('user_id'), json.dumps(normalized_items), refund_amount, reason))
        refund_id = cursor.lastrowid

        _record_finance_transaction(
            cursor,
            shop_id=get_current_shop_id(),
            direction='OUT',
            amount=refund_amount,
            transaction_type='refund',
            source_table='sale_returns',
            source_id=refund_id,
            reference=f'REF-{refund_id}',
            notes=reason or f'Refund for sale #{sale_id}',
            created_by=session.get('user_id'),
        )

        conn.commit()
        return jsonify({'success': True, 'refund_id': refund_id, 'refund_amount': refund_amount}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Sales: Export CSV ---
@app.route('/api/sales/export-csv', methods=['GET'])
@login_required
def export_sales_csv():
    import csv as csv_mod, io as io_mod
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('s')
    query = f'''SELECT s.*, u.username, u.full_name,
               i.status AS invoice_status, i.paid_amount AS invoice_paid_amount
        FROM sales s
        LEFT JOIN users u ON s.user_id = u.id
        LEFT JOIN invoices i ON i.sale_id = s.id
        WHERE 1=1 {flt_sql}'''
    params = list(flt_params)
    if start_date:
        query += ' AND DATE(s.sale_date) >= ?'; params.append(start_date)
    if end_date:
        query += ' AND DATE(s.sale_date) <= ?'; params.append(end_date)
    query += ' ORDER BY sale_date DESC'
    sales_rows = conn.execute(query, params).fetchall()

    # Build refund map for status computation
    sale_ids = [s['id'] for s in sales_rows]
    refund_map = {}
    if sale_ids:
        placeholders = ','.join('?' * len(sale_ids))
        refunds = conn.execute(f'''
            SELECT sale_id, SUM(refund_amount) as total_refunded
            FROM sale_returns WHERE sale_id IN ({placeholders})
            GROUP BY sale_id
        ''', sale_ids).fetchall()
        for r in refunds:
            refund_map[r['sale_id']] = r['total_refunded']

    conn.close()

    def compute_status(s):
        total_refunded = refund_map.get(s['id'], 0)
        if total_refunded >= s['total_amount']:
            return 'Refunded'
        elif total_refunded > 0:
            return 'Partial Refund'
        pm = (s['payment_method'] or '')
        if pm in ('credit', 'partial'):
            inv_status = s['invoice_status']
            if inv_status == 'paid':
                return 'Completed'
            elif inv_status == 'partial':
                return 'Partially Paid'
            return 'Unpaid'
        return 'Completed'

    is_admin = session.get('role') in ('admin', 'super_admin', 'shop_owner')
    output = io_mod.StringIO()
    writer = csv_mod.writer(output)

    if is_admin:
        writer.writerow(['Sale ID', 'Date', 'Sold By', 'Customer', 'Payment Method', 'Status', 'Items', 'Subtotal', 'Discount', 'Total', 'Cost', 'Profit'])
    else:
        writer.writerow(['Sale ID', 'Date', 'Sold By', 'Customer', 'Payment Method', 'Status', 'Items', 'Subtotal', 'Discount', 'Total'])

    for s in sales_rows:
        items = json.loads(s['items_json'])
        grouped_items = {}
        for item in items:
            name = item.get('name', 'Item')
            variant_parts = []
            if item.get('size'):
                variant_parts.append(f"Size: {item['size']}")
            if item.get('color'):
                variant_parts.append(f"Color: {item['color']}")
            variant_label = ', '.join(variant_parts)

            if name not in grouped_items:
                grouped_items[name] = []

            grouped_items[name].append({
                'variant': variant_label,
                'quantity': item.get('quantity', 0)
            })

        item_names = ', '.join(
            f"{name} (" + '; '.join(
                f"{entry['variant']} x{entry['quantity']}" if entry['variant'] else f"x{entry['quantity']}"
                for entry in entries
            ) + ")"
            for name, entries in grouped_items.items()
        )
        discount = s['discount_amount'] or 0
        subtotal = s['total_amount'] + discount
        row = [s['id'], s['sale_date'], s['full_name'] or s['username'] or '',
               s['customer_name'] or '', s['payment_method'] or 'cash',
               compute_status(s),
               item_names, f"{subtotal:.2f}", f"{discount:.2f}", f"{s['total_amount']:.2f}"]
        if is_admin:
            row.extend([f"{s['total_cost']:.2f}", f"{s['total_amount'] - s['total_cost']:.2f}"])
        writer.writerow(row)

    csv_bytes = output.getvalue().encode('utf-8')
    return send_file(BytesIO(csv_bytes), mimetype='text/csv', as_attachment=True,
                     download_name=f'sales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')

# --- Sales: Refund details for a sale ---
@app.route('/api/sales/<int:sale_id>/refunds', methods=['GET'])
@login_required
def get_sale_refunds(sale_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('sr')
    refunds = conn.execute(f'''
        SELECT sr.*, u.username, u.full_name
        FROM sale_returns sr
        LEFT JOIN users u ON sr.user_id = u.id
        WHERE sr.sale_id = ? {flt_sql}
        ORDER BY sr.return_date DESC
    ''', [sale_id] + flt_params).fetchall()
    conn.close()

    result = []
    for r in refunds:
        rd = dict(r)
        rd['items'] = json.loads(rd['items_json'])
        result.append(rd)
    return jsonify(result)

# API Routes - User Management
@app.route('/api/users', methods=['GET'])
@login_required
@admin_required
def get_users():
    conn = get_db_connection()
    role = session.get('role')
    if role == 'super_admin':
        active_shop_id = session.get('active_shop_id')
        if active_shop_id:
            users = conn.execute(
                'SELECT u.id, u.username, u.role, u.full_name, u.created_at, u.last_login, u.last_activity, u.shop_id, s.name as shop_name '
                'FROM users u LEFT JOIN shops s ON u.shop_id = s.id WHERE u.shop_id = ? ORDER BY u.created_at DESC',
                (active_shop_id,)
            ).fetchall()
        else:
            users = conn.execute(
                'SELECT u.id, u.username, u.role, u.full_name, u.created_at, u.last_login, u.last_activity, u.shop_id, s.name as shop_name '
                'FROM users u LEFT JOIN shops s ON u.shop_id = s.id ORDER BY u.created_at DESC'
            ).fetchall()
    elif role == 'shop_owner':
        active_shop_id = session.get('active_shop_id')
        allowed_ids = _get_user_shop_ids()
        if active_shop_id and active_shop_id in allowed_ids:
            users = conn.execute(
                'SELECT u.id, u.username, u.role, u.full_name, u.created_at, u.last_login, u.last_activity, u.shop_id, s.name as shop_name '
                'FROM users u LEFT JOIN shops s ON u.shop_id = s.id WHERE u.shop_id = ? ORDER BY u.created_at DESC',
                (active_shop_id,)
            ).fetchall()
        elif allowed_ids:
            placeholders = ','.join(['?'] * len(allowed_ids))
            users = conn.execute(
                f'SELECT u.id, u.username, u.role, u.full_name, u.created_at, u.last_login, u.last_activity, u.shop_id, s.name as shop_name '
                f'FROM users u LEFT JOIN shops s ON u.shop_id = s.id WHERE u.shop_id IN ({placeholders}) ORDER BY u.created_at DESC',
                allowed_ids
            ).fetchall()
        else:
            users = []
    else:
        shop_id = session.get('shop_id')
        users = conn.execute(
            'SELECT u.id, u.username, u.role, u.full_name, u.created_at, u.last_login, u.last_activity, u.shop_id, s.name as shop_name '
            'FROM users u LEFT JOIN shops s ON u.shop_id = s.id WHERE u.shop_id = ? ORDER BY u.created_at DESC',
            (shop_id,)
        ).fetchall()
    conn.close()
    result = [dict(row) for row in users]

    # For super_admin, enrich shop_owner users with their assigned shop names
    if role == 'super_admin':
        owner_ids = [u['id'] for u in result if u['role'] == 'shop_owner']
        if owner_ids:
            conn2 = get_db_connection()
            placeholders = ','.join(['?'] * len(owner_ids))
            rows = conn2.execute(
                f'SELECT us.user_id, s.name, s.icon FROM user_shops us '
                f'JOIN shops s ON us.shop_id = s.id WHERE us.user_id IN ({placeholders})',
                owner_ids
            ).fetchall()
            conn2.close()
            shop_map = {}
            for r in rows:
                shop_map.setdefault(r['user_id'], []).append(r['name'])
            for u in result:
                if u['role'] == 'shop_owner':
                    u['assigned_shops'] = shop_map.get(u['id'], [])

    return jsonify(result)


@app.route('/api/users/access-stats', methods=['GET'])
@login_required
@super_admin_required
def get_user_access_stats():
    try:
        days = int(request.args.get('days', 30))
    except Exception:
        days = 30
    if days < 1:
        days = 1
    if days > 365:
        days = 365

    cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()

    users = conn.execute(
        '''
        SELECT
            u.id,
            u.username,
            u.full_name,
            u.role,
            u.shop_id,
            s.name AS shop_name,
            COALESCE(u.portal_access_count, 0) AS portal_access_count,
            u.last_login,
            u.last_portal_access_at,
            u.last_access_ip,
            COALESCE(SUM(CASE WHEN l.access_at >= ? THEN 1 ELSE 0 END), 0) AS recent_access_count
        FROM users u
        LEFT JOIN shops s ON s.id = u.shop_id
        LEFT JOIN user_access_log l ON l.user_id = u.id
        GROUP BY
            u.id,
            u.username,
            u.full_name,
            u.role,
            u.shop_id,
            s.name,
            u.portal_access_count,
            u.last_login,
            u.last_portal_access_at,
            u.last_access_ip
        ORDER BY portal_access_count DESC, u.last_login DESC
        ''',
        (cutoff,)
    ).fetchall()

    summary = conn.execute(
        '''
        SELECT
            COUNT(*) AS total_access_events,
            COUNT(DISTINCT user_id) AS active_users,
            SUM(CASE WHEN access_at >= DATE_SUB(NOW(), INTERVAL 1 DAY) THEN 1 ELSE 0 END) AS last_24h_events
        FROM user_access_log
        WHERE access_at >= ?
        ''',
        (cutoff,)
    ).fetchone()

    recent_events = conn.execute(
        '''
        SELECT
            l.user_id,
            u.username,
            u.full_name,
            u.role,
            s.name AS shop_name,
            l.access_ip,
            l.access_at
        FROM user_access_log l
        JOIN users u ON u.id = l.user_id
        LEFT JOIN shops s ON s.id = u.shop_id
        ORDER BY l.access_at DESC
        LIMIT 100
        '''
    ).fetchall()

    conn.close()

    return jsonify({
        'days': days,
        'summary': dict(summary) if summary else {
            'total_access_events': 0,
            'active_users': 0,
            'last_24h_events': 0,
        },
        'users': [dict(row) for row in users],
        'recent_events': [dict(row) for row in recent_events],
    })

@app.route('/api/users', methods=['POST'])
@login_required
@admin_required
def create_user():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    role = data.get('role', '').strip()
    full_name = data.get('full_name', '').strip() or None

    if not username or not password:
        return jsonify({'success': False, 'error': 'Username and password are required'}), 400

    # Validate allowed roles based on who's creating
    requester_role = session.get('role')
    allowed_roles = ['admin', 'user']
    if requester_role == 'super_admin':
        allowed_roles.append('shop_owner')
    if role not in allowed_roles:
        return jsonify({'success': False, 'error': 'Invalid role'}), 400

    # Determine which shop this user belongs to
    new_user_shop_id = data.get('shop_id')
    if requester_role == 'super_admin':
        # super_admin can assign any shop; default to their active shop
        if not new_user_shop_id:
            new_user_shop_id = session.get('active_shop_id') or session.get('shop_id')
    elif requester_role == 'shop_owner':
        # shop_owner can only create users for their assigned shops
        allowed_ids = _get_user_shop_ids()
        if new_user_shop_id:
            new_user_shop_id = int(new_user_shop_id)
            if new_user_shop_id not in allowed_ids:
                return jsonify({'success': False, 'error': 'Cannot assign user to this shop'}), 403
        else:
            new_user_shop_id = session.get('active_shop_id') or (allowed_ids[0] if allowed_ids else None)
    else:
        # admin can only create users for their own shop
        new_user_shop_id = session.get('shop_id')

    # For shop_owner, collect shop_ids for multi-shop assignment
    shop_ids = data.get('shop_ids', [])
    if role == 'shop_owner' and requester_role == 'super_admin' and not shop_ids:
        return jsonify({'success': False, 'error': 'At least one shop must be assigned to a Shop Owner'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute('''
            INSERT INTO users (username, password, role, full_name, shop_id)
            VALUES (?, ?, ?, ?, ?)
        ''', (username, hash_password(password), role, full_name, new_user_shop_id))

        new_user_id = cursor.lastrowid

        # If shop_owner, insert into user_shops
        if role == 'shop_owner' and shop_ids:
            for sid in shop_ids:
                cursor.execute('INSERT INTO user_shops (user_id, shop_id) VALUES (?, ?)', (new_user_id, int(sid)))

        conn.commit()
        return jsonify({'success': True, 'id': new_user_id}), 201
    except Exception as e:
        conn.rollback()
        if 'UNIQUE constraint failed' in str(e) or 'Duplicate entry' in str(e):
            return jsonify({'success': False, 'error': 'Username already exists'}), 400
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_user(user_id):
    # Prevent deleting yourself
    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'error': 'Cannot delete your own account'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM users WHERE id=?', (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/users/<int:user_id>/password', methods=['PUT'])
@login_required
@admin_required
def change_user_password(user_id):
    data = request.json
    new_password = data.get('password', '').strip()

    if not new_password:
        return jsonify({'success': False, 'error': 'Password cannot be empty'}), 400

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute('''
            UPDATE users
            SET password = ?
            WHERE id = ?
        ''', (hash_password(new_password), user_id))

        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({'success': False, 'error': 'User not found'}), 404

        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# ── Shop assignment for shop_owner users (super_admin only) ───────────────────

@app.route('/api/users/<int:user_id>/shops', methods=['GET'])
@login_required
@super_admin_required
def get_user_shops(user_id):
    """Get shops assigned to a user (for shop_owner role)."""
    conn = get_db_connection()
    shops = conn.execute(
        'SELECT s.id, s.name, s.icon FROM shops s '
        'INNER JOIN user_shops us ON us.shop_id = s.id WHERE us.user_id = ? ORDER BY s.name',
        (user_id,)
    ).fetchall()
    conn.close()
    return jsonify([dict(s) for s in shops])

@app.route('/api/users/<int:user_id>/shops', methods=['PUT'])
@login_required
@super_admin_required
def set_user_shops(user_id):
    """Set shops assigned to a shop_owner user. Replaces existing assignments."""
    data = request.json
    shop_ids = data.get('shop_ids', [])

    conn = get_db_connection()
    cursor = conn.cursor()

    # Verify user exists and is shop_owner
    user = cursor.execute('SELECT role FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': 'User not found'}), 404
    if user['role'] != 'shop_owner':
        conn.close()
        return jsonify({'success': False, 'error': 'Shop assignment is only for shop_owner role'}), 400

    try:
        # Remove existing assignments
        cursor.execute('DELETE FROM user_shops WHERE user_id = ?', (user_id,))
        # Insert new assignments
        for sid in shop_ids:
            cursor.execute(
                'INSERT INTO user_shops (user_id, shop_id) VALUES (?, ?)',
                (user_id, int(sid))
            )
        conn.commit()
        return jsonify({'success': True, 'assigned': len(shop_ids)})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/users/<int:user_id>/shop', methods=['PUT'])
@login_required
@super_admin_required
def reassign_user_shop(user_id):
    """Reassign an admin or sales staff user to a different shop."""
    data = request.json
    shop_id = data.get('shop_id')
    if not shop_id:
        return jsonify({'success': False, 'error': 'shop_id is required'}), 400

    conn = get_db_connection()
    user = conn.execute('SELECT role FROM users WHERE id = ?', (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'success': False, 'error': 'User not found'}), 404
    if user['role'] not in ('admin', 'user'):
        conn.close()
        return jsonify({'success': False, 'error': 'Only admin and sales staff can be reassigned'}), 400

    try:
        conn.execute('UPDATE users SET shop_id = ? WHERE id = ?', (int(shop_id), user_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/reports/sales', methods=['GET'])
@login_required
def get_sales_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    is_admin = session.get('role') in ('admin', 'super_admin', 'shop_owner')
    flt_sql, flt_params = shop_filter('s')
    flt_sql_plain, flt_params_plain = shop_filter()
    flt_sql_sr, flt_params_sr = shop_filter('sr')

    try:
        # Add time to include the entire end date
        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # Get summary statistics
        summary_query = f'''
            SELECT
                COUNT(*) as transaction_count,
                SUM(total_amount) as total_sales,
                AVG(total_amount) as avg_transaction
        '''
        if is_admin:
            summary_query += ', SUM(total_amount - COALESCE(total_cost, 0)) as total_profit'

        summary_query += f'''
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql}
        '''

        summary = cursor.execute(summary_query, [start_date_time, end_date_time] + flt_params).fetchone()

        # Get daily sales breakdown
        daily_query = f'''
            SELECT
                DATE(sale_date) as sale_date,
                COUNT(*) as transaction_count,
                SUM(total_amount) as daily_sales
        '''
        if is_admin:
            daily_query += ', SUM(total_amount - COALESCE(total_cost, 0)) as daily_profit'

        daily_query += f'''
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql}
            GROUP BY DATE(sale_date)
            ORDER BY sale_date ASC
        '''

        daily_sales = cursor.execute(daily_query, [start_date_time, end_date_time] + flt_params).fetchall()

        # Get all sales in the date range to parse items
        sales_data = cursor.execute(f'''
            SELECT id, items_json, total_amount, total_cost
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql}
        ''', [start_date_time, end_date_time] + flt_params).fetchall()

        # Parse items and aggregate by product
        product_stats = {}
        for sale in sales_data:
            items = json.loads(sale['items_json'])
            for item in items:
                product_id = item['product_id']
                if product_id not in product_stats:
                    product_stats[product_id] = {
                        'product_id': product_id,
                        'name': item['name'],
                        'sku': item.get('sku', 'N/A'),
                        'quantity': 0,
                        'revenue': 0,
                        'cost': 0
                    }

                product_stats[product_id]['quantity'] += item['quantity']
                product_stats[product_id]['revenue'] += item['quantity'] * item['price']

                # Calculate cost if admin
                if is_admin and 'cost_price' in item:
                    product_stats[product_id]['cost'] += item['quantity'] * item['cost_price']

        # Convert to list and sort
        all_products = list(product_stats.values())

        # Top selling products (by quantity)
        top_products = sorted(all_products, key=lambda x: x['quantity'], reverse=True)[:10]

        # Format top products
        top_products_formatted = []
        for product in top_products:
            formatted = {
                'name': product['name'],
                'sku': product['sku'],
                'total_quantity': product['quantity'],
                'total_revenue': product['revenue']
            }
            if is_admin:
                formatted['total_profit'] = product['revenue'] - product['cost']
            top_products_formatted.append(formatted)

        # Product performance (by revenue)
        product_performance = sorted(all_products, key=lambda x: x['revenue'], reverse=True)

        # Format product performance
        performance_formatted = []
        for product in product_performance:
            formatted = {
                'name': product['name'],
                'sku': product['sku'],
                'quantity_sold': product['quantity'],
                'revenue': product['revenue']
            }
            if is_admin:
                profit = product['revenue'] - product['cost']
                formatted['profit'] = profit
                formatted['profit_margin'] = (profit / product['revenue'] * 100) if product['revenue'] > 0 else 0
            performance_formatted.append(formatted)

        # --- Payment method breakdown ---
        payment_breakdown = cursor.execute(f'''
            SELECT COALESCE(payment_method, 'cash') as method,
                   COUNT(*) as count, SUM(total_amount) as total
            FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql}
            GROUP BY COALESCE(payment_method, 'cash')
        ''', [start_date_time, end_date_time] + flt_params).fetchall()

        # --- Discount summary ---
        discount_summary = cursor.execute(f'''
            SELECT COUNT(CASE WHEN discount_amount > 0 THEN 1 END) as discounted_count,
                   SUM(COALESCE(discount_amount, 0)) as total_discount,
                   AVG(CASE WHEN discount_amount > 0 THEN discount_amount END) as avg_discount
            FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql}
        ''', [start_date_time, end_date_time] + flt_params).fetchone()

        # --- Refund summary ---
        refund_summary = cursor.execute(f'''
            SELECT COUNT(*) as refund_count,
                   SUM(refund_amount) as total_refunded,
                   COUNT(DISTINCT sale_id) as sales_with_refunds
            FROM sale_returns WHERE return_date BETWEEN ? AND ? {flt_sql_plain}
        ''', [start_date_time, end_date_time] + flt_params_plain).fetchone()

        refund_profit_impact = cursor.execute(f'''
            SELECT
                COALESCE(SUM(sr.refund_amount), 0) as total_refunded,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(s.total_amount, 0) > 0
                            THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                        ELSE 0
                    END
                ), 0) as refunded_cogs,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(s.total_amount, 0) > 0
                            THEN LEAST(sr.refund_amount, s.total_amount) / s.total_amount * COALESCE(s.discount_amount, 0)
                        ELSE 0
                    END
                ), 0) as refunded_discount
            FROM sale_returns sr
            JOIN sales s ON s.id = sr.sale_id
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
        ''', [start_date_time, end_date_time] + flt_params_sr).fetchone()

        daily_refund_impact = cursor.execute(f'''
            SELECT
                DATE(sr.return_date) as return_date,
                COALESCE(SUM(sr.refund_amount), 0) as refunded,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(s.total_amount, 0) > 0
                            THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                        ELSE 0
                    END
                ), 0) as refunded_cogs
            FROM sale_returns sr
            JOIN sales s ON s.id = sr.sale_id
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
            GROUP BY DATE(sr.return_date)
        ''', [start_date_time, end_date_time] + flt_params_sr).fetchall()

        # Top refunded products
        top_refunded_query = cursor.execute(f'''
            SELECT sr.items_json, sr.refund_amount
            FROM sale_returns sr WHERE sr.return_date BETWEEN ? AND ? {flt_sql_plain}
        ''', [start_date_time, end_date_time] + flt_params_plain).fetchall()

        refund_product_stats = {}
        for r in top_refunded_query:
            items = json.loads(r['items_json'])
            for item in items:
                pid = item.get('product_id', item.get('name', 'Unknown'))
                name = item.get('name', f'Product #{pid}')
                if name not in refund_product_stats:
                    refund_product_stats[name] = 0
                refund_product_stats[name] += item.get('quantity', 0)
        top_refunded = sorted(refund_product_stats.items(), key=lambda x: x[1], reverse=True)[:5]

        # --- Staff performance ---
        staff_query = f'''
            SELECT u.id, u.username, u.full_name,
                   COUNT(s.id) as transaction_count,
                   SUM(s.total_amount) as total_sales,
                   AVG(s.total_amount) as avg_sale
        '''
        if is_admin:
            staff_query += ', SUM(s.total_amount - COALESCE(s.total_cost, 0)) as total_profit'
        staff_query += f'''
            FROM sales s JOIN users u ON s.user_id = u.id
            WHERE s.sale_date BETWEEN ? AND ? {flt_sql}
            GROUP BY u.id ORDER BY total_sales DESC
        '''
        staff_performance = cursor.execute(staff_query, [start_date_time, end_date_time] + flt_params).fetchall()

        # --- Customer insights ---
        customer_query = cursor.execute(f'''
            SELECT COALESCE(customer_name, '') as customer_name,
                   COUNT(*) as visit_count,
                   SUM(total_amount) as total_spent,
                   AVG(total_amount) as avg_spent
            FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql}
            AND customer_name IS NOT NULL AND customer_name != ''
            GROUP BY customer_name ORDER BY total_spent DESC LIMIT 10
        ''', [start_date_time, end_date_time] + flt_params).fetchall()

        total_with_customer = cursor.execute(f'''
            SELECT COUNT(*) as c FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql}
            AND customer_name IS NOT NULL AND customer_name != ''
        ''', [start_date_time, end_date_time] + flt_params).fetchone()

        repeat_customers = cursor.execute(f'''
            SELECT COUNT(*) as c FROM (
                SELECT customer_name FROM sales s
                WHERE sale_date BETWEEN ? AND ? {flt_sql}
                AND customer_name IS NOT NULL AND customer_name != ''
                GROUP BY customer_name HAVING COUNT(*) > 1
            ) AS subq
        ''', [start_date_time, end_date_time] + flt_params).fetchone()

        # --- Category breakdown ---
        # Build product category lookup for fallback when historical items_json is missing category.
        flt_sql_p, flt_params_p = shop_filter('p')
        product_rows = cursor.execute(f'''
            SELECT p.id, p.sku, p.category
            FROM products p
            WHERE 1=1 {flt_sql_p}
        ''', flt_params_p).fetchall()

        product_category_by_id = {}
        product_category_by_sku = {}
        for row in product_rows:
            cat_val = normalize_category_label(row['category'])
            if cat_val:
                product_category_by_id[row['id']] = cat_val
                sku_val = (row['sku'] or '').strip().lower()
                if sku_val:
                    product_category_by_sku[sku_val] = cat_val

        category_stats = {}
        generic_categories = {'', 'other', 'uncategorized', 'n/a', 'na', 'none', '-', 'unknown'}
        for sale in sales_data:
            items = json.loads(sale['items_json'])
            for item in items:
                cat = normalize_category_label(item.get('category'))

                # If category is generic/missing, resolve from product_id first, then sku.
                if not cat or cat.lower() in generic_categories:
                    pid = item.get('product_id')
                    resolved = None

                    if pid is not None:
                        resolved = product_category_by_id.get(pid)
                        if resolved is None:
                            try:
                                resolved = product_category_by_id.get(int(pid))
                            except (TypeError, ValueError):
                                resolved = None

                    if resolved is None:
                        sku_key = (item.get('sku') or '').strip().lower()
                        if sku_key:
                            resolved = product_category_by_sku.get(sku_key)

                    cat = normalize_category_label(resolved or 'Other')

                if cat not in category_stats:
                    category_stats[cat] = {'revenue': 0, 'quantity': 0, 'cost': 0}
                category_stats[cat]['revenue'] += item['quantity'] * item['price']
                category_stats[cat]['quantity'] += item['quantity']
                if is_admin and 'cost_price' in item:
                    category_stats[cat]['cost'] += item['quantity'] * item['cost_price']

        category_breakdown = []
        for cat, stats in sorted(category_stats.items(), key=lambda x: x[1]['revenue'], reverse=True):
            entry = {'category': cat, 'revenue': stats['revenue'], 'quantity': stats['quantity']}
            if is_admin:
                entry['profit'] = stats['revenue'] - stats['cost']
            category_breakdown.append(entry)

        # --- Hourly distribution ---
        hourly_data = cursor.execute(f'''
            SELECT HOUR(sale_date) as hour,
                   COUNT(*) as count, SUM(total_amount) as total
            FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql}
            GROUP BY HOUR(sale_date) ORDER BY hour
        ''', [start_date_time, end_date_time] + flt_params).fetchall()

        # --- Period comparison ---
        # Calculate previous period of same length
        from datetime import datetime, timedelta
        sd = datetime.strptime(start_date, '%Y-%m-%d')
        ed = datetime.strptime(end_date, '%Y-%m-%d')
        period_days = (ed - sd).days + 1
        prev_end = sd - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_days - 1)
        prev_start_str = prev_start.strftime('%Y-%m-%d') + ' 00:00:00'
        prev_end_str = prev_end.strftime('%Y-%m-%d') + ' 23:59:59'

        prev_summary_query = f'''
            SELECT COUNT(*) as transaction_count,
                   COALESCE(SUM(total_amount), 0) as total_sales,
                   COALESCE(AVG(total_amount), 0) as avg_transaction
        '''
        if is_admin:
            prev_summary_query += ', COALESCE(SUM(total_amount - COALESCE(total_cost, 0)), 0) as total_profit'
        prev_summary_query += f' FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql}'
        prev_summary = cursor.execute(prev_summary_query, [prev_start_str, prev_end_str] + flt_params).fetchone()

        prev_refund_impact = cursor.execute(f'''
            SELECT
                COALESCE(SUM(sr.refund_amount), 0) as total_refunded,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(s.total_amount, 0) > 0
                            THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                        ELSE 0
                    END
                ), 0) as refunded_cogs
            FROM sale_returns sr
            JOIN sales s ON s.id = sr.sale_id
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
        ''', [prev_start_str, prev_end_str] + flt_params_sr).fetchone()

        gross_total_sales = float(summary['total_sales'] or 0)
        current_refunded_total = float(refund_summary['total_refunded'] or 0)
        net_total_sales = gross_total_sales - current_refunded_total
        prev_gross_total_sales = float(prev_summary['total_sales'] or 0)
        prev_refunded_total = float(prev_refund_impact['total_refunded'] or 0)
        prev_net_total_sales = prev_gross_total_sales - prev_refunded_total

        refund_by_date = {str(r['return_date']): dict(r) for r in daily_refund_impact}
        daily_sales_list = [dict(row) for row in daily_sales]
        for d in daily_sales_list:
            day_key = str(d['sale_date'])
            impact = refund_by_date.get(day_key)
            base_daily_sales = float(d.get('daily_sales') or 0)
            refunded = float((impact or {}).get('refunded') or 0)
            d['gross_sales'] = base_daily_sales
            d['daily_sales'] = base_daily_sales - refunded

            if is_admin:
                base_profit = float(d.get('daily_profit') or 0)
                refunded_cogs = float((impact or {}).get('refunded_cogs') or 0)
                d['daily_profit'] = base_profit - refunded + refunded_cogs

        # Build response
        response = {
            'summary': {
                'transaction_count': summary['transaction_count'] or 0,
                'gross_sales': gross_total_sales,
                'total_sales': net_total_sales,
                'avg_transaction': summary['avg_transaction'] or 0,
                'total_discount': max(0.0, float(discount_summary['total_discount'] or 0) - float(refund_profit_impact['refunded_discount'] or 0)),
                'discounted_count': discount_summary['discounted_count'] or 0,
                'avg_discount': discount_summary['avg_discount'] or 0,
            },
            'daily_sales': daily_sales_list,
            'top_products': top_products_formatted,
            'product_performance': performance_formatted,
            'payment_breakdown': [dict(row) for row in payment_breakdown],
            'refund_summary': {
                'refund_count': refund_summary['refund_count'] or 0,
                'total_refunded': refund_summary['total_refunded'] or 0,
                'sales_with_refunds': refund_summary['sales_with_refunds'] or 0,
                'top_refunded': [{'name': n, 'quantity': q} for n, q in top_refunded]
            },
            'staff_performance': [dict(row) for row in staff_performance],
            'customer_insights': {
                'top_customers': [dict(row) for row in customer_query],
                'total_with_customer': total_with_customer['c'] or 0,
                'repeat_customers': repeat_customers['c'] or 0
            },
            'category_breakdown': category_breakdown,
            'hourly_distribution': [dict(row) for row in hourly_data],
            'period_comparison': {
                'prev_start': prev_start.strftime('%Y-%m-%d'),
                'prev_end': prev_end.strftime('%Y-%m-%d'),
                'prev_transaction_count': prev_summary['transaction_count'] or 0,
                'prev_gross_sales': prev_gross_total_sales,
                'prev_total_sales': prev_net_total_sales,
                'prev_avg_transaction': prev_summary['avg_transaction'] or 0
            }
        }

        if is_admin:
            base_profit = float(summary['total_profit'] or 0)
            cur_refunded = float(refund_profit_impact['total_refunded'] or 0)
            cur_refunded_cogs = float(refund_profit_impact['refunded_cogs'] or 0)
            response['summary']['total_profit'] = base_profit - cur_refunded + cur_refunded_cogs

            prev_base_profit = float(prev_summary['total_profit'] or 0)
            prev_refunded = float(prev_refund_impact['total_refunded'] or 0)
            prev_refunded_cogs = float(prev_refund_impact['refunded_cogs'] or 0)
            response['period_comparison']['prev_total_profit'] = prev_base_profit - prev_refunded + prev_refunded_cogs

        conn.close()
        return jsonify(response)

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/settings/currency', methods=['GET', 'PUT'])
@login_required
def manage_currency():
    shop_id = get_current_shop_id() or session.get('shop_id')
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'GET':
        shop = cursor.execute('SELECT currency_symbol FROM shops WHERE id = ?', (shop_id,)).fetchone()
        currency = shop['currency_symbol'] if shop else '$'
        conn.close()
        return jsonify({'currency': currency})

    # PUT - Update currency (admin only)
    if session.get('role') not in ('admin', 'super_admin', 'shop_owner'):
        conn.close()
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    data = request.json
    currency = data.get('currency', '').strip()

    if not currency:
        conn.close()
        return jsonify({'success': False, 'error': 'Currency symbol cannot be empty'}), 400

    try:
        cursor.execute('UPDATE shops SET currency_symbol = ? WHERE id = ?', (currency, shop_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'currency': currency})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/settings/shop-name', methods=['GET', 'PUT'])
@login_required
def manage_shop_name():
    shop_id = get_current_shop_id() or session.get('shop_id')
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'GET':
        shop = cursor.execute('SELECT name, icon FROM shops WHERE id = ?', (shop_id,)).fetchone()
        conn.close()
        return jsonify({
            'shop_name': shop['name'] if shop else 'POS System',
            'shop_icon': shop['icon'] if shop else '🛒'
        })

    # PUT - Update shop name/icon (admin only)
    if session.get('role') not in ('admin', 'super_admin', 'shop_owner'):
        conn.close()
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    data = request.json
    shop_name = data.get('shop_name', '').strip()
    shop_icon = data.get('shop_icon', '').strip()

    if not shop_name:
        conn.close()
        return jsonify({'success': False, 'error': 'Shop name cannot be empty'}), 400

    try:
        cursor.execute(
            'UPDATE shops SET name = ?, icon = ? WHERE id = ?',
            (shop_name, shop_icon or '🛒', shop_id)
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'shop_name': shop_name, 'shop_icon': shop_icon})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/settings/brand', methods=['GET', 'PUT'])
@login_required
def manage_brand_name():
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'GET':
        row = cursor.execute(
            "SELECT value FROM settings WHERE `key` = 'brand_name' AND shop_id IS NULL"
        ).fetchone()
        conn.close()
        return jsonify({'brand_name': row['value'] if row else 'POS System'})

    # PUT - Update brand name (admin only)
    if session.get('role') not in ('admin', 'super_admin', 'shop_owner'):
        conn.close()
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    data = request.json
    brand_name = data.get('brand_name', '').strip()

    if not brand_name:
        conn.close()
        return jsonify({'success': False, 'error': 'Brand name cannot be empty'}), 400

    try:
        # Check if brand_name setting already exists
        existing = cursor.execute(
            "SELECT value FROM settings WHERE `key` = 'brand_name' AND shop_id IS NULL"
        ).fetchone()
        if existing:
            cursor.execute(
                "UPDATE settings SET value = ? WHERE `key` = 'brand_name' AND shop_id IS NULL",
                (brand_name,)
            )
        else:
            cursor.execute(
                "INSERT INTO settings (`key`, value, shop_id) VALUES ('brand_name', ?, NULL)",
                (brand_name,)
            )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'brand_name': brand_name})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/settings/session-timeout', methods=['GET', 'PUT'])
@login_required
def manage_session_timeout():
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'GET':
        timeout_row = cursor.execute(
            'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', ('session_timeout',)
        ).fetchone()
        timeout = int(timeout_row['value']) if timeout_row else 1800
        conn.close()
        return jsonify({'session_timeout': timeout})

    # PUT - Update session timeout (admin only)
    if session.get('role') != 'admin':
        conn.close()
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    data = request.json
    timeout = data.get('session_timeout')

    if timeout is None:
        conn.close()
        return jsonify({'success': False, 'error': 'Session timeout is required'}), 400

    try:
        timeout = int(timeout)
        if timeout < 300:  # Minimum 5 minutes
            conn.close()
            return jsonify({'success': False, 'error': 'Session timeout must be at least 5 minutes (300 seconds)'}), 400
        if timeout > 86400:  # Maximum 24 hours
            conn.close()
            return jsonify({'success': False, 'error': 'Session timeout cannot exceed 24 hours (86400 seconds)'}), 400
    except ValueError:
        conn.close()
        return jsonify({'success': False, 'error': 'Invalid timeout value'}), 400

    try:
        cursor.execute(
            """
            INSERT INTO settings (`key`, value, shop_id)
            VALUES (?, ?, NULL)
            ON DUPLICATE KEY UPDATE value = VALUES(value)
            """,
            ('session_timeout', str(timeout))
        )

        conn.commit()
        conn.close()
        return jsonify({'success': True, 'session_timeout': timeout})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/reports/inventory', methods=['GET'])
@login_required
def get_inventory_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    is_admin = session.get('role') in ('admin', 'super_admin', 'shop_owner')
    flt_sql, flt_params = shop_filter()
    flt_sql_sh, flt_params_sh = shop_filter('sh')

    try:
        # Add time to include the entire end date
        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # Current inventory status
        if is_admin:
            inventory_query = f'''
                SELECT
                    name, sku, category, stock_quantity, cost_price, price,
                    (stock_quantity * cost_price) as stock_value,
                    (stock_quantity * price) as potential_revenue,
                    (stock_quantity * (price - cost_price)) as potential_profit
                FROM products
                WHERE is_active = 1 {flt_sql}
                ORDER BY stock_value DESC
            '''
        else:
            inventory_query = f'''
                SELECT name, sku, category, stock_quantity, price
                FROM products
                WHERE is_active = 1 {flt_sql}
                ORDER BY stock_quantity DESC
            '''

        inventory = cursor.execute(inventory_query, flt_params).fetchall()

        # Stock movements in date range
        stock_movements_query = f'''
            SELECT
                DATE(sh.created_at) as date,
                sh.action_type,
                COUNT(*) as transaction_count,
                SUM(CASE WHEN sh.quantity_change > 0 THEN sh.quantity_change ELSE 0 END) as items_added,
                SUM(CASE WHEN sh.quantity_change < 0 THEN ABS(sh.quantity_change) ELSE 0 END) as items_removed
            FROM stock_history sh
            WHERE sh.created_at BETWEEN ? AND ? {flt_sql_sh}
            GROUP BY DATE(sh.created_at), sh.action_type
            ORDER BY date DESC
        '''

        stock_movements = cursor.execute(stock_movements_query, [start_date_time, end_date_time] + flt_params_sh).fetchall()

        # Top stock activities (products with most movement)
        top_movements_query = f'''
            SELECT
                p.name, p.sku,
                COUNT(*) as activity_count,
                SUM(CASE WHEN sh.quantity_change > 0 THEN sh.quantity_change ELSE 0 END) as total_added,
                SUM(CASE WHEN sh.quantity_change < 0 THEN ABS(sh.quantity_change) ELSE 0 END) as total_removed
            FROM stock_history sh
            JOIN products p ON sh.product_id = p.id
            WHERE sh.created_at BETWEEN ? AND ? {flt_sql_sh}
            GROUP BY sh.product_id, p.name, p.sku
            ORDER BY activity_count DESC
            LIMIT 10
        '''

        top_movements = cursor.execute(top_movements_query, [start_date_time, end_date_time] + flt_params_sh).fetchall()

        # Low stock items (count first, then fetch top 10 only)
        low_stock_count = cursor.execute(
            f'SELECT COUNT(*) as cnt FROM products WHERE is_active = 1 AND stock_quantity < 5 {flt_sql}',
            flt_params
        ).fetchone()['cnt']

        low_stock_query = f'''
            SELECT name, sku, stock_quantity, category
            FROM products
            WHERE is_active = 1 AND stock_quantity < 5 {flt_sql}
            ORDER BY stock_quantity ASC
            LIMIT 10
        '''

        low_stock = cursor.execute(low_stock_query, flt_params).fetchall()

        # Summary statistics
        summary = {
            'total_products': len(inventory),
            'low_stock_items': low_stock_count,
            'total_stock_quantity': sum(item['stock_quantity'] for item in inventory)
        }

        if is_admin:
            summary['total_stock_value'] = sum(item['stock_value'] for item in inventory)
            summary['potential_revenue'] = sum(item['potential_revenue'] for item in inventory)
            summary['potential_profit'] = sum(item['potential_profit'] for item in inventory)

        # Slow-moving stock: products with 0 or very few sales in the period
        slow_moving_query = f'''
            SELECT p.id, p.name, p.sku, p.category, p.stock_quantity, p.price
            FROM products p
            WHERE p.is_active = 1 AND p.stock_quantity > 0 {flt_sql}
            ORDER BY p.name
        '''
        all_products_list = cursor.execute(slow_moving_query, flt_params).fetchall()

        # Get sales counts per product in this period
        flt_sql_s, flt_params_s = shop_filter('s')
        sales_in_period = cursor.execute(f'''
            SELECT items_json FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
        ''', [start_date_time, end_date_time] + flt_params_s).fetchall()

        sold_qty = {}
        for sale_row in sales_in_period:
            items = json.loads(sale_row['items_json'])
            for item in items:
                pid = item['product_id']
                sold_qty[pid] = sold_qty.get(pid, 0) + item['quantity']

        slow_moving = []
        for p in all_products_list:
            qty_sold = sold_qty.get(p['id'], 0)
            if qty_sold <= 1:
                slow_moving.append({
                    'name': p['name'], 'sku': p['sku'] or '-',
                    'category': p['category'], 'stock_quantity': p['stock_quantity'],
                    'quantity_sold': qty_sold,
                    'stock_value': p['stock_quantity'] * p['price']
                })
        slow_moving.sort(key=lambda x: x['quantity_sold'])

        # Build response
        response = {
            'summary': summary,
            'inventory': [dict(row) for row in inventory],
            'stock_movements': [dict(row) for row in stock_movements],
            'top_movements': [dict(row) for row in top_movements],
            'low_stock': [dict(row) for row in low_stock],
            'low_stock_total': low_stock_count,
            'slow_moving': slow_moving
        }

        conn.close()
        return jsonify(response)

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/reports/cashflow', methods=['GET'])
@login_required
@admin_required
def get_cashflow_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql_s, flt_params_s = shop_filter('s')
    flt_sql_i, flt_params_i = shop_filter('i')
    flt_sql_p, flt_params_p = shop_filter('p')
    flt_sql_c, flt_params_c = shop_filter('c')
    flt_sql_sr, flt_params_sr = shop_filter('sr')

    try:
        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # Summary: Cash collected in period (walk-in sales + payments on invoices)
        walkin_cash = cursor.execute(f'''
            SELECT COALESCE(SUM(s.total_amount), 0) as total
            FROM sales s
            WHERE s.sale_date BETWEEN ? AND ? {flt_sql_s}
            AND s.id NOT IN (SELECT sale_id FROM invoices)
        ''', [start_date_time, end_date_time] + flt_params_s).fetchone()['total']

        payments_collected = cursor.execute(f'''
            SELECT COALESCE(SUM(amount), 0) as total
            FROM payments p
            WHERE created_at BETWEEN ? AND ? {flt_sql_p}
        ''', [start_date_time, end_date_time] + flt_params_p).fetchone()['total']

        total_refunds = cursor.execute(f'''
            SELECT COALESCE(SUM(sr.refund_amount), 0) as total
            FROM sale_returns sr
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
        ''', [start_date_time, end_date_time] + flt_params_sr).fetchone()['total']

        total_collected = walkin_cash + payments_collected - total_refunds

        # Credit given in period
        credit_given = cursor.execute(f'''
            SELECT COALESCE(SUM(total_amount - paid_amount), 0) as total,
                   COUNT(*) as count
            FROM invoices i
            WHERE created_at BETWEEN ? AND ? {flt_sql_i}
        ''', [start_date_time, end_date_time] + flt_params_i).fetchone()

        # Total outstanding (current)
        total_outstanding = cursor.execute(f'''
            SELECT COALESCE(SUM(balance), 0) as total
            FROM customers c WHERE balance > 0 {flt_sql_c}
        ''', flt_params_c).fetchone()['total']

        # Overdue invoices
        overdue = cursor.execute(f'''
            SELECT COUNT(*) as count, COALESCE(SUM(total_amount - paid_amount), 0) as amount
            FROM invoices i
WHERE status != 'paid' AND due_date < CURDATE() AND due_date IS NOT NULL {flt_sql_i}
        ''', flt_params_i).fetchone()

        # Daily cash flow breakdown
        daily_walkin = cursor.execute(f'''
            SELECT DATE(s.sale_date) as date, COALESCE(SUM(s.total_amount), 0) as cash_in
            FROM sales s
            WHERE s.sale_date BETWEEN ? AND ? {flt_sql_s}
            AND s.id NOT IN (SELECT sale_id FROM invoices)
            GROUP BY DATE(s.sale_date)
        ''', [start_date_time, end_date_time] + flt_params_s).fetchall()

        daily_payments = cursor.execute(f'''
            SELECT DATE(created_at) as date, COALESCE(SUM(amount), 0) as cash_in
            FROM payments p
            WHERE created_at BETWEEN ? AND ? {flt_sql_p}
            GROUP BY DATE(created_at)
        ''', [start_date_time, end_date_time] + flt_params_p).fetchall()

        daily_refunds = cursor.execute(f'''
            SELECT DATE(sr.return_date) as date, COALESCE(SUM(sr.refund_amount), 0) as cash_out
            FROM sale_returns sr
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
            GROUP BY DATE(sr.return_date)
        ''', [start_date_time, end_date_time] + flt_params_sr).fetchall()

        daily_credit = cursor.execute(f'''
            SELECT DATE(created_at) as date, COALESCE(SUM(total_amount - paid_amount), 0) as credit_out
            FROM invoices i
            WHERE created_at BETWEEN ? AND ? {flt_sql_i}
            GROUP BY DATE(created_at)
        ''', [start_date_time, end_date_time] + flt_params_i).fetchall()

        # Merge daily data
        daily_map = {}
        for row in daily_walkin:
            d = row['date']
            if d not in daily_map:
                daily_map[d] = {'date': d, 'cash_collected': 0, 'credit_given': 0}
            daily_map[d]['cash_collected'] += row['cash_in']
        for row in daily_payments:
            d = row['date']
            if d not in daily_map:
                daily_map[d] = {'date': d, 'cash_collected': 0, 'credit_given': 0}
            daily_map[d]['cash_collected'] += row['cash_in']
        for row in daily_credit:
            d = row['date']
            if d not in daily_map:
                daily_map[d] = {'date': d, 'cash_collected': 0, 'credit_given': 0}
            daily_map[d]['credit_given'] += row['credit_out']
        for row in daily_refunds:
            d = row['date']
            if d not in daily_map:
                daily_map[d] = {'date': d, 'cash_collected': 0, 'credit_given': 0}
            daily_map[d]['cash_collected'] -= row['cash_out']

        daily_cashflow = sorted(daily_map.values(), key=lambda x: x['date'])

        # Outstanding by customer
        customer_outstanding = cursor.execute(f'''
            SELECT c.id, c.name, c.phone, c.credit_limit, c.balance,
                   COUNT(i.id) as invoice_count,
                   MIN(i.due_date) as earliest_due
            FROM customers c
            LEFT JOIN invoices i ON i.customer_id = c.id AND i.status != 'paid'
            WHERE c.balance > 0 {flt_sql_c}
            GROUP BY c.id
            ORDER BY c.balance DESC
        ''', flt_params_c).fetchall()

        # Aging breakdown
        aging = cursor.execute(f'''
            SELECT
                SUM(CASE WHEN DATEDIFF(NOW(), created_at) <= 30
                    THEN total_amount - paid_amount ELSE 0 END) as current_amount,
                SUM(CASE WHEN DATEDIFF(NOW(), created_at) > 30
                    AND DATEDIFF(NOW(), created_at) <= 60
                    THEN total_amount - paid_amount ELSE 0 END) as days_30,
                SUM(CASE WHEN DATEDIFF(NOW(), created_at) > 60
                    AND DATEDIFF(NOW(), created_at) <= 90
                    THEN total_amount - paid_amount ELSE 0 END) as days_60,
                SUM(CASE WHEN DATEDIFF(NOW(), created_at) > 90
                    THEN total_amount - paid_amount ELSE 0 END) as days_90_plus
            FROM invoices i
            WHERE status != 'paid' {flt_sql_i}
        ''', flt_params_i).fetchone()

        # Payment method breakdown for collections
        payment_methods = cursor.execute(f'''
            SELECT payment_method, COUNT(*) as count, SUM(amount) as total
            FROM payments p
            WHERE created_at BETWEEN ? AND ? {flt_sql_p}
            GROUP BY payment_method
        ''', [start_date_time, end_date_time] + flt_params_p).fetchall()

        response = {
            'summary': {
                'total_collected': total_collected,
                'walkin_cash': walkin_cash,
                'payments_collected': payments_collected,
                'total_refunds': total_refunds,
                'credit_given': credit_given['total'],
                'credit_invoice_count': credit_given['count'],
                'total_outstanding': total_outstanding,
                'overdue_count': overdue['count'],
                'overdue_amount': overdue['amount'],
                'net_cash_flow': total_collected - credit_given['total']
            },
            'daily_cashflow': daily_cashflow,
            'customer_outstanding': [dict(row) for row in customer_outstanding],
            'aging': {
                'current': aging['current_amount'] or 0,
                'days_30': aging['days_30'] or 0,
                'days_60': aging['days_60'] or 0,
                'days_90_plus': aging['days_90_plus'] or 0
            },
            'payment_methods': [dict(row) for row in payment_methods]
        }

        conn.close()
        return jsonify(response)

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/change-password', methods=['PUT'])
@login_required
def change_own_password():
    data = request.json
    current_password = data.get('current_password', '').strip()
    new_password = data.get('new_password', '').strip()

    if not current_password or not new_password:
        return jsonify({'success': False, 'error': 'Current and new passwords are required'}), 400

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'New password must be at least 6 characters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verify current password
        user = cursor.execute('SELECT password FROM users WHERE id = ?', (session.get('user_id'),)).fetchone()

        if not user or user['password'] != hash_password(current_password):
            return jsonify({'success': False, 'error': 'Current password is incorrect'}), 401

        # Update to new password
        cursor.execute('''
            UPDATE users
            SET password = ?
            WHERE id = ?
        ''', (hash_password(new_password), session.get('user_id')))

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# API Routes - Dashboard Stats
@app.route('/api/dashboard-stats', methods=['GET'])
@login_required
def get_dashboard_stats():
    conn = get_db_connection()
    cursor = conn.cursor()
    is_admin = session.get('role') in ('admin', 'super_admin', 'shop_owner')
    flt_sql, flt_params = shop_filter()
    flt_sql_s, flt_params_s = shop_filter('s')
    flt_sql_sr, flt_params_sr = shop_filter('sr')

    # Total products
    total_products = conn.execute(
        f'SELECT COUNT(*) as count FROM products WHERE 1=1 {flt_sql}', flt_params
    ).fetchone()['count']

    # Low stock items (less than 5)
    low_stock = conn.execute(
        f'SELECT COUNT(*) as count FROM products WHERE stock_quantity > 0 AND stock_quantity < 5 {flt_sql}', flt_params
    ).fetchone()['count']

    # Out of stock items
    out_of_stock = conn.execute(
        f'SELECT COUNT(*) as count FROM products WHERE stock_quantity <= 0 {flt_sql}', flt_params
    ).fetchone()['count']

    # Today's sales
    today_sales = conn.execute(f'''
        SELECT COALESCE(SUM(total_amount), 0) as total, COALESCE(SUM(total_cost), 0) as cost, COALESCE(SUM(discount_amount), 0) as discount
        FROM sales s
        WHERE DATE(sale_date) = CURDATE() {flt_sql_s}
    ''', flt_params_s).fetchone()

    # Yesterday's sales
    yesterday_sales = conn.execute(f'''
        SELECT COALESCE(SUM(total_amount), 0) as total, COALESCE(SUM(total_cost), 0) as cost
        FROM sales s
        WHERE DATE(sale_date) = DATE_SUB(CURDATE(), INTERVAL 1 DAY) {flt_sql_s}
    ''', flt_params_s).fetchone()

    # Refund impact for today/yesterday (revenue and COGS reversal)
    today_refunds = conn.execute(f'''
        SELECT
            COALESCE(SUM(sr.refund_amount), 0) as refunded,
            COALESCE(SUM(
                CASE
                    WHEN COALESCE(s.total_amount, 0) > 0
                        THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                    ELSE 0
                END
            ), 0) as refunded_cogs,
            COALESCE(SUM(
                CASE
                    WHEN COALESCE(s.total_amount, 0) > 0
                        THEN LEAST(sr.refund_amount, s.total_amount) / s.total_amount * COALESCE(s.discount_amount, 0)
                    ELSE 0
                END
            ), 0) as refunded_discount
        FROM sale_returns sr
        JOIN sales s ON s.id = sr.sale_id
        WHERE DATE(sr.return_date) = CURDATE() {flt_sql_sr}
    ''', flt_params_sr).fetchone()

    yesterday_refunds = conn.execute(f'''
        SELECT
            COALESCE(SUM(sr.refund_amount), 0) as refunded,
            COALESCE(SUM(
                CASE
                    WHEN COALESCE(s.total_amount, 0) > 0
                        THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                    ELSE 0
                END
            ), 0) as refunded_cogs,
            COALESCE(SUM(
                CASE
                    WHEN COALESCE(s.total_amount, 0) > 0
                        THEN LEAST(sr.refund_amount, s.total_amount) / s.total_amount * COALESCE(s.discount_amount, 0)
                    ELSE 0
                END
            ), 0) as refunded_discount
        FROM sale_returns sr
        JOIN sales s ON s.id = sr.sale_id
        WHERE DATE(sr.return_date) = DATE_SUB(CURDATE(), INTERVAL 1 DAY) {flt_sql_sr}
    ''', flt_params_sr).fetchone()

    # Transaction counts should exclude sales that have been fully refunded.
    today_count = conn.execute(f'''
        SELECT COUNT(*) as count
        FROM sales s
        LEFT JOIN (
            SELECT sale_id, COALESCE(SUM(refund_amount), 0) as total_refunded
            FROM sale_returns
            GROUP BY sale_id
        ) sr ON sr.sale_id = s.id
        WHERE DATE(s.sale_date) = CURDATE() {flt_sql_s}
          AND COALESCE(sr.total_refunded, 0) < COALESCE(s.total_amount, 0)
    ''', flt_params_s).fetchone()['count']

    # Yesterday's count
    yesterday_count = conn.execute(f'''
        SELECT COUNT(*) as count
        FROM sales s
        LEFT JOIN (
            SELECT sale_id, COALESCE(SUM(refund_amount), 0) as total_refunded
            FROM sale_returns
            GROUP BY sale_id
        ) sr ON sr.sale_id = s.id
        WHERE DATE(s.sale_date) = DATE_SUB(CURDATE(), INTERVAL 1 DAY) {flt_sql_s}
          AND COALESCE(sr.total_refunded, 0) < COALESCE(s.total_amount, 0)
    ''', flt_params_s).fetchone()['count']

    # Last 7 days sales trend
    sales_trend = cursor.execute(f'''
        SELECT
            DATE(sale_date) as date,
            COALESCE(SUM(total_amount), 0) as total_sales,
            COALESCE(SUM(total_cost), 0) as total_cost,
            COUNT(*) as transaction_count
        FROM sales s
        WHERE DATE(sale_date) >= DATE_SUB(CURDATE(), INTERVAL 6 DAY) {flt_sql_s}
        GROUP BY DATE(sale_date)
        ORDER BY date ASC
    ''', flt_params_s).fetchall()

    refund_trend = cursor.execute(f'''
        SELECT
            DATE(sr.return_date) as date,
            COALESCE(SUM(sr.refund_amount), 0) as refunded,
            COALESCE(SUM(
                CASE
                    WHEN COALESCE(s.total_amount, 0) > 0
                        THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                    ELSE 0
                END
            ), 0) as refunded_cogs
        FROM sale_returns sr
        JOIN sales s ON s.id = sr.sale_id
        WHERE DATE(sr.return_date) >= DATE_SUB(CURDATE(), INTERVAL 6 DAY) {flt_sql_sr}
        GROUP BY DATE(sr.return_date)
        ORDER BY date ASC
    ''', flt_params_sr).fetchall()

    sales_trend_list = [dict(row) for row in sales_trend]
    refund_trend_map = {str(r['date']): dict(r) for r in refund_trend}
    for d in sales_trend_list:
        day_key = str(d['date'])
        impact = refund_trend_map.get(day_key)
        if not impact:
            continue
        d['total_sales'] = float(d.get('total_sales') or 0) - float(impact.get('refunded') or 0)
        d['total_cost'] = float(d.get('total_cost') or 0) - float(impact.get('refunded_cogs') or 0)

    today_total_net = float(today_sales['total'] or 0) - float(today_refunds['refunded'] or 0)
    today_cost_net = float(today_sales['cost'] or 0) - float(today_refunds['refunded_cogs'] or 0)
    yesterday_total_net = float(yesterday_sales['total'] or 0) - float(yesterday_refunds['refunded'] or 0)
    yesterday_cost_net = float(yesterday_sales['cost'] or 0) - float(yesterday_refunds['refunded_cogs'] or 0)

    # Best selling products (last 7 days)
    recent_items_json = cursor.execute(f'''
        SELECT items_json FROM sales s
        WHERE DATE(sale_date) >= DATE_SUB(CURDATE(), INTERVAL 6 DAY)
        AND items_json IS NOT NULL {flt_sql_s}
    ''', flt_params_s).fetchall()
    product_totals = {}
    for row in recent_items_json:
        try:
            for item in json.loads(row['items_json']):
                pid = item.get('product_id')
                if pid:
                    product_totals[pid] = product_totals.get(pid, 0) + item.get('quantity', 0)
        except Exception:
            pass
    top_pids = sorted(product_totals, key=product_totals.get, reverse=True)[:5]
    best_sellers = []
    for pid in top_pids:
        prod = cursor.execute(
            'SELECT name, sku, price FROM products WHERE id = ?', (pid,)
        ).fetchone()
        if prod:
            best_sellers.append({
                'name': prod['name'], 'sku': prod['sku'],
                'price': prod['price'], 'total_sold': product_totals[pid]
            })

    # Recent sales
    recent_sales = cursor.execute(f'''
        SELECT
            s.id,
            s.total_amount,
            s.discount_amount,
            s.sale_date,
            u.username
        FROM sales s
        LEFT JOIN users u ON s.user_id = u.id
        WHERE 1=1 {flt_sql_s}
        ORDER BY s.sale_date DESC
        LIMIT 5
    ''', flt_params_s).fetchall()

    # Average order value for today to match the daily dashboard KPIs.
    avg_transaction = (today_total_net / today_count) if today_count else 0

    # Cash flow stats (admin only)
    cash_flow = {}
    if is_admin:
        # Cash collected today = walk-in sales + payments received today
        walkin_cash_today = conn.execute(f'''
            SELECT COALESCE(SUM(s.total_amount), 0) as total
            FROM sales s
            WHERE DATE(s.sale_date) = CURDATE() {flt_sql_s}
            AND s.id NOT IN (SELECT sale_id FROM invoices)
        ''', flt_params_s).fetchone()['total']

        flt_sql_p, flt_params_p = shop_filter('p')
        payments_today = conn.execute(f'''
            SELECT COALESCE(SUM(amount), 0) as total
            FROM payments p
            WHERE DATE(created_at) = CURDATE() {flt_sql_p}
        ''', flt_params_p).fetchone()['total']

        cash_flow['cash_collected_today'] = walkin_cash_today + payments_today

        # Credit given today
        flt_sql_i, flt_params_i = shop_filter('i')
        credit_today = conn.execute(f'''
            SELECT COALESCE(SUM(total_amount - paid_amount), 0) as total
            FROM invoices i
            WHERE DATE(created_at) = CURDATE() {flt_sql_i}
        ''', flt_params_i).fetchone()['total']
        cash_flow['credit_sales_today'] = credit_today

        # Total outstanding across all customers
        flt_sql_c, flt_params_c = shop_filter('c')
        total_outstanding = conn.execute(f'''
            SELECT COALESCE(SUM(balance), 0) as total
            FROM customers c
            WHERE balance > 0 {flt_sql_c}
        ''', flt_params_c).fetchone()['total']
        cash_flow['total_outstanding'] = total_outstanding

        # Overdue invoices count and amount
        overdue = conn.execute(f'''
            SELECT COUNT(*) as count, COALESCE(SUM(total_amount - paid_amount), 0) as amount
            FROM invoices i
            WHERE status != 'paid' AND due_date < CURDATE() AND due_date IS NOT NULL {flt_sql_i}
        ''', flt_params_i).fetchone()
        cash_flow['overdue_count'] = overdue['count']
        cash_flow['overdue_amount'] = overdue['amount']

        funds = _get_available_funds(cursor)
        cash_flow['available_funds'] = funds['available_funds']
        cash_flow['money_in_total'] = funds['money_in']
        cash_flow['money_out_total'] = funds['money_out']

    conn.close()

    stats = {
        'total_products': total_products,
        'low_stock_items': low_stock,
        'out_of_stock_items': out_of_stock,
        'today_sales': today_total_net,
        'today_sales_count': today_count,
        'yesterday_sales': yesterday_total_net,
        'yesterday_count': yesterday_count,
        'sales_trend': sales_trend_list,
        'best_sellers': [dict(row) for row in best_sellers],
        'recent_sales': [dict(row) for row in recent_sales],
        'avg_transaction': avg_transaction,
        'total_discount_today': max(0.0, float(today_sales['discount'] or 0) - float(today_refunds['refunded_discount'] or 0))
    }

    # Only show profit to admins
    if is_admin:
        stats['today_cost'] = today_cost_net
        stats['today_profit'] = today_total_net - today_cost_net
        stats['yesterday_profit'] = yesterday_total_net - yesterday_cost_net
        stats['cash_flow'] = cash_flow

    return jsonify(stats)

# Email and Alert Settings
@app.route('/api/settings/email', methods=['GET', 'PUT'])
@admin_required
def manage_email_settings():
    """Manage email settings for alerts — scoped per shop (super_admin uses shop_id=NULL global)"""
    role = session.get('role')
    # super_admin with no active shop → edit global (NULL) settings
    # admin → edit their own shop's settings
    shop_id = get_settings_shop_id()  # super_admin no-shop→global; shop_owner→first assigned; admin→own shop

    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'GET':
        settings = get_shop_email_settings(cursor, shop_id)
        # Never return the actual password — just indicate if it is set
        settings['password_set'] = bool(settings.get('smtp_password'))
        settings.pop('smtp_password', None)
        conn.close()
        return jsonify(settings)

    # PUT - Update email settings
    data = request.json

    required_keys = ['smtp_server', 'smtp_port', 'smtp_username', 'alert_email', 'low_stock_threshold']
    for key in required_keys:
        if key not in data:
            conn.close()
            return jsonify({'success': False, 'error': f'{key} is required'}), 400

    try:
        smtp_port = int(data['smtp_port'])
        if smtp_port < 1 or smtp_port > 65535:
            raise ValueError('Invalid port number')

        threshold = int(data['low_stock_threshold'])
        if threshold < 0:
            raise ValueError('Threshold must be non-negative')

        settings_to_update = {
            'smtp_server': data['smtp_server'],
            'smtp_port': str(smtp_port),
            'smtp_username': data['smtp_username'],
            'alert_email': data['alert_email'],
            'low_stock_threshold': str(threshold),
        }
        if 'smtp_password' in data and data['smtp_password']:
            settings_to_update['smtp_password'] = data['smtp_password']

        for key, value in settings_to_update.items():
            if shop_id is not None:
                cursor.execute(
                    """
                    INSERT INTO settings (`key`, value, shop_id)
                    VALUES (?, ?, ?)
                    ON DUPLICATE KEY UPDATE value = VALUES(value)
                    """,
                    (key, value, shop_id)
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO settings (`key`, value, shop_id)
                    VALUES (?, ?, NULL)
                    ON DUPLICATE KEY UPDATE value = VALUES(value)
                    """,
                    (key, value)
                )

        conn.commit()
        conn.close()
        return jsonify({'success': True})

    except ValueError as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/alerts/test-email', methods=['POST'])
@admin_required
def send_test_email():
    """Send a test email to verify configuration — uses the current shop's settings"""
    role = session.get('role')
    shop_id = get_settings_shop_id()  # super_admin no-shop→global; shop_owner→first assigned; admin→own shop

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        settings = get_shop_email_settings(cursor, shop_id)

        # Get shop name
        if shop_id:
            shop_row = cursor.execute('SELECT name FROM shops WHERE id = ?', (shop_id,)).fetchone()
            shop_name = shop_row['name'] if shop_row else 'POS System'
        else:
            shop_name = 'POS System (Global)'

        conn.close()

        if not settings['smtp_server'] or not settings['alert_email']:
            return jsonify({'success': False, 'error': 'Email settings not configured for this shop'}), 400

        send_email(
            settings,
            settings['alert_email'],
            f'{shop_name} - Test Email',
            f'<h2>Email Configuration Test</h2><p>This is a test email from your POS system ({shop_name}).</p><p>If you received this, your email settings are configured correctly!</p>',
            from_name=shop_name
        )

        return jsonify({'success': True, 'message': 'Test email sent successfully'})

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/alerts/check-low-stock', methods=['POST'])
@admin_required
def check_low_stock():
    """Check for low stock items and send email alert — scoped to current shop"""
    role = session.get('role')
    shop_id = get_settings_shop_id()  # super_admin no-shop→global; shop_owner→first assigned; admin→own shop

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        email_settings = get_shop_email_settings(cursor, shop_id)
        alert_email = email_settings.get('alert_email', '')
        threshold = int(email_settings.get('low_stock_threshold') or 5)

        if not alert_email:
            conn.close()
            return jsonify({'success': False, 'error': 'Alert email not configured for this shop'}), 400

        # Get low stock products scoped to this shop
        flt_sql, flt_params = shop_filter()
        low_stock_products = cursor.execute(f'''
            SELECT id, name, sku, category, size, color, stock_quantity,
                   COALESCE(low_stock_threshold, ?) as threshold
            FROM products
                        WHERE is_active = 1
                            AND stock_quantity <= COALESCE(low_stock_threshold, ?)
            {('AND ' + flt_sql) if flt_sql else ''}
            ORDER BY stock_quantity ASC
        ''', (threshold, threshold, *flt_params)).fetchall()

        if not low_stock_products:
            conn.close()
            return jsonify({'success': True, 'message': 'No low stock items found', 'count': 0})

        # Get shop name
        if shop_id:
            shop_row = cursor.execute('SELECT name FROM shops WHERE id = ?', (shop_id,)).fetchone()
            shop_name = shop_row['name'] if shop_row else 'POS System'
        else:
            shop_name = 'POS System (Global)'

        conn.close()

        # Generate email body
        products_html = '<table class="low-stock-table" style="width: 100%; border-collapse: collapse; margin-top: 20px; table-layout: fixed;">'
        products_html += '<tr class="low-stock-head" style="background-color: #f8f9fa; border-bottom: 2px solid #dee2e6;">'
        products_html += '<th style="padding: 12px; text-align: left; width: 40%;">Product</th>'
        products_html += '<th style="padding: 12px; text-align: left; width: 20%;">SKU</th>'
        products_html += '<th style="padding: 12px; text-align: center; width: 14%;">Size</th>'
        products_html += '<th style="padding: 12px; text-align: left; width: 16%;">Color</th>'
        products_html += '<th style="padding: 12px; text-align: center; width: 10%;">Stock</th>'
        products_html += '</tr>'

        plain_lines = [
            f'Low Stock Alert - {shop_name}',
            '',
            f'Total low stock items: {len(low_stock_products)}',
            '',
            'Product | SKU | Size | Color | Stock',
            '-' * 90,
        ]

        for product in low_stock_products:
            products_html += '<tr class="low-stock-row" style="border-bottom: 1px solid #dee2e6;">'
            products_html += f'<td class="low-stock-cell" data-label="Product" style="padding: 10px; word-break: break-word; overflow-wrap: anywhere;">{product["name"]}</td>'
            products_html += f'<td class="low-stock-cell" data-label="SKU" style="padding: 10px; word-break: break-word; overflow-wrap: anywhere;">{product["sku"] or "-"}</td>'
            products_html += f'<td class="low-stock-cell" data-label="Size" style="padding: 10px; text-align: center; word-break: break-word;">{product["size"] or "-"}</td>'
            products_html += f'<td class="low-stock-cell" data-label="Color" style="padding: 10px; word-break: break-word; overflow-wrap: anywhere;">{product["color"] or "-"}</td>'
            products_html += f'<td class="low-stock-cell low-stock-value" data-label="Stock" style="padding: 10px; text-align: center; color: #e74c3c; font-weight: bold;">{product["stock_quantity"]}</td>'
            products_html += '</tr>'
            plain_lines.append(
                f'{product["name"]} | {product["sku"] or "-"} | {product["size"] or "-"} | '
                f'{product["color"] or "-"} | {product["stock_quantity"]}'
            )

        products_html += '</table>'
        plain_text = '\n'.join(plain_lines)

        email_body = f'''
        <html>
        <head>
            <style>
                @media only screen and (max-width: 600px) {{
                    .email-wrap {{ padding: 12px !important; }}
                    .low-stock-table {{ table-layout: auto !important; }}
                    .low-stock-head {{ display: none !important; }}
                    .low-stock-row {{
                        display: block !important;
                        border: 1px solid #e5e7eb !important;
                        border-radius: 8px !important;
                        margin-bottom: 10px !important;
                    }}
                    .low-stock-cell {{
                        display: block !important;
                        width: 100% !important;
                        box-sizing: border-box !important;
                        text-align: left !important;
                        padding: 8px 10px !important;
                        border-bottom: 1px solid #f1f5f9 !important;
                    }}
                    .low-stock-cell:last-child {{ border-bottom: none !important; }}
                    .low-stock-cell::before {{
                        content: attr(data-label) ': ';
                        font-weight: 700;
                        color: #374151;
                    }}
                    .low-stock-value {{ color: #e74c3c !important; font-weight: 700 !important; }}
                }}
            </style>
        </head>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div class="email-wrap" style="max-width: 800px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #e74c3c; border-bottom: 3px solid #e74c3c; padding-bottom: 10px;">
                    ⚠️ Low Stock Alert
                </h2>
                <p>The following products are running low on stock and need to be reordered:</p>
                <p><strong>Total low stock items: {len(low_stock_products)}</strong></p>
                {products_html}
                <div style="margin-top: 30px; padding: 15px; background-color: #fff3cd; border-left: 4px solid #ffc107;">
                    <p style="margin: 0;"><strong>Note:</strong> Please restock these items as soon as possible to avoid running out of inventory.</p>
                </div>
                <p style="margin-top: 20px; color: #6c757d; font-size: 0.9em;">
                    This is an automated alert from {shop_name}.
                </p>
            </div>
        </body>
        </html>
        '''

        # Send email
        send_email(
            email_settings,
            alert_email,
            f'{shop_name} - Low Stock Alert ({len(low_stock_products)} items)',
            email_body,
            from_name=shop_name,
            text_body=plain_text
        )

        return jsonify({
            'success': True,
            'message': f'Alert sent for {len(low_stock_products)} low stock items',
            'count': len(low_stock_products)
        })

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

def get_shop_email_settings(cursor, shop_id):
    """
    Load email settings for a specific shop (shop_id).
    Falls back to global settings (shop_id IS NULL) for any key not set per-shop.
    Returns a dict with keys: smtp_server, smtp_port, smtp_username, smtp_password, alert_email, low_stock_threshold.
    """
    keys = ['smtp_server', 'smtp_port', 'smtp_username', 'smtp_password', 'alert_email', 'low_stock_threshold']
    result = {}
    for key in keys:
        # Try shop-specific first
        if shop_id:
            row = cursor.execute(
                'SELECT value FROM settings WHERE `key` = ? AND shop_id = ?', (key, shop_id)
            ).fetchone()
            if row and row['value']:
                result[key] = row['value']
                continue
        # Fall back to global (shop_id IS NULL)
        row = cursor.execute(
            'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', (key,)
        ).fetchone()
        result[key] = row['value'] if row else ''
    return result

def send_email(settings, to_email, subject, html_body, from_name=None, text_body=None):
    """Send email using SMTP. `to_email` may be a single address, a comma/semicolon/space-separated string, or a list of addresses.
    `from_name` if provided will be used as the display name in the From header (e.g. "Shop Name <noreply@example.com>")."""
    try:
        # Normalize recipients to a list
        if isinstance(to_email, (list, tuple)):
            recipients = [str(x).strip() for x in to_email if str(x).strip()]
        else:
            # Split on commas, semicolons or whitespace
            recipients = [r.strip() for r in re.split('[,;\\s]+', str(to_email)) if r.strip()]

        if not recipients:
            raise Exception('No recipient email address provided')

        print(f"[EMAIL] Attempting to send email to: {recipients}")
        print(f"[EMAIL] SMTP Server: {settings['smtp_server']}:{settings['smtp_port']}")

        msg = MIMEMultipart('alternative')
        if from_name:
            try:
                msg['From'] = formataddr((str(from_name), settings['smtp_username']))
            except Exception:
                msg['From'] = f"{from_name} <{settings['smtp_username']}>"
        else:
            msg['From'] = settings['smtp_username']
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject

        if text_body is None:
            fallback_text = re.sub(r'(?i)<br\s*/?>', '\n', html_body)
            fallback_text = re.sub(r'(?i)</p>|</div>|</tr>|</h[1-6]>', '\n', fallback_text)
            fallback_text = re.sub(r'(?i)</td>|</th>', ' | ', fallback_text)
            fallback_text = re.sub(r'<[^>]+>', '', fallback_text)
            fallback_text = unescape(fallback_text)
            fallback_text = re.sub(r'\n\s*\n\s*\n+', '\n\n', fallback_text).strip()
            text_body = fallback_text

        text_part = MIMEText(text_body, 'plain')
        msg.attach(text_part)

        html_part = MIMEText(html_body, 'html')
        msg.attach(html_part)

        print(f"[EMAIL] Connecting to SMTP server...")
        port = int(settings['smtp_port'])

        if port == 465:
            server = smtplib.SMTP_SSL(settings['smtp_server'], port, timeout=10)
            print(f"[EMAIL] Connected via SSL (port 465)")
        else:
            server = smtplib.SMTP(settings['smtp_server'], port, timeout=10)
            print(f"[EMAIL] Starting TLS...")
            server.starttls()

        print(f"[EMAIL] Logging in as: {settings['smtp_username']}")
        server.login(settings['smtp_username'], settings['smtp_password'])
        print(f"[EMAIL] Sending message to: {recipients}")
        server.send_message(msg, to_addrs=recipients)
        print(f"[EMAIL] Closing connection...")
        server.quit()
        print(f"[EMAIL] Email sent successfully!")

        return True
    except smtplib.SMTPAuthenticationError as e:
        print(f"[EMAIL ERROR] Authentication failed: {str(e)}")
        raise Exception(f'SMTP Authentication failed. Check your username and password. For Gmail, use an App Password.')
    except smtplib.SMTPException as e:
        print(f"[EMAIL ERROR] SMTP error: {str(e)}")
        raise Exception(f'SMTP error: {str(e)}')
    except Exception as e:
        print(f"[EMAIL ERROR] Unexpected error: {str(e)}")
        raise Exception(f'Failed to send email: {str(e)}')

# Export endpoints
@app.route('/api/reports/sales/export/<format>', methods=['GET'])
@login_required
def export_sales_report(format):
    """Export sales report to Excel."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400

    if format != 'excel':
        return jsonify({'success': False, 'error': 'Only Excel export is supported'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    is_admin = session.get('role') in ('admin', 'super_admin', 'shop_owner')
    flt_sql, flt_params = shop_filter('s')

    try:
        # Get shop name and currency (from current shop)
        shop_id = get_current_shop_id() or session.get('shop_id')
        shop = cursor.execute('SELECT name, currency_symbol FROM shops WHERE id = ?', (shop_id,)).fetchone()
        shop_name = shop['name'] if shop else 'POS System'
        currency = shop['currency_symbol'] if shop else '$'

        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # Get summary statistics
        summary_query = f'''
            SELECT
                COUNT(*) as transaction_count,
                SUM(total_amount) as total_sales,
                AVG(total_amount) as avg_transaction
        '''
        if is_admin:
            summary_query += ', SUM(total_amount - COALESCE(total_cost, 0)) as total_profit'

        summary_query += f'''
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql}
        '''

        summary = cursor.execute(summary_query, [start_date_time, end_date_time] + flt_params).fetchone()

        # Get daily sales
        daily_query = f'''
            SELECT
                DATE(sale_date) as sale_date,
                COUNT(*) as transaction_count,
                SUM(total_amount) as daily_sales
        '''
        if is_admin:
            daily_query += ', SUM(total_amount - COALESCE(total_cost, 0)) as daily_profit'

        daily_query += f'''
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql}
            GROUP BY DATE(sale_date)
            ORDER BY sale_date ASC
        '''

        daily_sales = cursor.execute(daily_query, [start_date_time, end_date_time] + flt_params).fetchall()

        # Get product stats
        sales_data = cursor.execute(f'''
            SELECT id, items_json, total_amount, total_cost
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql}
        ''', [start_date_time, end_date_time] + flt_params).fetchall()

        product_stats = {}
        for sale in sales_data:
            items = json.loads(sale['items_json'])
            for item in items:
                product_id = item['product_id']
                if product_id not in product_stats:
                    product_stats[product_id] = {
                        'name': item['name'],
                        'sku': item.get('sku', 'N/A'),
                        'size': item.get('size') or '-',
                        'color': item.get('color') or '-',
                        'quantity': 0,
                        'revenue': 0,
                        'cost': 0
                    }

                product_stats[product_id]['quantity'] += item['quantity']
                product_stats[product_id]['revenue'] += item['quantity'] * item['price']

                if is_admin and 'cost_price' in item:
                    product_stats[product_id]['cost'] += item['quantity'] * item['cost_price']

        top_products = sorted(product_stats.values(), key=lambda x: x['quantity'], reverse=True)[:10]

        conn.close()

        return generate_sales_excel(shop_name, currency, start_date, end_date, summary, daily_sales, top_products, is_admin)

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/inventory/export/<format>', methods=['GET'])
@login_required
def export_inventory_report(format):
    """Export inventory report to Excel."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400

    if format != 'excel':
        return jsonify({'success': False, 'error': 'Only Excel export is supported'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    is_admin = session.get('role') in ('admin', 'super_admin', 'shop_owner')
    flt_sql, flt_params = shop_filter()
    flt_sql_sh, flt_params_sh = shop_filter('sh')
    flt_sql_s, flt_params_s = shop_filter('s')

    try:
        # Get shop name and currency (from current shop)
        shop_id = get_current_shop_id() or session.get('shop_id')
        shop = cursor.execute('SELECT name, currency_symbol FROM shops WHERE id = ?', (shop_id,)).fetchone()
        shop_name = shop['name'] if shop else 'POS System'
        currency = shop['currency_symbol'] if shop else '$'

        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # Current inventory
        if is_admin:
            inventory = cursor.execute(f'''
                SELECT name, sku, size, color, category, stock_quantity, cost_price, price,
                       (stock_quantity * cost_price) as stock_value,
                       (stock_quantity * price) as potential_revenue
                FROM products
                WHERE is_active = 1 {flt_sql}
                ORDER BY stock_value DESC
            ''', flt_params).fetchall()
        else:
            inventory = cursor.execute(f'''
                SELECT name, sku, size, color, category, stock_quantity, price
                FROM products
                WHERE is_active = 1 {flt_sql}
                ORDER BY stock_quantity DESC
            ''', flt_params).fetchall()

        # Stock movements
        stock_movements = cursor.execute(f'''
            SELECT
                DATE(sh.created_at) as date,
                sh.action_type,
                COUNT(*) as transaction_count,
                SUM(CASE WHEN sh.quantity_change > 0 THEN sh.quantity_change ELSE 0 END) as items_added,
                SUM(CASE WHEN sh.quantity_change < 0 THEN ABS(sh.quantity_change) ELSE 0 END) as items_removed
            FROM stock_history sh
            WHERE sh.created_at BETWEEN ? AND ? {flt_sql_sh}
            GROUP BY DATE(sh.created_at), sh.action_type
            ORDER BY date DESC
        ''', [start_date_time, end_date_time] + flt_params_sh).fetchall()

        # Top active products in period
        top_movements = cursor.execute(f'''
            SELECT
                p.name, p.sku, p.size, p.color,
                COUNT(*) as activity_count,
                SUM(CASE WHEN sh.quantity_change > 0 THEN sh.quantity_change ELSE 0 END) as total_added,
                SUM(CASE WHEN sh.quantity_change < 0 THEN ABS(sh.quantity_change) ELSE 0 END) as total_removed
            FROM stock_history sh
            JOIN products p ON sh.product_id = p.id
            WHERE sh.created_at BETWEEN ? AND ? {flt_sql_sh}
            GROUP BY sh.product_id, p.name, p.sku, p.size, p.color
            ORDER BY activity_count DESC
            LIMIT 10
        ''', [start_date_time, end_date_time] + flt_params_sh).fetchall()

        # Low stock
        low_stock = cursor.execute(f'''
            SELECT name, sku, size, color, stock_quantity, category
            FROM products
            WHERE is_active = 1 AND stock_quantity < 5 {flt_sql}
            ORDER BY stock_quantity ASC
        ''', flt_params).fetchall()

        # Slow-moving stock (0–1 sales in period)
        all_products_list = cursor.execute(f'''
            SELECT id, name, sku, size, color, category, stock_quantity, price
            FROM products
            WHERE is_active = 1 AND stock_quantity > 0 {flt_sql}
            ORDER BY name
        ''', flt_params).fetchall()

        sales_in_period = cursor.execute(f'''
            SELECT items_json FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
        ''', [start_date_time, end_date_time] + flt_params_s).fetchall()

        sold_qty = {}
        for sale_row in sales_in_period:
            items = json.loads(sale_row['items_json'])
            for item in items:
                pid = item['product_id']
                sold_qty[pid] = sold_qty.get(pid, 0) + item['quantity']

        slow_moving = []
        for p in all_products_list:
            qty_sold = sold_qty.get(p['id'], 0)
            if qty_sold <= 1:
                slow_moving.append({
                    'name': p['name'], 'sku': p['sku'] or '-',
                    'size': p['size'] or '-', 'color': p['color'] or '-',
                    'category': p['category'], 'stock_quantity': p['stock_quantity'],
                    'quantity_sold': qty_sold,
                    'stock_value': p['stock_quantity'] * p['price']
                })
        slow_moving.sort(key=lambda x: x['quantity_sold'])

        conn.close()

        return generate_inventory_excel(shop_name, currency, start_date, end_date, inventory, stock_movements, top_movements, low_stock, slow_moving, is_admin)

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/cashflow/export/<format>', methods=['GET'])
@login_required
@admin_required
def export_cashflow_report(format):
    """Export cash flow report to Excel."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400

    if format != 'excel':
        return jsonify({'success': False, 'error': 'Only Excel export is supported'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql_s, flt_params_s = shop_filter('s')
    flt_sql_p, flt_params_p = shop_filter('p')
    flt_sql_i, flt_params_i = shop_filter('i')
    flt_sql_c, flt_params_c = shop_filter('c')
    flt_sql_sr, flt_params_sr = shop_filter('sr')

    try:
        shop_id = get_current_shop_id() or session.get('shop_id')
        shop = cursor.execute('SELECT name, currency_symbol FROM shops WHERE id = ?', (shop_id,)).fetchone()
        shop_name = shop['name'] if shop else 'POS System'
        currency = shop['currency_symbol'] if shop else '$'

        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # Collect data
        walkin_cash = cursor.execute(f'''
            SELECT COALESCE(SUM(s.total_amount), 0) as total
            FROM sales s WHERE s.sale_date BETWEEN ? AND ? {flt_sql_s}
            AND s.id NOT IN (SELECT sale_id FROM invoices)
        ''', [start_date_time, end_date_time] + flt_params_s).fetchone()['total']

        payments_collected = cursor.execute(f'''
            SELECT COALESCE(SUM(amount), 0) as total
            FROM payments p WHERE created_at BETWEEN ? AND ? {flt_sql_p}
        ''', [start_date_time, end_date_time] + flt_params_p).fetchone()['total']

        total_refunds = cursor.execute(f'''
            SELECT COALESCE(SUM(sr.refund_amount), 0) as total
            FROM sale_returns sr WHERE return_date BETWEEN ? AND ? {flt_sql_sr}
        ''', [start_date_time, end_date_time] + flt_params_sr).fetchone()['total']

        credit_given = cursor.execute(f'''
            SELECT COALESCE(SUM(total_amount - paid_amount), 0) as total
            FROM invoices i WHERE created_at BETWEEN ? AND ? {flt_sql_i}
        ''', [start_date_time, end_date_time] + flt_params_i).fetchone()['total']

        total_outstanding = cursor.execute(f'''
            SELECT COALESCE(SUM(balance), 0) as total
            FROM customers c WHERE balance > 0 {flt_sql_c}
        ''', flt_params_c).fetchone()['total']

        customer_outstanding = cursor.execute(f'''
            SELECT c.name, c.phone, c.balance, c.credit_limit
            FROM customers c WHERE c.balance > 0 {flt_sql_c}
            ORDER BY c.balance DESC
        ''', flt_params_c).fetchall()

        conn.close()

        total_collected = walkin_cash + payments_collected - total_refunds
        summary = {
            'total_collected': total_collected,
            'walkin_cash': walkin_cash,
            'payments_collected': payments_collected,
            'total_refunds': total_refunds,
            'credit_given': credit_given,
            'total_outstanding': total_outstanding,
            'net_cash_flow': total_collected - credit_given
        }

        return generate_cashflow_excel(shop_name, currency, start_date, end_date, summary, customer_outstanding)

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_cashflow_pdf(shop_name, currency, start_date, end_date, summary, customer_outstanding):
    """Generate PDF for cash flow report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=20,
        textColor=colors.HexColor('#2c3e50'), spaceAfter=30, alignment=1
    )
    elements.append(Paragraph(f"{shop_name}", title_style))
    elements.append(Paragraph("Cash Flow Report", styles['Heading2']))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))

    # Summary
    elements.append(Paragraph("Cash Flow Summary", styles['Heading3']))
    summary_data = [
        ['Metric', 'Amount'],
        ['Cash Collected (Walk-in)', f"{currency}{summary['walkin_cash']:.2f}"],
        ['Payments Received (Credit)', f"{currency}{summary['payments_collected']:.2f}"],
        ['Less: Refunds', f"({currency}{summary['total_refunds']:.2f})"],
        ['Net Cash In', f"{currency}{summary['total_collected']:.2f}"],
        ['Credit Given', f"{currency}{summary['credit_given']:.2f}"],
        ['Net Cash Flow', f"{currency}{summary['net_cash_flow']:.2f}"],
        ['Total Outstanding', f"{currency}{summary['total_outstanding']:.2f}"],
    ]
    t = Table(summary_data, colWidths=[3*inch, 2.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.3*inch))

    # Customer outstanding
    if customer_outstanding:
        elements.append(Paragraph("Outstanding by Customer", styles['Heading3']))
        cust_data = [['Customer', 'Phone', 'Credit Limit', 'Outstanding']]
        for c in customer_outstanding:
            cust_data.append([
                c['name'], c['phone'] or '-',
                f"{currency}{c['credit_limit']:.2f}", f"{currency}{c['balance']:.2f}"
            ])
        ct = Table(cust_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        ct.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(ct)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True,
                     download_name=f'cashflow_report_{start_date}_to_{end_date}.pdf')

def generate_cashflow_excel(shop_name, currency, start_date, end_date, summary, customer_outstanding):
    """Generate Excel for cash flow report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Report"

    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws['A1'] = shop_name
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "Cash Flow Report"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A3'] = f"Period: {start_date} to {end_date}"

    row = 5
    ws[f'A{row}'] = "Cash Flow Summary"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for col_num, header in enumerate(['Metric', 'Amount'], 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    row += 1

    metrics = [
        ('Cash Collected (Walk-in)', summary['walkin_cash']),
        ('Payments Received (Credit)', summary['payments_collected']),
        ('Less: Refunds', -summary['total_refunds']),
        ('Net Cash In', summary['total_collected']),
        ('Credit Given', summary['credit_given']),
        ('Net Cash Flow', summary['net_cash_flow']),
        ('Total Outstanding', summary['total_outstanding']),
    ]
    for label, val in metrics:
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=2, value=f"{currency}{val:.2f}").border = border
        row += 1

    if customer_outstanding:
        row += 2
        ws[f'A{row}'] = "Outstanding by Customer"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        cust_headers = ['Customer', 'Phone', 'Credit Limit', 'Outstanding']
        for col_num, header in enumerate(cust_headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = border
        row += 1

        for c in customer_outstanding:
            ws.cell(row=row, column=1, value=c['name']).border = border
            ws.cell(row=row, column=2, value=c['phone'] or '-').border = border
            ws.cell(row=row, column=3, value=f"{currency}{c['credit_limit']:.2f}").border = border
            ws.cell(row=row, column=4, value=f"{currency}{c['balance']:.2f}").border = border
            row += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'cashflow_report_{start_date}_to_{end_date}.xlsx')

def generate_sales_pdf(shop_name, currency, start_date, end_date, summary, daily_sales, top_products, is_admin):
    """Generate PDF for sales report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"{shop_name}", title_style))
    elements.append(Paragraph(f"Sales Report", styles['Heading2']))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))

    # Summary
    elements.append(Paragraph("Summary", styles['Heading3']))
    summary_data = [
        ['Metric', 'Value'],
        ['Total Transactions', str(summary['transaction_count'] or 0)],
        ['Total Sales', f"{currency}{summary['total_sales'] or 0:.2f}"],
        ['Average Transaction', f"{currency}{summary['avg_transaction'] or 0:.2f}"]
    ]
    if is_admin and summary['total_profit'] is not None:
        summary_data.append(['Total Profit', f"{currency}{summary['total_profit']:.2f}"])

    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))

    # Top Products
    if top_products:
        elements.append(Paragraph("Top Selling Products", styles['Heading3']))
        product_headers = ['Product', 'SKU', 'Quantity', 'Revenue']
        if is_admin:
            product_headers.append('Profit')

        product_data = [product_headers]
        for p in top_products:
            row = [p['name'], p['sku'], str(p['quantity']), f"{currency}{p['revenue']:.2f}"]
            if is_admin:
                profit = p['revenue'] - p['cost']
                row.append(f"{currency}{profit:.2f}")
            product_data.append(row)

        col_widths = [2*inch, 1.5*inch, 1*inch, 1.2*inch]
        if is_admin:
            col_widths.append(1.2*inch)

        product_table = Table(product_data, colWidths=col_widths)
        product_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(product_table)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'sales_report_{start_date}_to_{end_date}.pdf')

def generate_sales_excel(shop_name, currency, start_date, end_date, summary, daily_sales, top_products, is_admin):
    """Generate Excel for sales report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    def _fmt_variant(value):
        """Normalize variant values so size/color can be text or numeric."""
        if value is None:
            return '-'
        text = str(value).strip()
        return text if text else '-'

    # Header styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = shop_name
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "Sales Report"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A3'] = f"Period: {start_date} to {end_date}"

    row = 5
    # Summary
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Value"
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].fill = header_fill
    ws[f'B{row}'].font = header_font
    row += 1

    ws[f'A{row}'] = "Total Transactions"
    ws[f'B{row}'] = summary['transaction_count'] or 0
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

    ws[f'A{row}'] = "Total Sales"
    ws[f'B{row}'] = f"{currency}{summary['total_sales'] or 0:.2f}"
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

    ws[f'A{row}'] = "Average Transaction"
    ws[f'B{row}'] = f"{currency}{summary['avg_transaction'] or 0:.2f}"
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

    if is_admin and summary['total_profit'] is not None:
        ws[f'A{row}'] = "Total Profit"
        ws[f'B{row}'] = f"{currency}{summary['total_profit']:.2f}"
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1

    # Top Products
    if top_products:
        row += 2
        ws[f'A{row}'] = "Top Selling Products"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        headers = ['Product', 'SKU', 'Size', 'Color', 'Quantity', 'Revenue']
        if is_admin:
            headers.append('Profit')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = border
        row += 1

        for product in top_products:
            ws.cell(row=row, column=1, value=product['name']).border = border
            ws.cell(row=row, column=2, value=product['sku']).border = border
            ws.cell(row=row, column=3, value=_fmt_variant(product.get('size'))).border = border
            ws.cell(row=row, column=4, value=_fmt_variant(product.get('color'))).border = border
            ws.cell(row=row, column=5, value=product['quantity']).border = border
            ws.cell(row=row, column=6, value=f"{currency}{product['revenue']:.2f}").border = border
            if is_admin:
                profit = product['revenue'] - product['cost']
                ws.cell(row=row, column=7, value=f"{currency}{profit:.2f}").border = border
            row += 1

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'sales_report_{start_date}_to_{end_date}.xlsx')

def generate_inventory_pdf(shop_name, currency, start_date, end_date, inventory, stock_movements, top_movements, low_stock, slow_moving, is_admin):
    """Generate PDF for inventory report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=20,
                                 textColor=colors.HexColor('#2c3e50'), spaceAfter=6, alignment=1)
    heading_style = ParagraphStyle('SectionHeading', parent=styles['Heading3'], fontSize=12,
                                   textColor=colors.HexColor('#2c3e50'), spaceBefore=14, spaceAfter=6)
    normal_style = styles['Normal']

    def make_table(data, col_widths, header_color='#3498db'):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_color)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ]))
        return t

    # Title
    elements.append(Paragraph(shop_name, title_style))
    elements.append(Paragraph("Inventory Report", styles['Heading2']))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", normal_style))
    elements.append(Spacer(1, 0.3*inch))

    # --- Summary Stats ---
    total_products = len(inventory)
    total_stock = sum(item['stock_quantity'] for item in inventory)
    elements.append(Paragraph("Inventory Summary", heading_style))
    summary_rows = [['Total Products', 'Total Stock Units', 'Low Stock Items']]
    summary_row = [str(total_products), str(total_stock), str(len(low_stock))]
    if is_admin:
        total_value = sum(item['stock_value'] for item in inventory)
        pot_revenue = sum(item['potential_revenue'] for item in inventory)
        summary_rows[0] += ['Stock Value', 'Potential Revenue']
        summary_row += [f"{currency}{total_value:,.2f}", f"{currency}{pot_revenue:,.2f}"]
    summary_rows.append(summary_row)
    col_w = [1.4*inch] * len(summary_rows[0])
    elements.append(make_table(summary_rows, col_w))
    elements.append(Spacer(1, 0.2*inch))

    # --- Current Inventory Table ---
    elements.append(Paragraph("Current Inventory Status", heading_style))
    if is_admin:
        inv_header = ['Product', 'SKU', 'Category', 'Stock', 'Cost Price', 'Sell Price', 'Stock Value']
        inv_col_w = [2*inch, 1.1*inch, 1*inch, 0.6*inch, 0.9*inch, 0.9*inch, 0.9*inch]
    else:
        inv_header = ['Product', 'SKU', 'Category', 'Stock', 'Sell Price']
        inv_col_w = [2.5*inch, 1.3*inch, 1.3*inch, 0.8*inch, 1*inch]
    inv_data = [inv_header]
    for item in inventory:
        if is_admin:
            inv_data.append([
                item['name'], item['sku'] or '-', item['category'],
                str(item['stock_quantity']),
                f"{currency}{item['cost_price']:,.2f}",
                f"{currency}{item['price']:,.2f}",
                f"{currency}{item['stock_value']:,.2f}"
            ])
        else:
            inv_data.append([
                item['name'], item['sku'] or '-', item['category'],
                str(item['stock_quantity']),
                f"{currency}{item['price']:,.2f}"
            ])
    elements.append(make_table(inv_data, inv_col_w))
    elements.append(Spacer(1, 0.2*inch))

    # --- Stock Movements ---
    elements.append(Paragraph(f"Stock Movements ({start_date} to {end_date})", heading_style))
    if stock_movements:
        sm_data = [['Date', 'Action Type', 'Transactions', 'Items Added', 'Items Removed']]
        for row in stock_movements:
            sm_data.append([
                str(row['date']),
                str(row['action_type']).capitalize(),
                str(row['transaction_count']),
                str(int(row['items_added'])),
                str(int(row['items_removed']))
            ])
        elements.append(make_table(sm_data, [1.3*inch, 1.3*inch, 1.2*inch, 1.2*inch, 1.4*inch]))
    else:
        elements.append(Paragraph("No stock movements in this period.", normal_style))
    elements.append(Spacer(1, 0.2*inch))

    # --- Most Active Products ---
    elements.append(Paragraph("Most Active Products", heading_style))
    if top_movements:
        tm_data = [['Product', 'SKU', 'Activities', 'Total Added', 'Total Removed']]
        for row in top_movements:
            tm_data.append([
                str(row['name']), str(row['sku'] or '-'),
                str(row['activity_count']),
                str(int(row['total_added'])),
                str(int(row['total_removed']))
            ])
        elements.append(make_table(tm_data, [2.2*inch, 1.2*inch, 1*inch, 1.1*inch, 1.2*inch]))
    else:
        elements.append(Paragraph("No product activity in this period.", normal_style))
    elements.append(Spacer(1, 0.2*inch))

    # --- Slow-Moving Stock ---
    elements.append(Paragraph("Slow-Moving Stock (0–1 sales in period)", heading_style))
    if slow_moving:
        sml_data = [['Product', 'SKU', 'Category', 'Stock', 'Qty Sold', 'Stock Value']]
        for item in slow_moving:
            sml_data.append([
                item['name'], item['sku'], item['category'],
                str(item['stock_quantity']),
                str(item['quantity_sold']),
                f"{currency}{item['stock_value']:,.2f}"
            ])
        elements.append(make_table(sml_data, [2*inch, 1*inch, 1*inch, 0.7*inch, 0.8*inch, 1.2*inch], header_color='#e67e22'))
    else:
        elements.append(Paragraph("No slow-moving stock in this period.", normal_style))
    elements.append(Spacer(1, 0.2*inch))

    # --- Low Stock Alerts ---
    elements.append(Paragraph("Low Stock Alerts (< 5 units)", heading_style))
    if low_stock:
        ls_data = [['Product', 'SKU', 'Category', 'Stock']]
        for item in low_stock:
            ls_data.append([item['name'], item['sku'] or '-', item['category'], str(item['stock_quantity'])])
        elements.append(make_table(ls_data, [2.5*inch, 1.3*inch, 1.5*inch, 0.9*inch], header_color='#e74c3c'))
    else:
        elements.append(Paragraph("No low stock items.", normal_style))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True,
                     download_name=f'inventory_report_{start_date}_to_{end_date}.pdf')

def generate_inventory_excel(shop_name, currency, start_date, end_date, inventory, stock_movements, top_movements, low_stock, slow_moving, is_admin):
    """Generate Excel for inventory report"""
    wb = Workbook()

    def _fmt_variant(value):
        """Normalize variant values so size/color can be text or numeric."""
        if value is None:
            return '-'
        text = str(value).strip()
        return text if text else '-'

    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    red_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
    orange_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alt_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")

    def write_section(ws, row, title, headers, rows_data, fill=None, title_color=None):
        """Write a titled section with header row and data rows; returns next row number."""
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.font = Font(bold=True, size=12, color=title_color or "2c3e50")
        row += 1
        for col_num, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_num, value=h)
            c.fill = fill or header_fill
            c.font = header_font
            c.border = border
        row += 1
        for i, data_row in enumerate(rows_data):
            for col_num, val in enumerate(data_row, 1):
                c = ws.cell(row=row, column=col_num, value=val)
                c.border = border
                if i % 2 == 1:
                    c.fill = alt_fill
            row += 1
        return row + 1

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws['A1'] = shop_name
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "Inventory Report"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A3'] = f"Period: {start_date} to {end_date}"
    row = 5

    total_products = len(inventory)
    total_stock = sum(item['stock_quantity'] for item in inventory)
    ws.cell(row=row, column=1, value="Total Products").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_products)
    row += 1
    ws.cell(row=row, column=1, value="Total Stock Units").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_stock)
    row += 1
    ws.cell(row=row, column=1, value="Low Stock Items").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(low_stock))
    row += 1
    if is_admin:
        total_value = sum(item['stock_value'] for item in inventory)
        pot_revenue = sum(item['potential_revenue'] for item in inventory)
        ws.cell(row=row, column=1, value="Stock Value").font = Font(bold=True)
        ws.cell(row=row, column=2, value=round(total_value, 2))
        row += 1
        ws.cell(row=row, column=1, value="Potential Revenue").font = Font(bold=True)
        ws.cell(row=row, column=2, value=round(pot_revenue, 2))
        row += 1

    # ---- Sheet 2: Current Inventory ----
    ws2 = wb.create_sheet("Current Inventory")
    ws2['A1'] = "Current Inventory Status"
    ws2['A1'].font = Font(bold=True, size=14)
    row2 = 3
    if is_admin:
        inv_headers = ['Product', 'SKU', 'Size', 'Color', 'Category', 'Stock', 'Cost Price', 'Sell Price', 'Stock Value']
    else:
        inv_headers = ['Product', 'SKU', 'Size', 'Color', 'Category', 'Stock', 'Sell Price']
    inv_rows = []
    for item in inventory:
        if is_admin:
            inv_rows.append([
                item['name'], item['sku'] or '-', _fmt_variant(item.get('size')), _fmt_variant(item.get('color')),
                item['category'], item['stock_quantity'],
                round(item['cost_price'], 2), round(item['price'], 2), round(item['stock_value'], 2)
            ])
        else:
            inv_rows.append([
                item['name'], item['sku'] or '-', _fmt_variant(item.get('size')), _fmt_variant(item.get('color')),
                item['category'], item['stock_quantity'], round(item['price'], 2)
            ])
    write_section(ws2, row2, "Current Inventory", inv_headers, inv_rows)

    # ---- Sheet 3: Stock Movements ----
    ws3 = wb.create_sheet("Stock Movements")
    ws3['A1'] = f"Stock Movements: {start_date} to {end_date}"
    ws3['A1'].font = Font(bold=True, size=14)
    sm_headers = ['Date', 'Action Type', 'Transactions', 'Items Added', 'Items Removed']
    sm_rows = [[str(r['date']), str(r['action_type']).capitalize(),
                r['transaction_count'], int(r['items_added']), int(r['items_removed'])]
               for r in stock_movements] if stock_movements else [['No data', '', '', '', '']]
    write_section(ws3, 3, "Stock Movements", sm_headers, sm_rows)

    # ---- Sheet 4: Most Active Products ----
    ws4 = wb.create_sheet("Most Active Products")
    ws4['A1'] = f"Most Active Products: {start_date} to {end_date}"
    ws4['A1'].font = Font(bold=True, size=14)
    tm_headers = ['Product', 'SKU', 'Size', 'Color', 'Activities', 'Total Added', 'Total Removed']
    tm_rows = [[
        r['name'], r['sku'] or '-', _fmt_variant(r.get('size')), _fmt_variant(r.get('color')),
        r['activity_count'], int(r['total_added']), int(r['total_removed'])
    ] for r in top_movements] if top_movements else [['No data', '', '', '', '', '', '']]
    write_section(ws4, 3, "Most Active Products", tm_headers, tm_rows)

    # ---- Sheet 5: Slow-Moving Stock ----
    ws5 = wb.create_sheet("Slow-Moving Stock")
    ws5['A1'] = "Slow-Moving Stock (0–1 sales in period)"
    ws5['A1'].font = Font(bold=True, size=14)
    sml_headers = ['Product', 'SKU', 'Size', 'Color', 'Category', 'Stock', 'Qty Sold', 'Stock Value']
    sml_rows = [[
        i['name'], i['sku'], _fmt_variant(i.get('size')), _fmt_variant(i.get('color')),
        i['category'], i['stock_quantity'], i['quantity_sold'], round(i['stock_value'], 2)
    ] for i in slow_moving] if slow_moving else [['No data', '', '', '', '', '', '', '']]
    write_section(ws5, 3, "Slow-Moving Stock", sml_headers, sml_rows, fill=orange_fill)

    # ---- Sheet 6: Low Stock Alerts ----
    ws6 = wb.create_sheet("Low Stock Alerts")
    ws6['A1'] = "Low Stock Alerts (< 5 units)"
    ws6['A1'].font = Font(bold=True, size=14)
    ls_headers = ['Product', 'SKU', 'Size', 'Color', 'Category', 'Stock']
    ls_rows = [[
        i['name'], i['sku'] or '-', _fmt_variant(i.get('size')), _fmt_variant(i.get('color')), i['category'], i['stock_quantity']
    ] for i in low_stock] if low_stock else [['No low stock items', '', '', '', '', '']]
    write_section(ws6, 3, "Low Stock Alerts", ls_headers, ls_rows, fill=red_fill)

    # Auto-size all sheets
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            sheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'inventory_report_{start_date}_to_{end_date}.xlsx')

@app.route('/api/scheduler/status', methods=['GET'])
@admin_required
def get_scheduler_status():
    """Get status of scheduled jobs - reads from database settings for consistency across workers"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        shop_id = get_settings_shop_id()

        def _get(key):
            if shop_id is not None:
                row = cursor.execute(
                    'SELECT value FROM settings WHERE `key` = ? AND shop_id = ?', (key, shop_id)
                ).fetchone()
                if row and row['value']:
                    return row['value']
            row = cursor.execute(
                'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', (key,)
            ).fetchone()
            return row['value'] if row and row['value'] else None

        times_val = _get('scheduler_times')
        tz_val = _get('scheduler_timezone')
        conn.close()

        times = json.loads(times_val) if times_val else ["09:00", "12:00", "18:00"]
        timezone = tz_val if tz_val else 'UTC'

        # Build job info from settings (consistent across all workers)
        import pytz
        from datetime import datetime as dt
        tz = pytz.timezone(timezone)
        now = dt.now(tz)

        jobs = []
        for idx, time_str in enumerate(times):
            hour, minute = map(int, time_str.split(':'))
            next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                next_run = next_run.replace(day=next_run.day + 1) if next_run.month == now.month else next_run
                try:
                    from datetime import timedelta
                    next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0) + timedelta(days=1)
                except Exception:
                    pass
            jobs.append({
                'id': f'low_stock_check_{idx}',
                'name': f'Low Stock Check - {time_str}',
                'next_run': next_run.strftime('%Y-%m-%d %H:%M:%S'),
                'trigger': f'cron[hour={hour}, minute={minute}]'
            })

        return jsonify({
            'scheduler_running': True,
            'jobs': jobs,
            'timezone': timezone
        })
    except Exception as e:
        return jsonify({
            'scheduler_running': False,
            'jobs': [],
            'error': str(e)
        })

@app.route('/api/scheduler/trigger-now', methods=['POST'])
@admin_required
def trigger_low_stock_now():
    """Manually trigger low stock check immediately (for testing)"""
    try:
        scheduled_low_stock_check()
        return jsonify({'success': True, 'message': 'Low stock check triggered successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/scheduler', methods=['GET', 'PUT'])
@admin_required
def manage_scheduler_settings():
    """Manage scheduler times and timezone for low stock alerts"""
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'GET':
        shop_id = get_settings_shop_id()

        def _get_setting(key):
            # Try shop-specific first, fall back to global
            if shop_id is not None:
                row = cursor.execute(
                    'SELECT value FROM settings WHERE `key` = ? AND shop_id = ?', (key, shop_id)
                ).fetchone()
                if row and row['value']:
                    return row['value']
            row = cursor.execute(
                'SELECT value FROM settings WHERE `key` = ? AND shop_id IS NULL', (key,)
            ).fetchone()
            return row['value'] if row and row['value'] else None

        times_val = _get_setting('scheduler_times')
        tz_val = _get_setting('scheduler_timezone')
        times = json.loads(times_val) if times_val else ["09:00", "12:00", "18:00"]
        timezone = tz_val if tz_val else 'UTC'

        conn.close()
        return jsonify({'times': times, 'timezone': timezone})

    # PUT - Update scheduler times and timezone
    data = request.json
    times = data.get('times', [])
    timezone = data.get('timezone', 'UTC')

    # Validate times
    if not times or not isinstance(times, list):
        conn.close()
        return jsonify({'success': False, 'error': 'Times must be a non-empty list'}), 400

    # Validate each time format (HH:MM)
    for time_str in times:
        try:
            hour, minute = map(int, time_str.split(':'))
            if hour < 0 or hour > 23 or minute < 0 or minute > 59:
                raise ValueError()
        except:
            conn.close()
            return jsonify({'success': False, 'error': f'Invalid time format: {time_str}. Use HH:MM (24-hour format)'}), 400

    # Validate timezone
    import pytz
    if timezone not in pytz.all_timezones:
        conn.close()
        return jsonify({'success': False, 'error': f'Invalid timezone: {timezone}'}), 400

    try:
        # Update times using DELETE + INSERT to prevent duplicate key errors
        shop_id = get_settings_shop_id()

        def _upsert_setting(key, value):
            # Always delete first (no-op if doesn't exist), then insert
            if shop_id is not None:
                cursor.execute("DELETE FROM settings WHERE `key` = ? AND shop_id = ?", (key, shop_id))
                cursor.execute(
                    "INSERT INTO settings (`key`, value, shop_id) VALUES (?, ?, ?)",
                    (key, value, shop_id)
                )
            else:
                cursor.execute("DELETE FROM settings WHERE `key` = ? AND shop_id IS NULL", (key,))
                cursor.execute(
                    "INSERT INTO settings (`key`, value, shop_id) VALUES (?, ?, NULL)",
                    (key, value)
                )

        _upsert_setting('scheduler_times', json.dumps(times))
        _upsert_setting('scheduler_timezone', timezone)

        conn.commit()
        conn.close()

        # Reload scheduler jobs with new times and timezone
        load_scheduler_jobs()

        return jsonify({'success': True, 'times': times, 'timezone': timezone})

    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/timezones', methods=['GET'])
@login_required
def list_timezones():
    """Return list of common timezones for the UI dropdown"""
    import pytz
    common = sorted(pytz.common_timezones)
    return jsonify({'timezones': common})

# --- Products: Export CSV ---
@app.route('/api/products/export-csv', methods=['GET'])
@login_required
def export_products_csv():
    import csv, io
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    products = conn.execute(
        f'SELECT * FROM products WHERE is_active = 1 {flt_sql} ORDER BY name', flt_params
    ).fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    is_admin = session.get('role') in ('admin', 'super_admin', 'shop_owner')

    if is_admin:
        writer.writerow(['SKU', 'Name', 'Category', 'Size', 'Color', 'Cost Price', 'Sell Price', 'Stock', 'Low Stock Threshold', 'Description'])
        for p in products:
            writer.writerow([p['sku'] or '', p['name'], p['category'], p['size'] or '', p['color'] or '',
                             p['cost_price'], p['price'], p['stock_quantity'],
                             p['low_stock_threshold'] or '', p['description'] or ''])
    else:
        writer.writerow(['SKU', 'Name', 'Category', 'Size', 'Color', 'Sell Price', 'Stock'])
        for p in products:
            writer.writerow([p['sku'] or '', p['name'], p['category'], p['size'] or '', p['color'] or '',
                             p['price'], p['stock_quantity']])

    csv_bytes = output.getvalue().encode('utf-8')
    return send_file(BytesIO(csv_bytes), mimetype='text/csv', as_attachment=True,
                     download_name=f'products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')

# --- Products: Import CSV ---
@app.route('/api/products/import-csv', methods=['POST'])
@login_required
@admin_required
def import_products_csv():
    import csv, io
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400

    try:
        content = file.stream.read().decode('utf-8-sig')  # utf-8-sig strips BOM
        reader = csv.DictReader(io.StringIO(content))

        # Normalize headers to handle case variations
        if reader.fieldnames:
            header_map = {}
            for h in reader.fieldnames:
                hl = h.strip().lower()
                header_map[hl] = h
        else:
            return jsonify({'success': False, 'error': 'CSV has no headers'}), 400

        def get_col(row, *names):
            for n in names:
                # Try exact match first
                if n in row and row[n]:
                    return row[n].strip()
                # Try case-insensitive match via header_map
                nl = n.strip().lower()
                if nl in header_map and header_map[nl] in row and row[header_map[nl]]:
                    return row[header_map[nl]].strip()
            return ''

        conn = get_db_connection()
        cursor = conn.cursor()
        flt_sql, flt_params = shop_filter()
        added = 0
        updated = 0
        errors = []

        for i, row in enumerate(reader, start=2):
            try:
                name = get_col(row, 'Name', 'name', 'Product Name', 'product_name')
                category = get_col(row, 'Category', 'category')
                price_str = get_col(row, 'Sell Price', 'sell_price', 'Price', 'price')
                cost_str = get_col(row, 'Cost Price', 'cost_price', 'Cost')
                stock_str = get_col(row, 'Stock', 'stock', 'Stock Quantity', 'stock_quantity', 'Quantity')
                sku = get_col(row, 'SKU', 'sku', 'Sku')
                description = get_col(row, 'Description', 'description')
                threshold_str = get_col(row, 'Low Stock Threshold', 'low_stock_threshold', 'Threshold')
                size = get_col(row, 'Size', 'size')
                color = get_col(row, 'Color', 'color')

                price = float(price_str) if price_str else 0
                cost_price = float(cost_str) if cost_str else 0
                stock = int(float(stock_str)) if stock_str else 0
                threshold = int(threshold_str) if threshold_str else None

                if not name or not category:
                    errors.append(f'Row {i}: Name and Category are required')
                    continue

                # Check if SKU exists — update if so, insert if not
                if sku:
                    existing = cursor.execute(
                        f'SELECT id FROM products WHERE sku=? {flt_sql}', [sku] + flt_params
                    ).fetchone()
                    if existing:
                        cursor.execute(f'''
                            UPDATE products SET name=?, category=?, cost_price=?, price=?,
                            stock_quantity=?, description=?, low_stock_threshold=?, size=?, color=? WHERE sku=? {flt_sql}
                        ''', [name, category, cost_price, price, stock, description, threshold,
                              size or None, color or None, sku] + flt_params)
                        updated += 1
                        continue

                if not sku:
                    shop_id = get_current_shop_id()
                    prefix = category[:3].upper()
                    existing_count = cursor.execute(
                        f"SELECT COUNT(*) FROM products WHERE sku LIKE ? {flt_sql}", [f'{prefix}-%'] + flt_params
                    ).fetchone()[0]
                    sku = f'{prefix}-{existing_count + 1:03d}'
                    while cursor.execute(
                        f"SELECT COUNT(*) FROM products WHERE sku=? {flt_sql}", [sku] + flt_params
                    ).fetchone()[0] > 0:
                        existing_count += 1
                        sku = f'{prefix}-{existing_count + 1:03d}'

                cursor.execute('''
                    INSERT INTO products (shop_id, name, category, cost_price, price, stock_quantity, sku, description, low_stock_threshold, size, color)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (get_current_shop_id(), name, category, cost_price, price, stock, sku, description, threshold,
                      size or None, color or None))
                added += 1

            except Exception as e:
                errors.append(f'Row {i}: {str(e)}')

        conn.commit()
        conn.close()

        detected = list(reader.fieldnames) if reader.fieldnames else []
        return jsonify({'success': True, 'added': added, 'updated': updated, 'errors': errors, 'detected_headers': detected})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

# --- Products: Duplicate ---
@app.route('/api/products/<int:product_id>/duplicate', methods=['POST'])
@login_required
@admin_required
def duplicate_product(product_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    flt_sql, flt_params = shop_filter()
    product = cursor.execute(
        f'SELECT * FROM products WHERE id=? {flt_sql}', [product_id] + flt_params
    ).fetchone()
    if not product:
        conn.close()
        return jsonify({'success': False, 'error': 'Product not found'}), 404

    # Generate new SKU within this shop
    shop_id = get_current_shop_id()
    prefix = product['category'][:3].upper()
    existing = cursor.execute(
        f"SELECT COUNT(*) FROM products WHERE sku LIKE ? {flt_sql}", [f'{prefix}-%'] + flt_params
    ).fetchone()[0]
    new_sku = f'{prefix}-{existing + 1:03d}'
    while cursor.execute(
        f"SELECT COUNT(*) FROM products WHERE sku=? {flt_sql}", [new_sku] + flt_params
    ).fetchone()[0] > 0:
        existing += 1
        new_sku = f'{prefix}-{existing + 1:03d}'

    try:
        cursor.execute('''
            INSERT INTO products (shop_id, name, category, cost_price, price, stock_quantity, sku, description, low_stock_threshold, size, color)
            VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?)
        ''', (shop_id, product['name'] + ' (Copy)', product['category'], product['cost_price'],
              product['price'], new_sku, product['description'], product['low_stock_threshold'],
              product['size'], product['color']))

        new_id = cursor.lastrowid

        cursor.execute('''
            INSERT INTO product_audit_log (shop_id, product_id, user_id, username, action, field_name, old_value, new_value)
            VALUES (?, ?, ?, ?, 'duplicate', 'source', ?, ?)
        ''', (shop_id, new_id, session.get('user_id'), session.get('username'), str(product_id), str(new_id)))

        conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


# --- Products: Create Variants ---
@app.route('/api/products/create-variants', methods=['POST'])
@login_required
@admin_required
def create_product_variants():
    """Bulk-create product variants for all size/color combinations."""
    data = request.json
    name = normalize_product_name(data.get('name', '').strip())
    category = data.get('category', '').strip()
    cost_price = float(data.get('cost_price', 0))
    price = float(data.get('price', 0))
    stock_per_variant = int(data.get('stock_quantity', 0))  # fallback uniform stock
    description = data.get('description', '').strip()
    sizes = data.get('sizes', [])
    colors = data.get('colors', [])
    stock_map = data.get('stock_map', [])  # [{size, color, stock}, ...]

    if not name or not category:
        return jsonify({'success': False, 'error': 'Name and Category are required'}), 400
    if not sizes or not colors:
        return jsonify({'success': False, 'error': 'At least one size and one color are required'}), 400

    # Build combinations. Bulk variants must always include both dimensions.
    combos = []
    for size in sizes:
        for color in colors:
            combos.append(((size or '').strip() or None, normalize_color_value(color)))

    if len(combos) > 100:
        return jsonify({'success': False, 'error': 'Too many variants (max 100)'}), 400

    # Normalize and de-duplicate incoming combinations from this request.
    # Reject if the same size/color pair is repeated in payload.
    normalized_combos = []
    seen_payload = set()
    duplicate_payload = []
    for size, color in combos:
        size_clean = (size or '').strip() or None
        color_clean = normalize_color_value(color)
        if not size_clean or not color_clean:
            return jsonify({
                'success': False,
                'error': 'Each variant must include both a size and a color'
            }), 400
        pair_key = (size_clean, color_clean)
        if pair_key in seen_payload:
            duplicate_payload.append(pair_key)
            continue
        seen_payload.add(pair_key)
        normalized_combos.append((size_clean, color_clean))

    if duplicate_payload:
        sample = ', '.join([f"size={d[0] or '-'}, color={d[1] or '-'}" for d in duplicate_payload[:5]])
        return jsonify({
            'success': False,
            'error': f'Duplicate size/color rows in request: {sample}'
        }), 400

    combos = normalized_combos

    # Build stock lookup from stock_map (matches by index)
    stock_values = []
    if stock_map and len(stock_map) == len(combos):
        stock_values = [max(0, int(entry.get('stock', 0))) for entry in stock_map]
    else:
        stock_values = [stock_per_variant] * len(combos)

    if not any(v > 0 for v in stock_values):
        return jsonify({
            'success': False,
            'error': 'All variant stock values are 0. Set stock for at least one variant.'
        }), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    shop_id = get_current_shop_id()
    flt_sql, flt_params = shop_filter()

    try:
        # Reuse canonical existing name/group when normalized names match.
        canonical_name = name
        canonical_category = category
        existing_variant_group = None
        siblings = cursor.execute(
            '''
            SELECT id, name, category, variant_group
            FROM products
            WHERE shop_id = ? AND is_active = 1
            ORDER BY id
            ''',
            (shop_id,)
        ).fetchall()
        for sibling in siblings:
            if normalize_product_name(sibling['name']) == name:
                canonical_name = sibling['name']
                if sibling['category']:
                    canonical_category = sibling['category']
                if sibling['variant_group']:
                    existing_variant_group = sibling['variant_group']
                break

        name = canonical_name
        category = canonical_category

        if existing_variant_group:
            variant_group = existing_variant_group
        else:
            import uuid
            variant_group = f"{name[:20].lower().replace(' ', '-')}-{uuid.uuid4().hex[:8]}"

        # Block creation only for variants that will actually be created (stock > 0).
        # Zero-stock rows are treated as skipped inputs.
        existing_conflicts = []
        for idx, (size, color) in enumerate(combos):
            if stock_values[idx] <= 0:
                continue
            existing = find_existing_variant_by_normalized_color(
                cursor,
                shop_id,
                name,
                size,
                color,
            )
            if existing:
                existing_conflicts.append((size, color))

        if existing_conflicts:
            sample = ', '.join([f"size={c[0] or '-'}, color={c[1] or '-'}" for c in existing_conflicts[:8]])
            return jsonify({
                'success': False,
                'error': f'Cannot create duplicate variants for "{name}". Already exists: {sample}'
            }), 400

        created_ids = []
        for idx, (size, color) in enumerate(combos):
            variant_stock = stock_values[idx]
            # Skip variants with 0 stock — don't create them
            if variant_stock <= 0:
                continue
            # Auto-generate SKU
            prefix = category[:3].upper()
            existing_count = cursor.execute(
                f"SELECT COUNT(*) FROM products WHERE sku LIKE ? {flt_sql}", [f'{prefix}-%'] + flt_params
            ).fetchone()[0]
            sku = f'{prefix}-{existing_count + 1:03d}'
            while cursor.execute(
                f"SELECT COUNT(*) FROM products WHERE sku=? {flt_sql}", [sku] + flt_params
            ).fetchone()[0] > 0:
                existing_count += 1
                sku = f'{prefix}-{existing_count + 1:03d}'

            cursor.execute('''
                INSERT INTO products (shop_id, name, category, cost_price, price, stock_quantity, sku, description, size, color, variant_group)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (shop_id, name, category, cost_price, price, variant_stock, sku, description,
                  size, color, variant_group))

            product_id = cursor.lastrowid
            created_ids.append(product_id)

            if variant_stock > 0:
                cursor.execute('''
                    INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                    VALUES (?, ?, ?, ?, ?)
                ''', (shop_id, product_id, variant_stock, 'initial', f'Initial stock | Variant: {size or ""} {color or ""}'))

        conn.commit()
        if not created_ids:
            return jsonify({
                'success': False,
                'error': 'No variants were created. Check stock values and duplicate combinations.'
            }), 400
        return jsonify({
            'success': True,
            'created': len(created_ids),
            'variant_group': variant_group,
            'ids': created_ids
        }), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


# --- Products: Create Zero-Stock Variant Blueprints ---
@app.route('/api/products/create-variant-blueprint', methods=['POST'])
@login_required
@admin_required
def create_variant_blueprint():
    """Create all size/color combinations as zero-stock variants for future PO planning."""
    data = request.json
    name = normalize_product_name(data.get('name', '').strip())
    category = data.get('category', '').strip()
    cost_price = float(data.get('cost_price', 0))
    price = float(data.get('price', 0))
    description = data.get('description', '').strip()
    sizes = data.get('sizes', [])
    colors = data.get('colors', [])
    selected_pairs = data.get('selected_pairs', [])

    if not name or not category:
        return jsonify({'success': False, 'error': 'Name and Category are required'}), 400

    combos = []
    if selected_pairs:
        # Explicit pair mode: create only the selected size/color rows.
        for pair in selected_pairs:
            combos.append(((pair.get('size') or '').strip() or None, normalize_color_value(pair.get('color'))))
    else:
        # Backward-compatible mode: full cartesian product from sizes and colors.
        if not sizes or not colors:
            return jsonify({'success': False, 'error': 'At least one size and one color are required'}), 400
        for size in sizes:
            for color in colors:
                combos.append(((size or '').strip() or None, normalize_color_value(color)))

    if len(combos) > 100:
        return jsonify({'success': False, 'error': 'Too many variants (max 100)'}), 400

    if not combos:
        return jsonify({'success': False, 'error': 'No variant combinations selected'}), 400

    normalized_combos = []
    seen_payload = set()
    duplicate_payload = []
    for size, color in combos:
        size_clean = (size or '').strip() or None
        color_clean = normalize_color_value(color)
        if not size_clean or not color_clean:
            return jsonify({
                'success': False,
                'error': 'Each variant must include both a size and a color'
            }), 400
        pair_key = (size_clean, color_clean)
        if pair_key in seen_payload:
            duplicate_payload.append(pair_key)
            continue
        seen_payload.add(pair_key)
        normalized_combos.append((size_clean, color_clean))

    if duplicate_payload:
        sample = ', '.join([f"size={d[0] or '-'}, color={d[1] or '-'}" for d in duplicate_payload[:5]])
        return jsonify({
            'success': False,
            'error': f'Duplicate size/color rows in request: {sample}'
        }), 400

    combos = normalized_combos

    conn = get_db_connection()
    cursor = conn.cursor()
    shop_id = get_current_shop_id()
    flt_sql, flt_params = shop_filter()

    try:
        canonical_name = name
        canonical_category = category
        existing_variant_group = None
        siblings = cursor.execute(
            '''
            SELECT id, name, category, variant_group
            FROM products
            WHERE shop_id = ? AND is_active = 1
            ORDER BY id
            ''',
            (shop_id,)
        ).fetchall()
        for sibling in siblings:
            if normalize_product_name(sibling['name']) == name:
                canonical_name = sibling['name']
                if sibling['category']:
                    canonical_category = sibling['category']
                if sibling['variant_group']:
                    existing_variant_group = sibling['variant_group']
                break

        name = canonical_name
        category = canonical_category

        if existing_variant_group:
            variant_group = existing_variant_group
        else:
            import uuid
            variant_group = f"{name[:20].lower().replace(' ', '-')}-{uuid.uuid4().hex[:8]}"

        existing_conflicts = []
        for size, color in combos:
            existing = find_existing_variant_by_normalized_color(
                cursor,
                shop_id,
                name,
                size,
                color,
            )
            if existing:
                existing_conflicts.append((size, color))

        if existing_conflicts:
            sample = ', '.join([f"size={c[0] or '-'}, color={c[1] or '-'}" for c in existing_conflicts[:8]])
            return jsonify({
                'success': False,
                'error': f'Cannot create duplicate variants for "{name}". Already exists: {sample}'
            }), 400

        created_ids = []
        for size, color in combos:
            prefix = category[:3].upper()
            existing_count = cursor.execute(
                f"SELECT COUNT(*) FROM products WHERE sku LIKE ? {flt_sql}", [f'{prefix}-%'] + flt_params
            ).fetchone()[0]
            sku = f'{prefix}-{existing_count + 1:03d}'
            while cursor.execute(
                f"SELECT COUNT(*) FROM products WHERE sku=? {flt_sql}", [sku] + flt_params
            ).fetchone()[0] > 0:
                existing_count += 1
                sku = f'{prefix}-{existing_count + 1:03d}'

            cursor.execute('''
                INSERT INTO products (shop_id, name, category, cost_price, price, stock_quantity, sku, description, size, color, variant_group)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (shop_id, name, category, cost_price, price, 0, sku, description,
                  size, color, variant_group))

            created_ids.append(cursor.lastrowid)

        conn.commit()
        return jsonify({
            'success': True,
            'created': len(created_ids),
            'variant_group': variant_group,
            'ids': created_ids,
            'zero_stock': True
        }), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


# --- Products: Bulk stock update ---
@app.route('/api/products/bulk-stock', methods=['POST'])
@login_required
@admin_required
def bulk_stock_update():
    data = request.json
    product_ids = data.get('product_ids', [])
    quantity_change = data.get('quantity_change', 0)
    action_type = data.get('action_type', 'import')
    note = data.get('note', '')

    if not product_ids or quantity_change == 0:
        return jsonify({'success': False, 'error': 'No products or zero quantity'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql, flt_params = shop_filter()

    try:
        updated = 0
        for pid in product_ids:
            product = cursor.execute(
                f'SELECT stock_quantity FROM products WHERE id=? {flt_sql}', [pid] + flt_params
            ).fetchone()
            if not product:
                continue
            new_stock = product['stock_quantity'] + quantity_change
            if new_stock < 0:
                continue  # skip products that would go negative
            cursor.execute(
                f'UPDATE products SET stock_quantity=? WHERE id=? {flt_sql}',
                [new_stock, pid] + flt_params
            )
            cursor.execute('''
                INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                VALUES (?, ?, ?, ?, ?)
            ''', (get_current_shop_id(), pid, quantity_change, action_type, note or f'Bulk {action_type} by {session.get("username")}'))
            updated += 1

        conn.commit()
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Products: Update reorder level ---
@app.route('/api/products/<int:product_id>/reorder-level', methods=['PUT'])
@login_required
@admin_required
def update_reorder_level(product_id):
    data = request.json
    threshold = data.get('low_stock_threshold')
    if threshold is not None and threshold != '':
        threshold = int(threshold)
    else:
        threshold = None

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql, flt_params = shop_filter()
    try:
        cursor.execute(
            f'UPDATE products SET low_stock_threshold=? WHERE id=? {flt_sql}',
            [threshold, product_id] + flt_params
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Products: Audit log ---
@app.route('/api/products/<int:product_id>/audit-log', methods=['GET'])
@login_required
@admin_required
def get_product_audit_log(product_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    logs = conn.execute(
        f'SELECT * FROM product_audit_log WHERE product_id=? {flt_sql} ORDER BY created_at DESC LIMIT 100',
        [product_id] + flt_params
    ).fetchall()
    conn.close()
    return jsonify([dict(row) for row in logs])

# --- Products: Inline quick-edit ---
@app.route('/api/products/<int:product_id>/quick-edit', methods=['PATCH'])
@login_required
@admin_required
def quick_edit_product(product_id):
    data = request.json
    field = data.get('field')
    value = data.get('value')

    allowed = {'name': str, 'price': float, 'cost_price': float, 'stock_quantity': int, 'category': str}
    if field not in allowed:
        return jsonify({'success': False, 'error': f'Field {field} not editable'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql, flt_params = shop_filter()

    try:
        old = cursor.execute(
            f'SELECT {field} FROM products WHERE id=? {flt_sql}', [product_id] + flt_params
        ).fetchone()
        if not old:
            return jsonify({'success': False, 'error': 'Product not found'}), 404

        old_val = old[field]
        new_val = allowed[field](value)

        # For stock changes, also log in stock_history
        if field == 'stock_quantity':
            diff = new_val - old_val
            if diff != 0:
                cursor.execute('''
                    INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                    VALUES (?, ?, ?, ?, ?)
                ''', (get_current_shop_id(), product_id, diff, 'import' if diff > 0 else 'export',
                      f'Inline edit by {session.get("username")}'))

        cursor.execute(f'UPDATE products SET {field}=? WHERE id=? {flt_sql}', [new_val, product_id] + flt_params)

        # Audit log
        cursor.execute('''
            INSERT INTO product_audit_log (shop_id, product_id, user_id, username, action, field_name, old_value, new_value)
            VALUES (?, ?, ?, ?, 'inline-edit', ?, ?, ?)
        ''', (get_current_shop_id(), product_id, session.get('user_id'), session.get('username'), field, str(old_val), str(new_val)))

        conn.commit()
        return jsonify({'success': True, 'new_value': new_val})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# ===== SUPPLIER & PURCHASE ORDER ROUTES =====

@app.route('/suppliers')
@login_required
@admin_required
def suppliers_page():
    return render_template('suppliers.html')

@app.route('/purchase-orders')
@login_required
@admin_required
def purchase_orders_page():
    return render_template('purchase_orders.html')

# --- Product-Supplier Links API ---

@app.route('/api/product-suppliers/<int:product_id>', methods=['GET'])
@login_required
@admin_required
def get_product_suppliers(product_id):
    """Get all suppliers linked to a product"""
    conn = get_db_connection()
    rows = conn.execute('''
        SELECT ps.*, s.name as supplier_name
        FROM product_suppliers ps
        JOIN suppliers s ON ps.supplier_id = s.id
        WHERE ps.product_id = ?
        ORDER BY s.name
    ''', (product_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/product-suppliers', methods=['POST'])
@login_required
@admin_required
def add_product_supplier():
    """Link a product to a supplier"""
    data = request.json
    product_id = data.get('product_id')
    supplier_id = data.get('supplier_id')
    if not product_id or not supplier_id:
        return jsonify({'success': False, 'error': 'product_id and supplier_id required'}), 400
    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO product_suppliers (product_id, supplier_id, supplier_cost, supplier_sku)
            VALUES (?, ?, ?, ?)
        ''', (product_id, supplier_id, data.get('supplier_cost'), data.get('supplier_sku')))
        conn.commit()
        return jsonify({'success': True}), 201
    except Exception as e:
        conn.rollback()
        if 'Duplicate' in str(e):
            return jsonify({'success': False, 'error': 'Already linked'}), 409
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/product-suppliers/<int:link_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_product_supplier(link_id):
    """Remove a product-supplier link"""
    conn = get_db_connection()
    conn.execute('DELETE FROM product_suppliers WHERE id=?', (link_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/suppliers/<int:supplier_id>/products', methods=['GET'])
@login_required
@admin_required
def get_supplier_products(supplier_id):
    """Get all products linked to a supplier (for PO creation)"""
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('p')
    rows = conn.execute(f'''
        SELECT p.*, ps.supplier_cost, ps.supplier_sku
        FROM products p
        JOIN product_suppliers ps ON ps.product_id = p.id
        WHERE ps.supplier_id = ? {flt_sql}
        ORDER BY p.name
    ''', [supplier_id] + flt_params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# --- Suppliers API ---

@app.route('/api/suppliers', methods=['GET'])
@login_required
@admin_required
def get_suppliers():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('s')
    suppliers = conn.execute(f'''
        SELECT s.*,
            COUNT(po.id) as order_count,
            COALESCE(SUM(po.total_amount), 0) as total_purchased,
            MAX(po.created_at) as last_order_date,
            SUM(CASE WHEN po.status IN ('draft','ordered','partial') THEN 1 ELSE 0 END) as pending_orders
        FROM suppliers s
        LEFT JOIN purchase_orders po ON po.supplier_id = s.id
        WHERE 1=1 {flt_sql}
        GROUP BY s.id
        ORDER BY s.name
    ''', flt_params).fetchall()
    conn.close()
    return jsonify([dict(s) for s in suppliers])

@app.route('/api/suppliers', methods=['POST'])
@login_required
@admin_required
def add_supplier():
    data = request.json
    if not data.get('name', '').strip():
        return jsonify({'success': False, 'error': 'Supplier name is required'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        shop_id = get_current_shop_id()
        cursor.execute('''
            INSERT INTO suppliers (shop_id, name, contact_person, email, phone, address, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (shop_id, data['name'].strip(), data.get('contact_person', '').strip(),
              data.get('email', '').strip(), data.get('phone', '').strip(),
              data.get('address', '').strip(), data.get('notes', '').strip()))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/suppliers/<int:supplier_id>', methods=['GET'])
@login_required
@admin_required
def get_supplier(supplier_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    supplier = conn.execute(
        f'SELECT * FROM suppliers WHERE id=? {flt_sql}', [supplier_id] + flt_params
    ).fetchone()
    conn.close()
    if not supplier:
        return jsonify({'error': 'Supplier not found'}), 404
    return jsonify(dict(supplier))

@app.route('/api/suppliers/<int:supplier_id>', methods=['PUT'])
@login_required
@admin_required
def update_supplier(supplier_id):
    data = request.json
    if not data.get('name', '').strip():
        return jsonify({'success': False, 'error': 'Supplier name is required'}), 400

    conn = get_db_connection()
    try:
        flt_sql, flt_params = shop_filter()
        status = data.get('status', '').strip()
        if status and status in ('active', 'inactive'):
            conn.execute(f'''
                UPDATE suppliers SET name=?, contact_person=?, email=?, phone=?, address=?, notes=?, status=?
                WHERE id=? {flt_sql}
            ''', [data['name'].strip(), data.get('contact_person', '').strip(),
                  data.get('email', '').strip(), data.get('phone', '').strip(),
                  data.get('address', '').strip(), data.get('notes', '').strip(), status, supplier_id] + flt_params)
        else:
            conn.execute(f'''
                UPDATE suppliers SET name=?, contact_person=?, email=?, phone=?, address=?, notes=?
                WHERE id=? {flt_sql}
            ''', [data['name'].strip(), data.get('contact_person', '').strip(),
                  data.get('email', '').strip(), data.get('phone', '').strip(),
                  data.get('address', '').strip(), data.get('notes', '').strip(), supplier_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/suppliers/<int:supplier_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_supplier(supplier_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    # Check if supplier has purchase orders
    po_count = conn.execute(
        f'SELECT COUNT(*) as c FROM purchase_orders WHERE supplier_id=? {flt_sql}',
        [supplier_id] + flt_params
    ).fetchone()['c']
    if po_count > 0:
        conn.close()
        return jsonify({'success': False, 'error': f'Cannot delete supplier with {po_count} purchase order(s). Delete orders first.'}), 400
    try:
        conn.execute(f'DELETE FROM suppliers WHERE id=? {flt_sql}', [supplier_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Purchase Orders API ---

@app.route('/api/purchase-orders', methods=['GET'])
@login_required
@admin_required
def get_purchase_orders():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('po')
    orders = conn.execute(f'''
        SELECT po.*, s.name as supplier_name, u.username as created_by_name
        FROM purchase_orders po
        LEFT JOIN suppliers s ON po.supplier_id = s.id
        LEFT JOIN users u ON po.created_by = u.id
        WHERE 1=1 {flt_sql}
        ORDER BY po.created_at DESC
    ''', flt_params).fetchall()
    result = []
    for o in orders:
        od = dict(o)
        items = conn.execute('''
            SELECT pi.*, p.name as product_name, p.sku, p.size, p.color
            FROM po_items pi
            LEFT JOIN products p ON pi.product_id = p.id
            WHERE pi.po_id = ?
        ''', (o['id'],)).fetchall()
        od['items'] = [dict(i) for i in items]
        result.append(od)
    conn.close()
    return jsonify(result)

@app.route('/api/purchase-orders', methods=['POST'])
@login_required
@admin_required
def create_purchase_order():
    data = request.json
    if not data.get('supplier_id'):
        return jsonify({'success': False, 'error': 'Supplier is required'}), 400
    items = data.get('items', [])
    if not items:
        return jsonify({'success': False, 'error': 'At least one item is required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        shop_id = get_current_shop_id()
        total = sum(item['quantity'] * item['unit_cost'] for item in items)
        status = data.get('status', 'ordered')
        if status not in ('draft', 'ordered'):
            status = 'ordered'
        expected_date = data.get('expected_date') or None
        reference = (data.get('reference') or '').strip() or None

        # Backward-compatible insert: some older DBs may not yet have expected_date/reference/shop_id/po_number.
        po_cols = {row[0] for row in cursor.execute('SHOW COLUMNS FROM purchase_orders').fetchall()}
        candidate_values = {
            'shop_id': shop_id,
            'supplier_id': data['supplier_id'],
            'status': status,
            'total_amount': total,
            'notes': data.get('notes', '').strip(),
            'expected_date': expected_date,
            'reference': reference,
            'created_by': session.get('user_id')
        }
        insert_cols = [c for c in candidate_values.keys() if c in po_cols]
        if 'supplier_id' not in insert_cols or 'status' not in insert_cols or 'total_amount' not in insert_cols:
            return jsonify({'success': False, 'error': 'purchase_orders schema is missing required columns'}), 500

        placeholders = ', '.join(['?'] * len(insert_cols))
        col_sql = ', '.join(insert_cols)
        insert_values = [candidate_values[c] for c in insert_cols]
        cursor.execute(f'INSERT INTO purchase_orders ({col_sql}) VALUES ({placeholders})', insert_values)
        po_id = cursor.lastrowid

        # Generate PO number
        po_number = f'PO-{po_id:05d}'
        if 'po_number' in po_cols:
            cursor.execute('UPDATE purchase_orders SET po_number=? WHERE id=?', (po_number, po_id))

        for item in items:
            if not item.get('product_id') or not item.get('quantity') or item['quantity'] <= 0:
                continue
            cursor.execute('''
                INSERT INTO po_items (po_id, product_id, quantity_ordered, unit_cost)
                VALUES (?, ?, ?, ?)
            ''', (po_id, item['product_id'], item['quantity'], item['unit_cost']))

        conn.commit()
        return jsonify({'success': True, 'id': po_id, 'po_number': po_number}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/purchase-orders/<int:po_id>', methods=['GET'])
@login_required
@admin_required
def get_purchase_order(po_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('po')
    order = conn.execute(f'''
        SELECT po.*, s.name as supplier_name, u.username as created_by_name
        FROM purchase_orders po
        LEFT JOIN suppliers s ON po.supplier_id = s.id
        LEFT JOIN users u ON po.created_by = u.id
        WHERE po.id = ? {flt_sql}
    ''', [po_id] + flt_params).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Purchase order not found'}), 404
    od = dict(order)
    items = conn.execute('''
        SELECT pi.*, p.name as product_name, p.sku, p.size, p.color
        FROM po_items pi
        LEFT JOIN products p ON pi.product_id = p.id
        WHERE pi.po_id = ?
    ''', (po_id,)).fetchall()
    od['items'] = [dict(i) for i in items]
    conn.close()
    return jsonify(od)

@app.route('/api/purchase-orders/<int:po_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_purchase_order(po_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    order = conn.execute(
        f'SELECT status FROM purchase_orders WHERE id=? {flt_sql}', [po_id] + flt_params
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Purchase order not found'}), 404
    if order['status'] in ('received', 'paid'):
        conn.close()
        return jsonify({'success': False, 'error': 'Cannot delete a received/paid purchase order'}), 400
    try:
        conn.execute('DELETE FROM po_items WHERE po_id=?', (po_id,))
        conn.execute('DELETE FROM purchase_orders WHERE id=?', (po_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/purchase-orders/<int:po_id>/submit', methods=['POST'])
@login_required
@admin_required
def submit_purchase_order(po_id):
    """Submit a draft PO — changes status from draft to ordered"""
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    order = conn.execute(
        f'SELECT status FROM purchase_orders WHERE id=? {flt_sql}', [po_id] + flt_params
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Purchase order not found'}), 404
    if order['status'] != 'draft':
        conn.close()
        return jsonify({'success': False, 'error': 'Only draft POs can be submitted'}), 400
    try:
        conn.execute("UPDATE purchase_orders SET status='ordered' WHERE id=?", (po_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/purchase-orders/<int:po_id>/receive', methods=['POST'])
@login_required
@admin_required
def receive_purchase_order(po_id):
    """Receive stock from a purchase order — updates inventory and cost prices"""
    data = request.json
    received_items = data.get('items', [])

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql, flt_params = shop_filter()

    order = cursor.execute(
        f'SELECT * FROM purchase_orders WHERE id=? {flt_sql}', [po_id] + flt_params
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'error': 'Purchase order not found'}), 404
    if order['status'] in ('received', 'paid'):
        conn.close()
        return jsonify({'success': False, 'error': 'This order has already been fully received'}), 400

    try:
        all_fully_received = True
        for ri in received_items:
            po_item = cursor.execute('SELECT * FROM po_items WHERE id=? AND po_id=?', (ri['po_item_id'], po_id)).fetchone()
            if not po_item:
                continue

            qty_receiving = int(_as_float(ri.get('quantity_received', 0)))
            item_received = int(_as_float(po_item.get('quantity_received', 0)))
            item_ordered = int(_as_float(po_item.get('quantity_ordered', 0)))
            if qty_receiving <= 0:
                if item_received < item_ordered:
                    all_fully_received = False
                continue

            new_received = item_received + qty_receiving
            if new_received > item_ordered:
                conn.rollback()
                conn.close()
                return jsonify({'success': False, 'error': f'Cannot receive more than ordered for item {po_item["product_id"]}'}), 400

            # Update po_item received quantity
            cursor.execute('UPDATE po_items SET quantity_received=? WHERE id=?', (new_received, po_item['id']))

            # Get current stock and cost for weighted average calculation
            product = cursor.execute('SELECT stock_quantity, cost_price FROM products WHERE id=?',
                                     (po_item['product_id'],)).fetchone()
            if not product:
                conn.rollback()
                conn.close()
                return jsonify({'success': False, 'error': f'Product not found for PO item {po_item["id"]}'}), 404

            old_stock = int(_as_float(product.get('stock_quantity', 0)))
            old_cost = _as_float(product.get('cost_price', 0))
            new_cost = _as_float(po_item.get('unit_cost', 0))

            # Weighted Average Cost: (old_qty × old_cost + new_qty × new_cost) / total_qty
            total_qty = old_stock + qty_receiving
            if total_qty > 0:
                weighted_avg_cost = round((old_stock * old_cost + qty_receiving * new_cost) / total_qty, 2)
            else:
                weighted_avg_cost = new_cost

            # Update product stock and cost price
            cursor.execute('UPDATE products SET stock_quantity = ?, cost_price = ? WHERE id=?',
                           (total_qty, weighted_avg_cost, po_item['product_id']))

            # Log stock history with cost tracking
            cost_note = f'PO #{po_id} received by {session.get("username")}'
            if old_cost != weighted_avg_cost:
                cost_note += f' | Cost: {old_cost:.2f} → {weighted_avg_cost:.2f} (PO unit cost: {new_cost:.2f})'
            cursor.execute('''
                INSERT INTO stock_history (shop_id, product_id, quantity_change, action_type, note)
                VALUES (?, ?, ?, ?, ?)
            ''', (get_current_shop_id(), po_item['product_id'], qty_receiving, 'purchase', cost_note))

            if new_received < po_item['quantity_ordered']:
                all_fully_received = False

        # Check remaining items not in received_items
        all_items = cursor.execute('SELECT * FROM po_items WHERE po_id=?', (po_id,)).fetchall()
        for item in all_items:
            if int(_as_float(item.get('quantity_received', 0))) < int(_as_float(item.get('quantity_ordered', 0))):
                all_fully_received = False
                break

        # Update PO status
        if all_fully_received:
            cursor.execute("UPDATE purchase_orders SET status='received', received_at=CURRENT_TIMESTAMP WHERE id=?", (po_id,))
        else:
            cursor.execute("UPDATE purchase_orders SET status='partial' WHERE id=?", (po_id,))

        conn.commit()
        return jsonify({'success': True, 'status': 'received' if all_fully_received else 'partial'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/purchase-orders/<int:po_id>/mark-paid', methods=['POST'])
@login_required
@admin_required
def mark_purchase_order_paid(po_id):
    """Mark a fully received purchase order as paid and post cash outflow to finance ledger."""
    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql, flt_params = shop_filter('po')

    order = cursor.execute(
        f'SELECT po.* FROM purchase_orders po WHERE po.id = ? {flt_sql}', [po_id] + flt_params
    ).fetchone()
    if not order:
        conn.close()
        return jsonify({'success': False, 'error': 'Purchase order not found'}), 404

    if order['status'] == 'paid':
        conn.close()
        return jsonify({'success': False, 'error': 'Purchase order is already paid'}), 400
    if order['status'] != 'received':
        conn.close()
        return jsonify({'success': False, 'error': 'Only fully received purchase orders can be marked paid'}), 400

    try:
        cursor.execute("UPDATE purchase_orders SET status='paid' WHERE id=?", (po_id,))
        _record_finance_transaction(
            cursor,
            shop_id=order.get('shop_id'),
            direction='OUT',
            amount=order.get('total_amount'),
            transaction_type='po_payment',
            source_table='purchase_orders',
            source_id=po_id,
            reference=order.get('po_number') or f'PO-{po_id:05d}',
            notes='Purchase order payment',
            created_by=session.get('user_id'),
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Customers & Credit System ---

@app.route('/customers')
@login_required
@admin_required
def customers_page():
    return render_template('customers.html')

@app.route('/api/customers', methods=['GET'])
@login_required
@admin_required
def get_customers():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('c')
    customers = conn.execute(f'''
        SELECT c.*,
            COUNT(DISTINCT i.id) as invoice_count,
            COALESCE(SUM(CASE WHEN i.status != 'paid' THEN i.total_amount - i.paid_amount ELSE 0 END), 0) as total_outstanding,
            COALESCE(SUM(i.total_amount), 0) as lifetime_sales,
            MAX(i.created_at) as last_purchase_date
        FROM customers c
        LEFT JOIN invoices i ON i.customer_id = c.id
        WHERE 1=1 {flt_sql}
        GROUP BY c.id
        ORDER BY c.name
    ''', flt_params).fetchall()
    conn.close()
    return jsonify([dict(c) for c in customers])

@app.route('/api/customers', methods=['POST'])
@login_required
@admin_required
def add_customer():
    data = request.json
    if not data.get('name', '').strip():
        return jsonify({'success': False, 'error': 'Customer name is required'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        shop_id = get_current_shop_id()
        cursor.execute('''
            INSERT INTO customers (shop_id, name, phone, email, address, customer_type, credit_limit, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (shop_id, data['name'].strip(), data.get('phone', '').strip(),
              data.get('email', '').strip(), data.get('address', '').strip(),
              data.get('customer_type', 'walk-in'), float(data.get('credit_limit', 0)),
              data.get('notes', '').strip()))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/customers/<int:customer_id>', methods=['GET'])
@login_required
def get_customer(customer_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    customer = conn.execute(
        f'SELECT * FROM customers WHERE id=? {flt_sql}', [customer_id] + flt_params
    ).fetchone()
    if not customer:
        conn.close()
        return jsonify({'error': 'Customer not found'}), 404

    # Get invoices
    invoices = conn.execute('''
        SELECT i.*, s.sale_date
        FROM invoices i
        LEFT JOIN sales s ON i.sale_id = s.id
        WHERE i.customer_id = ?
        ORDER BY i.created_at DESC
    ''', (customer_id,)).fetchall()

    # Get payment history
    payments = conn.execute('''
        SELECT p.*, u.username as received_by_name
        FROM payments p
        LEFT JOIN users u ON p.received_by = u.id
        WHERE p.customer_id = ?
        ORDER BY p.created_at DESC
    ''', (customer_id,)).fetchall()

    conn.close()
    return jsonify({
        'customer': dict(customer),
        'invoices': [dict(i) for i in invoices],
        'payments': [dict(p) for p in payments]
    })

@app.route('/api/customers/<int:customer_id>', methods=['PUT'])
@login_required
@admin_required
def update_customer(customer_id):
    data = request.json
    if not data.get('name', '').strip():
        return jsonify({'success': False, 'error': 'Customer name is required'}), 400

    conn = get_db_connection()
    try:
        flt_sql, flt_params = shop_filter()
        conn.execute(f'''
            UPDATE customers SET name=?, phone=?, email=?, address=?, customer_type=?, credit_limit=?, notes=?
            WHERE id=? {flt_sql}
        ''', [data['name'].strip(), data.get('phone', '').strip(),
              data.get('email', '').strip(), data.get('address', '').strip(),
              data.get('customer_type', 'walk-in'), float(data.get('credit_limit', 0)),
              data.get('notes', '').strip(), customer_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/customers/<int:customer_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_customer(customer_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    # Check for outstanding invoices
    outstanding = conn.execute(
        f"SELECT COUNT(*) as c FROM invoices WHERE customer_id=? AND status != 'paid' {flt_sql}",
        [customer_id] + flt_params
    ).fetchone()['c']
    if outstanding > 0:
        conn.close()
        return jsonify({'success': False, 'error': f'Cannot delete customer with {outstanding} unpaid invoice(s). Settle all invoices first.'}), 400
    try:
        conn.execute(f'DELETE FROM payments WHERE customer_id=? {flt_sql}', [customer_id] + flt_params)
        conn.execute(f'DELETE FROM invoices WHERE customer_id=? {flt_sql}', [customer_id] + flt_params)
        conn.execute(f'DELETE FROM customers WHERE id=? {flt_sql}', [customer_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# Get wholesale customers for POS dropdown
@app.route('/api/customers/wholesale', methods=['GET'])
@login_required
def get_wholesale_customers():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    customers = conn.execute(f'''
        SELECT id, name, credit_limit, balance
        FROM customers
        WHERE customer_type = 'wholesale' {flt_sql}
        ORDER BY name
    ''', flt_params).fetchall()
    conn.close()
    return jsonify([dict(c) for c in customers])

# --- Payments API ---

@app.route('/api/payments', methods=['POST'])
@login_required
@admin_required
def record_payment():
    data = request.json
    customer_id = data.get('customer_id')
    invoice_id = data.get('invoice_id')
    amount = float(data.get('amount', 0))
    payment_method = data.get('payment_method', 'cash')
    notes = data.get('notes', '').strip()

    if not customer_id or amount <= 0:
        return jsonify({'success': False, 'error': 'Customer and positive amount required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Validate customer exists (scoped to shop)
        flt_sql, flt_params = shop_filter()
        customer = cursor.execute(
            f'SELECT * FROM customers WHERE id=? {flt_sql}', [customer_id] + flt_params
        ).fetchone()
        if not customer:
            return jsonify({'success': False, 'error': 'Customer not found'}), 404

        if invoice_id:
            # Payment against specific invoice
            invoice = cursor.execute(
                f'SELECT * FROM invoices WHERE id=? AND customer_id=? {flt_sql}',
                [invoice_id, customer_id] + flt_params
            ).fetchone()
            if not invoice:
                return jsonify({'success': False, 'error': 'Invoice not found'}), 404
            remaining = invoice['total_amount'] - invoice['paid_amount']
            if amount > remaining:
                return jsonify({'success': False, 'error': f'Amount exceeds invoice balance of {remaining:.2f}'}), 400

            # Update invoice
            new_paid = invoice['paid_amount'] + amount
            new_status = 'paid' if new_paid >= invoice['total_amount'] else 'partial'
            cursor.execute('UPDATE invoices SET paid_amount=?, status=? WHERE id=?',
                           (new_paid, new_status, invoice_id))
        else:
            # General payment - apply to oldest unpaid invoices
            unpaid = cursor.execute(f'''
                SELECT * FROM invoices
                WHERE customer_id=? AND status != 'paid' {flt_sql}
                ORDER BY created_at ASC
            ''', [customer_id] + flt_params).fetchall()

            remaining_amount = amount
            for inv in unpaid:
                if remaining_amount <= 0:
                    break
                inv_remaining = inv['total_amount'] - inv['paid_amount']
                apply_amount = min(remaining_amount, inv_remaining)
                new_paid = inv['paid_amount'] + apply_amount
                new_status = 'paid' if new_paid >= inv['total_amount'] else 'partial'
                cursor.execute('UPDATE invoices SET paid_amount=?, status=? WHERE id=?',
                               (new_paid, new_status, inv['id']))
                remaining_amount -= apply_amount

        # Record the payment
        cursor.execute('''
            INSERT INTO payments (shop_id, customer_id, invoice_id, amount, payment_method, received_by, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (get_current_shop_id(), customer_id, invoice_id, amount, payment_method, session.get('user_id'), notes))
        payment_id = cursor.lastrowid

        _record_finance_transaction(
            cursor,
            shop_id=get_current_shop_id(),
            direction='IN',
            amount=amount,
            transaction_type='invoice_payment',
            source_table='payments',
            source_id=payment_id,
            reference=f'PAY-{payment_id}',
            notes=notes or 'Invoice payment received',
            created_by=session.get('user_id'),
        )

        # Update customer balance
        cursor.execute('UPDATE customers SET balance = balance - ? WHERE id=?', (amount, customer_id))

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Customer Aging Report ---
@app.route('/api/customers/aging-report', methods=['GET'])
@login_required
@admin_required
def customer_aging_report():
    conn = get_db_connection()
    flt_sql_c, flt_params_c = shop_filter('c')
    flt_sql_i, flt_params_i = shop_filter('i')
    customers = conn.execute(f'''
        SELECT c.id, c.name, c.phone, c.credit_limit, c.balance
        FROM customers c
        WHERE c.customer_type = 'wholesale' AND c.balance > 0 {flt_sql_c}
        ORDER BY c.balance DESC
    ''', flt_params_c).fetchall()

    result = []
    for c in customers:
        invoices = conn.execute(f'''
            SELECT i.id, i.total_amount, i.paid_amount, i.status, i.due_date, i.created_at,
                   DATEDIFF(NOW(), i.created_at) as days_old
            FROM invoices i
            WHERE i.customer_id = ? AND i.status != 'paid' {flt_sql_i}
            ORDER BY i.created_at ASC
        ''', [c['id']] + flt_params_i).fetchall()

        current = 0
        days_30 = 0
        days_60 = 0
        days_90_plus = 0
        for inv in invoices:
            outstanding = inv['total_amount'] - inv['paid_amount']
            days = inv['days_old'] or 0
            if days <= 30:
                current += outstanding
            elif days <= 60:
                days_30 += outstanding
            elif days <= 90:
                days_60 += outstanding
            else:
                days_90_plus += outstanding

        result.append({
            'id': c['id'],
            'name': c['name'],
            'phone': c['phone'],
            'credit_limit': c['credit_limit'],
            'balance': c['balance'],
            'current': round(current, 2),
            'days_30': round(days_30, 2),
            'days_60': round(days_60, 2),
            'days_90_plus': round(days_90_plus, 2)
        })

    conn.close()
    return jsonify(result)


# ─── Backup & Restore ────────────────────────────────────────────────────────

@app.route('/api/admin/backup', methods=['GET'])
@login_required
@super_admin_required
def backup_database():
    """Stream a mysqldump of the current database as a downloadable .sql file."""
    import subprocess
    from database import DB_CONFIG, DB_NAME
    from datetime import datetime as _dt
    import io

    host = DB_CONFIG['host']
    port = str(DB_CONFIG['port'])
    user = DB_CONFIG['user']
    password = DB_CONFIG['password']
    db = DB_NAME

    cmd = [
        'mysqldump',
        f'--host={host}',
        f'--port={port}',
        f'--user={user}',
        '--set-gtid-purged=OFF',
        '--single-transaction',
        '--routines',
        '--triggers',
        db
    ]
    if password:
        cmd.insert(4, f'--password={password}')

    try:
        result = subprocess.run(cmd, capture_output=True, timeout=120)
        if result.returncode != 0:
            err = result.stderr.decode('utf-8', errors='replace')
            return jsonify({'error': f'mysqldump failed: {err}'}), 500

        filename = f"backup_{db}_{_dt.now().strftime('%Y%m%d_%H%M%S')}.sql"
        return send_file(
            io.BytesIO(result.stdout),
            mimetype='application/octet-stream',
            as_attachment=True,
            download_name=filename
        )
    except FileNotFoundError:
        return jsonify({'error': 'mysqldump not found. Make sure MySQL client tools are installed.'}), 500
    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Backup timed out.'}), 500


@app.route('/api/admin/restore', methods=['POST'])
@login_required
@super_admin_required
def restore_database():
    """Restore the database from an uploaded .sql file."""
    import subprocess
    import tempfile
    from database import DB_CONFIG, DB_NAME, init_db

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    f = request.files['file']
    if not f.filename.endswith('.sql'):
        return jsonify({'error': 'Only .sql files are accepted'}), 400

    # Read and validate — must look like a MySQL dump
    content = f.read()
    if len(content) == 0:
        return jsonify({'error': 'Uploaded file is empty'}), 400
    header = content[:500].decode('utf-8', errors='replace')
    if 'mysql dump' not in header.lower() and 'mysqldump' not in header.lower():
        return jsonify({'error': 'File does not appear to be a valid MySQL dump'}), 400
    if len(content) > 200 * 1024 * 1024:  # 200 MB cap
        return jsonify({'error': 'File too large (max 200 MB)'}), 400

    host = DB_CONFIG['host']
    port = str(DB_CONFIG['port'])
    user = DB_CONFIG['user']
    password = DB_CONFIG['password']
    db = DB_NAME

    # Drop and recreate the database, then restore
    admin_conn_args = [
        f'--host={host}',
        f'--port={port}',
        f'--user={user}',
    ]
    if password:
        admin_conn_args.append(f'--password={password}')

    try:
        # Drop & recreate
        drop_create = f"DROP DATABASE IF EXISTS `{db}`; CREATE DATABASE `{db}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;"
        drop_result = subprocess.run(
            ['mysql'] + admin_conn_args + ['-e', drop_create],
            capture_output=True, timeout=30
        )
        if drop_result.returncode != 0:
            err = drop_result.stderr.decode('utf-8', errors='replace')
            return jsonify({'error': f'Failed to recreate database: {err}'}), 500

        # Write to temp file and restore
        with tempfile.NamedTemporaryFile(suffix='.sql', delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        restore_result = subprocess.run(
            ['mysql'] + admin_conn_args + [db],
            stdin=open(tmp_path, 'rb'),
            capture_output=True, timeout=300
        )
        os.unlink(tmp_path)

        if restore_result.returncode != 0:
            err = restore_result.stderr.decode('utf-8', errors='replace')
            return jsonify({'error': f'Restore failed: {err}'}), 500

        # Bring older backups up to the current schema before the app serves them.
        init_db()

        return jsonify({'message': 'Database restored successfully. Please log in again.'})

    except FileNotFoundError:
        return jsonify({'error': 'mysql client not found. Make sure MySQL client tools are installed.'}), 500
    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Restore timed out.'}), 500


# ══════════════════════════════════════════════════════════════════════════════
# EXPENSE MANAGEMENT MODULE
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/finance')
@login_required
@admin_required
def finance_page():
    return render_template('finance.html')


@app.route('/api/finance/summary', methods=['GET'])
@login_required
@admin_required
def get_finance_summary():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Keep period-scoped totals for analytics, but expose absolute cash-on-hand
        # so "Current Cash Balance" is stable regardless of selected date range.
        period_summary = _get_available_funds(cursor, date_from=start_date, date_to=end_date)
        absolute_summary = _get_available_funds(cursor)

        summary = dict(period_summary)
        summary['period_available_funds'] = period_summary['available_funds']
        summary['available_funds'] = absolute_summary['available_funds']

        flt_sql, flt_params = shop_filter('ft')
        range_sql = ''
        params = list(flt_params)
        if start_date:
            range_sql += ' AND ft.transaction_at >= ?'
            params.append(f'{start_date} 00:00:00')
        if end_date:
            range_sql += ' AND ft.transaction_at <= ?'
            params.append(f'{end_date} 23:59:59')

        by_type = cursor.execute(
            f'''
                SELECT ft.transaction_type,
                       ft.direction,
                       COUNT(*) as tx_count,
                       COALESCE(SUM(ft.amount), 0) as total_amount
                FROM finance_transactions ft
                WHERE 1=1 {flt_sql} {range_sql}
                GROUP BY ft.transaction_type, ft.direction
                ORDER BY total_amount DESC
            ''',
            params,
        ).fetchall()

        return jsonify({
            'summary': summary,
            'by_type': [dict(row) for row in by_type],
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/finance/opening-balance', methods=['GET', 'PUT'])
@login_required
@admin_required
def manage_finance_opening_balance():
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'GET':
        try:
            opening_balance = _get_opening_balance(cursor)
            return jsonify({'opening_balance': opening_balance})
        finally:
            conn.close()

    data = request.json or {}
    amount = round(_as_float(data.get('opening_balance')), 2)
    if amount < 0:
        conn.close()
        return jsonify({'success': False, 'error': 'Opening balance cannot be negative'}), 400

    shop_id = get_current_shop_id()
    if shop_id is None:
        conn.close()
        return jsonify({'success': False, 'error': 'Select a specific shop before setting opening balance'}), 400

    try:
        existing = cursor.execute(
            "SELECT value FROM settings WHERE `key` = 'finance_opening_balance' AND shop_id = ?",
            (shop_id,),
        ).fetchone()
        previous_amount = _as_float(existing['value'] if existing else 0)

        cursor.execute(
            '''
                INSERT INTO settings (`key`, value, shop_id)
                VALUES ('finance_opening_balance', ?, ?)
                ON DUPLICATE KEY UPDATE value = VALUES(value)
            ''',
            (str(amount), shop_id),
        )

        delta = round(amount - previous_amount, 2)
        _record_finance_audit(
            cursor,
            shop_id=shop_id,
            action_type='opening_balance_updated',
            entity_type='settings',
            entity_id=shop_id,
            amount=amount,
            direction='IN' if delta >= 0 else 'OUT',
            reference='finance_opening_balance',
            notes='Opening balance updated',
            details={
                'previous_opening_balance': previous_amount,
                'new_opening_balance': amount,
                'delta': delta,
            },
            created_by=session.get('user_id'),
        )

        conn.commit()
        return jsonify({'success': True, 'opening_balance': amount})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/finance/transactions', methods=['GET'])
@login_required
@admin_required
def list_finance_transactions():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    direction = (request.args.get('direction') or '').strip().upper()
    transaction_type = (request.args.get('transaction_type') or '').strip()
    page = max(1, int(request.args.get('page', 1) or 1))
    page_size = min(200, max(1, int(request.args.get('page_size', 50) or 50)))
    offset = (page - 1) * page_size

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        flt_sql, flt_params = shop_filter('ft')
        where_sql = ''
        params = list(flt_params)

        if start_date:
            where_sql += ' AND ft.transaction_at >= ?'
            params.append(f'{start_date} 00:00:00')
        if end_date:
            where_sql += ' AND ft.transaction_at <= ?'
            params.append(f'{end_date} 23:59:59')
        if direction in ('IN', 'OUT'):
            where_sql += ' AND ft.direction = ?'
            params.append(direction)
        if transaction_type:
            where_sql += ' AND ft.transaction_type = ?'
            params.append(transaction_type)

        total = cursor.execute(
            f'''SELECT COUNT(*) as cnt FROM finance_transactions ft WHERE 1=1 {flt_sql} {where_sql}''',
            params,
        ).fetchone()['cnt']

        rows = cursor.execute(
            f'''
                SELECT ft.*, u.username, u.full_name,
                       COALESCE(NULLIF(u.full_name, ''), u.username, 'System') as actor_name
                FROM finance_transactions ft
                LEFT JOIN users u ON u.id = ft.created_by
                WHERE 1=1 {flt_sql} {where_sql}
                ORDER BY ft.transaction_at DESC, ft.id DESC
                LIMIT ? OFFSET ?
            ''',
            params + [page_size, offset],
        ).fetchall()

        return jsonify({
            'page': page,
            'page_size': page_size,
            'total': total,
            'transactions': [dict(row) for row in rows],
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/finance/audit', methods=['GET'])
@login_required
@admin_required
def list_finance_audit():
    limit = min(300, max(1, int(request.args.get('limit', 80) or 80)))
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        flt_sql, flt_params = shop_filter('fa')
        where_sql = ''
        params = list(flt_params)

        if start_date:
            where_sql += ' AND fa.created_at >= ?'
            params.append(f'{start_date} 00:00:00')
        if end_date:
            where_sql += ' AND fa.created_at <= ?'
            params.append(f'{end_date} 23:59:59')

        rows = cursor.execute(
            f'''
                SELECT fa.*, u.username, u.full_name,
                       COALESCE(NULLIF(u.full_name, ''), u.username, 'System') as actor_name
                FROM finance_audit_log fa
                LEFT JOIN users u ON u.id = fa.created_by
                WHERE 1=1 {flt_sql} {where_sql}
                ORDER BY fa.created_at DESC, fa.id DESC
                LIMIT ?
            ''',
            params + [limit],
        ).fetchall()

        return jsonify({'audit': [dict(row) for row in rows]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


@app.route('/api/finance/owner-transaction', methods=['POST'])
@login_required
@admin_required
def create_owner_transaction():
    data = request.json or {}
    movement_type = (data.get('movement_type') or '').strip().lower()
    amount = _as_float(data.get('amount'))
    notes = (data.get('notes') or '').strip()
    reference = (data.get('reference') or '').strip()
    transaction_date = (data.get('transaction_date') or '').strip()

    if movement_type not in ('investment', 'withdrawal'):
        return jsonify({'success': False, 'error': 'movement_type must be investment or withdrawal'}), 400
    if amount <= 0:
        return jsonify({'success': False, 'error': 'Amount must be greater than 0'}), 400

    direction = 'IN' if movement_type == 'investment' else 'OUT'
    tx_type = 'owner_investment' if movement_type == 'investment' else 'owner_withdrawal'
    shop_id = get_current_shop_id()

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        _record_finance_transaction(
            cursor,
            shop_id=shop_id,
            direction=direction,
            amount=amount,
            transaction_type=tx_type,
            source_table='manual',
            source_id=None,
            reference=reference or None,
            notes=notes or None,
            created_by=session.get('user_id'),
            transaction_at=transaction_date or None,
        )
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/expenses')
@login_required
@admin_required
def expenses_page():
    return render_template('expenses.html')

# --- Expense Categories API ---

@app.route('/api/expense-categories', methods=['GET'])
@login_required
@admin_required
def get_expense_categories():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('ec')

    # When in all-shops mode (super_admin), deduplicate by name to avoid showing
    # the same category name from multiple shops
    if not flt_sql:
        categories = conn.execute('''
            SELECT MIN(ec.id) as id, ec.name, ec.icon, ec.description, MIN(ec.shop_id) as shop_id,
                   MIN(ec.created_at) as created_at,
                   (SELECT COUNT(*) FROM expenses ex WHERE ex.category_id IN
                       (SELECT id FROM expense_categories WHERE name = ec.name)) as expense_count
            FROM expense_categories ec
            GROUP BY ec.name, ec.icon, ec.description
            ORDER BY ec.name
        ''').fetchall()
    else:
        categories = conn.execute(f'''
            SELECT ec.id, ec.name, ec.icon, ec.description, ec.shop_id, ec.created_at,
                   (SELECT COUNT(*) FROM expenses ex WHERE ex.category_id = ec.id) as expense_count
            FROM expense_categories ec
            WHERE 1=1 {flt_sql}
            ORDER BY ec.name
        ''', flt_params).fetchall()
    conn.close()
    return jsonify([dict(c) for c in categories])

@app.route('/api/expense-categories', methods=['POST'])
@login_required
@admin_required
def add_expense_category():
    data = request.json
    if not data.get('name', '').strip():
        return jsonify({'success': False, 'error': 'Category name is required'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        shop_id = get_current_shop_id()
        cursor.execute('''
            INSERT INTO expense_categories (name, icon, description, shop_id)
            VALUES (?, ?, ?, ?)
        ''', (data['name'].strip(), data.get('icon', '📝').strip(),
              data.get('description', '').strip(), shop_id))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/expense-categories/<int:cat_id>', methods=['PUT'])
@login_required
@admin_required
def update_expense_category(cat_id):
    data = request.json
    if not data.get('name', '').strip():
        return jsonify({'success': False, 'error': 'Category name is required'}), 400

    conn = get_db_connection()
    try:
        flt_sql, flt_params = shop_filter()
        conn.execute(f'''
            UPDATE expense_categories SET name=?, icon=?, description=?
            WHERE id=? {flt_sql}
        ''', [data['name'].strip(), data.get('icon', '📝').strip(),
              data.get('description', '').strip(), cat_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/expense-categories/<int:cat_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_expense_category(cat_id):
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter()
    # Check if category has expenses
    exp_count = conn.execute(
        f'SELECT COUNT(*) as c FROM expenses WHERE category_id=? {flt_sql}',
        [cat_id] + flt_params
    ).fetchone()['c']
    if exp_count > 0:
        conn.close()
        return jsonify({'success': False, 'error': f'Cannot delete category with {exp_count} expense(s). Delete or reassign expenses first.'}), 400
    try:
        conn.execute(f'DELETE FROM expense_categories WHERE id=? {flt_sql}', [cat_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Expenses API ---

@app.route('/api/expenses', methods=['GET'])
@login_required
@admin_required
def get_expenses():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('e')

    # Optional filters
    category_id = request.args.get('category_id')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    search = request.args.get('search', '').strip()

    extra_sql = ''
    extra_params = []

    if category_id:
        extra_sql += ' AND e.category_id = ?'
        extra_params.append(int(category_id))
    if date_from:
        extra_sql += ' AND e.expense_date >= ?'
        extra_params.append(date_from)
    if date_to:
        extra_sql += ' AND e.expense_date <= ?'
        extra_params.append(date_to)
    if search:
        extra_sql += ' AND (e.description LIKE ? OR ec.name LIKE ? OR e.receipt_ref LIKE ?)'
        like_term = f'%{search}%'
        extra_params.extend([like_term, like_term, like_term])

    expenses = conn.execute(f'''
        SELECT e.*, ec.name as category_name, ec.icon as category_icon,
               u.username as created_by_name
        FROM expenses e
        LEFT JOIN expense_categories ec ON e.category_id = ec.id
        LEFT JOIN users u ON e.created_by = u.id
        WHERE 1=1 {flt_sql} {extra_sql}
        ORDER BY e.expense_date DESC, e.created_at DESC
    ''', flt_params + extra_params).fetchall()
    conn.close()
    return jsonify([dict(e) for e in expenses])

@app.route('/api/expenses/summary', methods=['GET'])
@login_required
@admin_required
def get_expenses_summary():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('e')

    today = datetime.now().strftime('%Y-%m-%d')
    first_of_month = datetime.now().strftime('%Y-%m-01')

    # Today's total
    today_total = conn.execute(f'''
        SELECT COALESCE(SUM(e.amount), 0) as total
        FROM expenses e WHERE e.expense_date = ? {flt_sql}
    ''', [today] + flt_params).fetchone()['total']

    # This month's total
    month_total = conn.execute(f'''
        SELECT COALESCE(SUM(e.amount), 0) as total
        FROM expenses e WHERE e.expense_date >= ? {flt_sql}
    ''', [first_of_month] + flt_params).fetchone()['total']

    # Recurring monthly total
    recurring_total = conn.execute(f'''
        SELECT COALESCE(SUM(re.amount), 0) as total
        FROM recurring_expenses re WHERE re.is_active = 1 AND re.frequency = 'monthly' {flt_sql.replace('e.shop_id', 're.shop_id')}
    ''', flt_params).fetchone()['total']

    # Top category this month
    top_category = conn.execute(f'''
        SELECT ec.name, COALESCE(SUM(e.amount), 0) as total
        FROM expenses e
        LEFT JOIN expense_categories ec ON e.category_id = ec.id
        WHERE e.expense_date >= ? {flt_sql}
        GROUP BY ec.id, ec.name
        ORDER BY total DESC LIMIT 1
    ''', [first_of_month] + flt_params).fetchone()

    conn.close()
    return jsonify({
        'today_total': today_total,
        'month_total': month_total,
        'recurring_monthly': recurring_total,
        'top_category': dict(top_category) if top_category else {'name': '-', 'total': 0}
    })

@app.route('/api/expenses/category-breakdown', methods=['GET'])
@login_required
@admin_required
def get_expenses_category_breakdown():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('e')
    first_of_month = datetime.now().strftime('%Y-%m-01')

    breakdown = conn.execute(f'''
        SELECT ec.id, ec.name, ec.icon, COALESCE(SUM(e.amount), 0) as total
        FROM expenses e
        LEFT JOIN expense_categories ec ON e.category_id = ec.id
        WHERE e.expense_date >= ? {flt_sql}
        GROUP BY ec.id, ec.name, ec.icon
        ORDER BY total DESC
    ''', [first_of_month] + flt_params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in breakdown])

@app.route('/api/expenses', methods=['POST'])
@login_required
@admin_required
def add_expense():
    data = request.json
    if not data.get('category_id'):
        return jsonify({'success': False, 'error': 'Category is required'}), 400
    if not data.get('amount') or float(data['amount']) <= 0:
        return jsonify({'success': False, 'error': 'Valid amount is required'}), 400
    if not data.get('expense_date'):
        return jsonify({'success': False, 'error': 'Date is required'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        shop_id = get_current_shop_id()
        cursor.execute('''
            INSERT INTO expenses (category_id, amount, description, expense_date, payment_method, receipt_ref, created_by, shop_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (int(data['category_id']), float(data['amount']),
              data.get('description', '').strip(),
              data['expense_date'],
              data.get('payment_method', 'cash'),
              data.get('receipt_ref', '').strip(),
              session.get('user_id'),
              shop_id))
        expense_id = cursor.lastrowid

        _record_finance_transaction(
            cursor,
            shop_id=shop_id,
            direction='OUT',
            amount=float(data['amount']),
            transaction_type='expense_payment',
            source_table='expenses',
            source_id=expense_id,
            reference=f'EXP-{expense_id}',
            notes=data.get('description', '').strip(),
            created_by=session.get('user_id'),
            transaction_at=data['expense_date'],
        )
        conn.commit()
        return jsonify({'success': True, 'id': expense_id}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/expenses/<int:expense_id>', methods=['PUT'])
@login_required
@admin_required
def update_expense(expense_id):
    data = request.json
    if not data.get('category_id'):
        return jsonify({'success': False, 'error': 'Category is required'}), 400
    if not data.get('amount') or float(data['amount']) <= 0:
        return jsonify({'success': False, 'error': 'Valid amount is required'}), 400
    if not data.get('expense_date'):
        return jsonify({'success': False, 'error': 'Date is required'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        flt_sql, flt_params = shop_filter('e')

        expense_row = cursor.execute(
            f'''SELECT e.id, e.shop_id FROM expenses e WHERE e.id=? {flt_sql}''',
            [expense_id] + flt_params,
        ).fetchone()
        if not expense_row:
            return jsonify({'success': False, 'error': 'Expense not found'}), 404

        amount = round(float(data['amount']), 2)
        description = data.get('description', '').strip()
        expense_date = data['expense_date']
        payment_method = data.get('payment_method', 'cash')
        receipt_ref = data.get('receipt_ref', '').strip()

        cursor.execute(
            '''
            UPDATE expenses
            SET category_id=?, amount=?, description=?, expense_date=?, payment_method=?, receipt_ref=?
            WHERE id=?
            ''',
            [
                int(data['category_id']),
                amount,
                description,
                expense_date,
                payment_method,
                receipt_ref,
                expense_id,
            ],
        )

        # Keep linked ledger row in sync so available funds stays accurate.
        cursor.execute(
            '''
            UPDATE finance_transactions
            SET amount=?, transaction_at=?, notes=?, reference=?
            WHERE source_table='expenses'
              AND source_id=?
              AND shop_id=?
              AND transaction_type='expense_payment'
              AND direction='OUT'
            ''',
            [amount, expense_date, description or None, f'EXP-{expense_id}', expense_id, expense_row['shop_id']],
        )

        # Self-heal older/missing ledger links for existing expenses.
        if cursor.rowcount == 0:
            _record_finance_transaction(
                cursor,
                shop_id=expense_row['shop_id'],
                direction='OUT',
                amount=amount,
                transaction_type='expense_payment',
                source_table='expenses',
                source_id=expense_id,
                reference=f'EXP-{expense_id}',
                notes=description,
                created_by=session.get('user_id'),
                transaction_at=expense_date,
            )

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_expense(expense_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        flt_sql, flt_params = shop_filter('e')

        expense_row = cursor.execute(
            f'''SELECT e.id, e.shop_id FROM expenses e WHERE e.id=? {flt_sql}''',
            [expense_id] + flt_params,
        ).fetchone()
        if not expense_row:
            return jsonify({'success': False, 'error': 'Expense not found'}), 404

        cursor.execute('DELETE FROM expenses WHERE id=?', [expense_id])

        # Remove linked OUT movement so available funds restores on delete.
        cursor.execute(
            '''
            DELETE FROM finance_transactions
            WHERE source_table='expenses'
              AND source_id=?
              AND shop_id=?
              AND transaction_type='expense_payment'
              AND direction='OUT'
            ''',
            [expense_id, expense_row['shop_id']],
        )

        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

# --- Recurring Expenses API ---

@app.route('/api/recurring-expenses', methods=['GET'])
@login_required
@admin_required
def get_recurring_expenses():
    conn = get_db_connection()
    flt_sql, flt_params = shop_filter('re')
    recurring = conn.execute(f'''
        SELECT re.*, ec.name as category_name, ec.icon as category_icon
        FROM recurring_expenses re
        LEFT JOIN expense_categories ec ON re.category_id = ec.id
        WHERE 1=1 {flt_sql}
        ORDER BY re.next_due_date ASC
    ''', flt_params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in recurring])

@app.route('/api/recurring-expenses', methods=['POST'])
@login_required
@admin_required
def add_recurring_expense():
    data = request.json
    if not data.get('category_id'):
        return jsonify({'success': False, 'error': 'Category is required'}), 400
    if not data.get('amount') or float(data['amount']) <= 0:
        return jsonify({'success': False, 'error': 'Valid amount is required'}), 400
    if not data.get('next_due_date'):
        return jsonify({'success': False, 'error': 'Next due date is required'}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        shop_id = get_current_shop_id()
        cursor.execute('''
            INSERT INTO recurring_expenses (category_id, amount, description, frequency, next_due_date, is_active, shop_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (int(data['category_id']), float(data['amount']),
              data.get('description', '').strip(),
              data.get('frequency', 'monthly'),
              data['next_due_date'],
              1, shop_id))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/recurring-expenses/<int:rec_id>', methods=['PUT'])
@login_required
@admin_required
def update_recurring_expense(rec_id):
    data = request.json
    if not data.get('category_id'):
        return jsonify({'success': False, 'error': 'Category is required'}), 400
    if not data.get('amount') or float(data['amount']) <= 0:
        return jsonify({'success': False, 'error': 'Valid amount is required'}), 400
    if not data.get('next_due_date'):
        return jsonify({'success': False, 'error': 'Next due date is required'}), 400

    conn = get_db_connection()
    try:
        flt_sql, flt_params = shop_filter()
        conn.execute(f'''
            UPDATE recurring_expenses SET category_id=?, amount=?, description=?, frequency=?, next_due_date=?, is_active=?
            WHERE id=? {flt_sql}
        ''', [int(data['category_id']), float(data['amount']),
              data.get('description', '').strip(),
              data.get('frequency', 'monthly'),
              data['next_due_date'],
              1 if data.get('is_active', True) else 0,
              rec_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/recurring-expenses/<int:rec_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_recurring_expense(rec_id):
    conn = get_db_connection()
    try:
        flt_sql, flt_params = shop_filter()
        conn.execute(f'DELETE FROM recurring_expenses WHERE id=? {flt_sql}', [rec_id] + flt_params)
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/recurring-expenses/<int:rec_id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_recurring_expense(rec_id):
    conn = get_db_connection()
    try:
        flt_sql, flt_params = shop_filter()
        current = conn.execute(
            f'SELECT is_active FROM recurring_expenses WHERE id=? {flt_sql}', [rec_id] + flt_params
        ).fetchone()
        if not current:
            conn.close()
            return jsonify({'success': False, 'error': 'Not found'}), 404
        new_state = 0 if current['is_active'] else 1
        conn.execute(f'UPDATE recurring_expenses SET is_active=? WHERE id=? {flt_sql}', [new_state, rec_id] + flt_params)
        conn.commit()
        return jsonify({'success': True, 'is_active': bool(new_state)})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()


# ── Manual Recurring Expense Trigger ─────────────────────────────────────────

@app.route('/api/scheduler/process-recurring', methods=['POST'])
@admin_required
def process_recurring_now():
    """Manually trigger recurring expense processing for testing/immediate use."""
    try:
        scheduled_recurring_expenses()
        return jsonify({'success': True, 'message': 'Recurring expenses processed successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── Profit & Loss Report ─────────────────────────────────────────────────────

@app.route('/api/reports/pnl', methods=['GET'])
@login_required
@admin_required
def get_pnl_report():
    """Generate Profit & Loss report showing gross profit, expenses, and net profit."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql_s, flt_params_s = shop_filter('s')
    flt_sql_e, flt_params_e = shop_filter('e')
    flt_sql_sr, flt_params_sr = shop_filter('sr')

    try:
        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # ─── Revenue ───
        revenue_data = cursor.execute(f'''
            SELECT COALESCE(SUM(total_amount), 0) as total_revenue,
                   COALESCE(SUM(total_cost), 0) as total_cogs,
                   COUNT(*) as transaction_count
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
        ''', [start_date_time, end_date_time] + flt_params_s).fetchone()

        total_revenue = float(revenue_data['total_revenue'] or 0)
        total_cogs = float(revenue_data['total_cogs'] or 0)
        transaction_count = revenue_data['transaction_count']
        gross_profit = total_revenue - total_cogs

        # ─── Refunds (reduce revenue) ───
        flt_sql_plain, flt_params_plain = shop_filter()
        refund_data = cursor.execute(f'''
            SELECT COALESCE(SUM(refund_amount), 0) as total_refunds
            FROM sale_returns WHERE return_date BETWEEN ? AND ? {flt_sql_plain}
        ''', [start_date_time, end_date_time] + flt_params_plain).fetchone()
        total_refunds = float(refund_data['total_refunds'] or 0)

        refunded_cogs_data = cursor.execute(f'''
            SELECT COALESCE(SUM(
                CASE
                    WHEN COALESCE(s.total_amount, 0) > 0
                        THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                    ELSE 0
                END
            ), 0) as refunded_cogs
            FROM sale_returns sr
            JOIN sales s ON s.id = sr.sale_id
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
        ''', [start_date_time, end_date_time] + flt_params_sr).fetchone()
        total_refunded_cogs = float(refunded_cogs_data['refunded_cogs'] or 0)

        # ─── Discounts ───
        discount_data = cursor.execute(f'''
            SELECT COALESCE(SUM(discount_amount), 0) as total_discounts
            FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
            AND discount_amount > 0
        ''', [start_date_time, end_date_time] + flt_params_s).fetchone()
        total_discounts = float(discount_data['total_discounts'] or 0)

        # Net revenue = revenue - refunds (discounts already factored in total_amount)
        net_revenue = total_revenue - total_refunds
        net_cogs = total_cogs - total_refunded_cogs
        gross_profit = net_revenue - net_cogs

        # ─── Expenses by Category ───
        expenses_by_category = cursor.execute(f'''
            SELECT ec.name as category_name, ec.icon as category_icon,
                   COALESCE(SUM(e.amount), 0) as total_amount,
                   COUNT(e.id) as expense_count
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE e.expense_date BETWEEN ? AND ? {flt_sql_e}
            GROUP BY ec.id, ec.name, ec.icon
            ORDER BY total_amount DESC
        ''', [start_date, end_date] + flt_params_e).fetchall()

        total_expenses = sum(float(row['total_amount']) for row in expenses_by_category)

        # ─── Net Profit ───
        net_profit = gross_profit - total_expenses

        # ─── Daily P&L Breakdown ───
        daily_revenue = cursor.execute(f'''
            SELECT DATE(sale_date) as the_date,
                   COALESCE(SUM(total_amount), 0) as revenue,
                   COALESCE(SUM(total_cost), 0) as cogs
            FROM sales s
            WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
            GROUP BY DATE(sale_date)
            ORDER BY the_date ASC
        ''', [start_date_time, end_date_time] + flt_params_s).fetchall()

        daily_expenses = cursor.execute(f'''
            SELECT expense_date as the_date,
                   COALESCE(SUM(amount), 0) as expenses
            FROM expenses e
            WHERE expense_date BETWEEN ? AND ? {flt_sql_e}
            GROUP BY expense_date
            ORDER BY the_date ASC
        ''', [start_date, end_date] + flt_params_e).fetchall()

        daily_refunds = cursor.execute(f'''
            SELECT
                DATE(sr.return_date) as the_date,
                COALESCE(SUM(sr.refund_amount), 0) as refunds,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(s.total_amount, 0) > 0
                            THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                        ELSE 0
                    END
                ), 0) as refunded_cogs
            FROM sale_returns sr
            JOIN sales s ON s.id = sr.sale_id
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
            GROUP BY DATE(sr.return_date)
            ORDER BY the_date ASC
        ''', [start_date_time, end_date_time] + flt_params_sr).fetchall()

        # Merge daily data
        daily_map = {}
        for row in daily_revenue:
            d = str(row['the_date'])
            daily_map[d] = {
                'date': d,
                'revenue': float(row['revenue']),
                'cogs': float(row['cogs']),
                'gross_profit': float(row['revenue']) - float(row['cogs']),
                'expenses': 0,
                'net_profit': 0
            }
        for row in daily_expenses:
            d = str(row['the_date'])
            if d not in daily_map:
                daily_map[d] = {'date': d, 'revenue': 0, 'cogs': 0, 'gross_profit': 0, 'expenses': 0, 'net_profit': 0}
            daily_map[d]['expenses'] = float(row['expenses'])

        for row in daily_refunds:
            d = str(row['the_date'])
            if d not in daily_map:
                daily_map[d] = {'date': d, 'revenue': 0, 'cogs': 0, 'gross_profit': 0, 'expenses': 0, 'net_profit': 0}
            daily_map[d]['revenue'] -= float(row['refunds'] or 0)
            daily_map[d]['cogs'] -= float(row['refunded_cogs'] or 0)
            daily_map[d]['gross_profit'] = daily_map[d]['revenue'] - daily_map[d]['cogs']

        for d in daily_map:
            daily_map[d]['net_profit'] = daily_map[d]['gross_profit'] - daily_map[d]['expenses']

        daily_pnl = sorted(daily_map.values(), key=lambda x: x['date'])

        # ─── Monthly P&L (if period > 31 days) ───
        from datetime import datetime as dt_cls, timedelta
        sd = dt_cls.strptime(start_date, '%Y-%m-%d')
        ed = dt_cls.strptime(end_date, '%Y-%m-%d')
        period_days = (ed - sd).days + 1

        monthly_pnl = []
        if period_days > 31:
            monthly_rev = cursor.execute(f'''
                SELECT DATE_FORMAT(sale_date, '%Y-%m') as month,
                       COALESCE(SUM(total_amount), 0) as revenue,
                       COALESCE(SUM(total_cost), 0) as cogs
                FROM sales s
                WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
                GROUP BY DATE_FORMAT(sale_date, '%Y-%m')
                ORDER BY month ASC
            ''', [start_date_time, end_date_time] + flt_params_s).fetchall()

            monthly_exp = cursor.execute(f'''
                SELECT DATE_FORMAT(expense_date, '%Y-%m') as month,
                       COALESCE(SUM(amount), 0) as expenses
                FROM expenses e
                WHERE expense_date BETWEEN ? AND ? {flt_sql_e}
                GROUP BY DATE_FORMAT(expense_date, '%Y-%m')
                ORDER BY month ASC
            ''', [start_date, end_date] + flt_params_e).fetchall()

            monthly_refunds = cursor.execute(f'''
                SELECT
                    DATE_FORMAT(sr.return_date, '%Y-%m') as month,
                    COALESCE(SUM(sr.refund_amount), 0) as refunds,
                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(s.total_amount, 0) > 0
                                THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                            ELSE 0
                        END
                    ), 0) as refunded_cogs
                FROM sale_returns sr
                JOIN sales s ON s.id = sr.sale_id
                WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
                GROUP BY DATE_FORMAT(sr.return_date, '%Y-%m')
                ORDER BY month ASC
            ''', [start_date_time, end_date_time] + flt_params_sr).fetchall()

            monthly_map = {}
            for row in monthly_rev:
                m = row['month']
                monthly_map[m] = {
                    'month': m,
                    'revenue': float(row['revenue']),
                    'cogs': float(row['cogs']),
                    'gross_profit': float(row['revenue']) - float(row['cogs']),
                    'expenses': 0, 'net_profit': 0
                }
            for row in monthly_exp:
                m = row['month']
                if m not in monthly_map:
                    monthly_map[m] = {'month': m, 'revenue': 0, 'cogs': 0, 'gross_profit': 0, 'expenses': 0, 'net_profit': 0}
                monthly_map[m]['expenses'] = float(row['expenses'])
            for row in monthly_refunds:
                m = row['month']
                if m not in monthly_map:
                    monthly_map[m] = {'month': m, 'revenue': 0, 'cogs': 0, 'gross_profit': 0, 'expenses': 0, 'net_profit': 0}
                monthly_map[m]['revenue'] -= float(row['refunds'] or 0)
                monthly_map[m]['cogs'] -= float(row['refunded_cogs'] or 0)
                monthly_map[m]['gross_profit'] = monthly_map[m]['revenue'] - monthly_map[m]['cogs']
            for m in monthly_map:
                monthly_map[m]['net_profit'] = monthly_map[m]['gross_profit'] - monthly_map[m]['expenses']
            monthly_pnl = sorted(monthly_map.values(), key=lambda x: x['month'])

        # ─── Period comparison ───
        prev_end = sd - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_days - 1)
        prev_start_str = prev_start.strftime('%Y-%m-%d') + ' 00:00:00'
        prev_end_str = prev_end.strftime('%Y-%m-%d') + ' 23:59:59'

        prev_rev = cursor.execute(f'''
            SELECT COALESCE(SUM(total_amount), 0) as revenue,
                   COALESCE(SUM(total_cost), 0) as cogs
            FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
        ''', [prev_start_str, prev_end_str] + flt_params_s).fetchone()

        prev_refunds = cursor.execute(f'''
            SELECT
                COALESCE(SUM(sr.refund_amount), 0) as refunds,
                COALESCE(SUM(
                    CASE
                        WHEN COALESCE(s.total_amount, 0) > 0
                            THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                        ELSE 0
                    END
                ), 0) as refunded_cogs
            FROM sale_returns sr
            JOIN sales s ON s.id = sr.sale_id
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_sr}
        ''', [prev_start_str, prev_end_str] + flt_params_sr).fetchone()

        prev_exp = cursor.execute(f'''
            SELECT COALESCE(SUM(amount), 0) as expenses
            FROM expenses e WHERE expense_date BETWEEN ? AND ? {flt_sql_e}
        ''', [prev_start.strftime('%Y-%m-%d'), prev_end.strftime('%Y-%m-%d')] + flt_params_e).fetchone()

        prev_revenue = float(prev_rev['revenue'] or 0)
        prev_cogs = float(prev_rev['cogs'] or 0)
        prev_refunded_amt = float(prev_refunds['refunds'] or 0)
        prev_refunded_cogs = float(prev_refunds['refunded_cogs'] or 0)
        prev_gross = (prev_revenue - prev_refunded_amt) - (prev_cogs - prev_refunded_cogs)
        prev_expenses = float(prev_exp['expenses'] or 0)
        prev_net = prev_gross - prev_expenses

        conn.close()

        # Margins
        gross_margin = (gross_profit / net_revenue * 100) if net_revenue > 0 else 0
        net_margin = (net_profit / net_revenue * 100) if net_revenue > 0 else 0

        return jsonify({
            'success': True,
            'summary': {
                'total_revenue': total_revenue,
                'total_refunds': total_refunds,
                'net_revenue': net_revenue,
                'total_cogs': net_cogs,
                'gross_profit': gross_profit,
                'gross_margin': gross_margin,
                'total_expenses': total_expenses,
                'net_profit': net_profit,
                'net_margin': net_margin,
                'total_discounts': total_discounts,
                'transaction_count': transaction_count
            },
            'expenses_by_category': [dict(row) for row in expenses_by_category],
            'daily_pnl': daily_pnl,
            'monthly_pnl': monthly_pnl,
            'period_comparison': {
                'prev_start': prev_start.strftime('%Y-%m-%d'),
                'prev_end': prev_end.strftime('%Y-%m-%d'),
                'prev_revenue': prev_revenue,
                'prev_gross_profit': prev_gross,
                'prev_expenses': prev_expenses,
                'prev_net_profit': prev_net
            }
        })

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reports/pnl/export/<format>', methods=['GET'])
@login_required
@admin_required
def export_pnl_report(format):
    """Export P&L report to Excel."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        return jsonify({'success': False, 'error': 'Start and end dates are required'}), 400
    if format != 'excel':
        return jsonify({'success': False, 'error': 'Only Excel export is supported'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    flt_sql_s, flt_params_s = shop_filter('s')
    flt_sql_e, flt_params_e = shop_filter('e')

    try:
        shop_id = get_current_shop_id() or session.get('shop_id')
        shop = cursor.execute('SELECT name, currency_symbol FROM shops WHERE id = ?', (shop_id,)).fetchone()
        shop_name = shop['name'] if shop else 'POS System'
        currency = shop['currency_symbol'] if shop else '$'

        end_date_time = end_date + ' 23:59:59'
        start_date_time = start_date + ' 00:00:00'

        # Revenue
        rev = cursor.execute(f'''
            SELECT COALESCE(SUM(total_amount), 0) as revenue, COALESCE(SUM(total_cost), 0) as cogs
            FROM sales s WHERE sale_date BETWEEN ? AND ? {flt_sql_s}
        ''', [start_date_time, end_date_time] + flt_params_s).fetchone()
        total_revenue = float(rev['revenue'] or 0)
        total_cogs = float(rev['cogs'] or 0)
        gross_profit = total_revenue - total_cogs

        # Refunds
        flt_sql_plain, flt_params_plain = shop_filter()
        refunds = cursor.execute(f'''
            SELECT COALESCE(SUM(refund_amount), 0) as total
            FROM sale_returns WHERE return_date BETWEEN ? AND ? {flt_sql_plain}
        ''', [start_date_time, end_date_time] + flt_params_plain).fetchone()
        total_refunds = float(refunds['total'] or 0)

        refunded_cogs_data = cursor.execute(f'''
            SELECT COALESCE(SUM(
                CASE
                    WHEN COALESCE(s.total_amount, 0) > 0
                        THEN LEAST(sr.refund_amount, s.total_amount) * (COALESCE(s.total_cost, 0) / s.total_amount)
                    ELSE 0
                END
            ), 0) as refunded_cogs
            FROM sale_returns sr
            JOIN sales s ON s.id = sr.sale_id
            WHERE sr.return_date BETWEEN ? AND ? {flt_sql_s}
        ''', [start_date_time, end_date_time] + flt_params_s).fetchone()
        total_refunded_cogs = float(refunded_cogs_data['refunded_cogs'] or 0)

        net_revenue = total_revenue - total_refunds
        net_cogs = total_cogs - total_refunded_cogs
        gross_profit = net_revenue - net_cogs

        # Expenses by category
        exp_cats = cursor.execute(f'''
            SELECT ec.name as category_name, COALESCE(SUM(e.amount), 0) as total_amount
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE e.expense_date BETWEEN ? AND ? {flt_sql_e}
            GROUP BY ec.id, ec.name ORDER BY total_amount DESC
        ''', [start_date, end_date] + flt_params_e).fetchall()
        total_expenses = sum(float(row['total_amount']) for row in exp_cats)
        net_profit = gross_profit - total_expenses

        conn.close()

        gross_margin = (gross_profit / net_revenue * 100) if net_revenue > 0 else 0
        net_margin = (net_profit / net_revenue * 100) if net_revenue > 0 else 0

        pnl_data = {
            'total_revenue': total_revenue,
            'total_refunds': total_refunds,
            'net_revenue': net_revenue,
            'total_cogs': net_cogs,
            'gross_profit': gross_profit,
            'gross_margin': gross_margin,
            'total_expenses': total_expenses,
            'net_profit': net_profit,
            'net_margin': net_margin,
            'expenses_by_category': [dict(row) for row in exp_cats]
        }

        return generate_pnl_excel(shop_name, currency, start_date, end_date, pnl_data)

    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


def generate_pnl_pdf(shop_name, currency, start_date, end_date, data):
    """Generate PDF for Profit & Loss report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=20,
        textColor=colors.HexColor('#2c3e50'), spaceAfter=30, alignment=1
    )
    elements.append(Paragraph(f"{shop_name}", title_style))
    elements.append(Paragraph("Profit & Loss Statement", styles['Heading2']))
    elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))

    # P&L Statement Table
    elements.append(Paragraph("Income", styles['Heading3']))
    pnl_rows = [
        ['', 'Amount'],
        ['Total Revenue (Sales)', f"{currency}{data['total_revenue']:,.2f}"],
        ['Less: Refunds', f"({currency}{data['total_refunds']:,.2f})"],
        ['Net Sales', f"{currency}{data['net_revenue']:,.2f}"],
    ]
    t = Table(pnl_rows, colWidths=[3.5*inch, 2.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.2*inch))

    # COGS & Gross Profit
    elements.append(Paragraph("Cost of Goods Sold", styles['Heading3']))
    cogs_rows = [
        ['', 'Amount'],
        ['Cost of Goods Sold (COGS)', f"({currency}{data['total_cogs']:,.2f})"],
        ['Gross Profit', f"{currency}{data['gross_profit']:,.2f}"],
        ['Gross Margin', f"{data['gross_margin']:.1f}%"],
    ]
    t2 = Table(cogs_rows, colWidths=[3.5*inch, 2.5*inch])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, 2), (1, 2), colors.HexColor('#27ae60')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 0.2*inch))

    # Operating Expenses
    elements.append(Paragraph("Operating Expenses", styles['Heading3']))
    exp_rows = [['Category', 'Amount']]
    for cat in data['expenses_by_category']:
        exp_rows.append([cat['category_name'] or 'Uncategorized', f"{currency}{float(cat['total_amount']):,.2f}"])
    exp_rows.append(['Total Expenses', f"({currency}{data['total_expenses']:,.2f})"])

    t3 = Table(exp_rows, colWidths=[3.5*inch, 2.5*inch])
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, -1), (1, -1), colors.HexColor('#e74c3c')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]
    t3.setStyle(TableStyle(style_cmds))
    elements.append(t3)
    elements.append(Spacer(1, 0.3*inch))

    # Net Profit Summary
    elements.append(Paragraph("Net Profit", styles['Heading3']))
    net_color = colors.HexColor('#27ae60') if data['net_profit'] >= 0 else colors.HexColor('#e74c3c')
    net_rows = [
        ['', 'Amount'],
        ['Gross Profit', f"{currency}{data['gross_profit']:,.2f}"],
        ['Less: Total Expenses', f"({currency}{data['total_expenses']:,.2f})"],
        ['Net Profit', f"{currency}{data['net_profit']:,.2f}"],
        ['Net Margin', f"{data['net_margin']:.1f}%"],
    ]
    t4 = Table(net_rows, colWidths=[3.5*inch, 2.5*inch])
    t4.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, -2), (1, -2), net_color),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t4)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=f'pnl_report_{start_date}_to_{end_date}.pdf')


def generate_pnl_excel(shop_name, currency, start_date, end_date, data):
    """Generate Excel for Profit & Loss report"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Profit & Loss'

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    bold_font = Font(bold=True, size=11)
    green_font = Font(bold=True, size=11, color='27AE60')
    red_font = Font(bold=True, size=11, color='E74C3C')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = f'{shop_name} - Profit & Loss Statement'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f'Period: {start_date} to {end_date}'
    ws['A2'].font = Font(size=11, italic=True)

    row = 4
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # Income Section
    ws.cell(row=row, column=1, value='INCOME').font = bold_font
    row += 1
    ws.cell(row=row, column=1, value='Total Revenue (Sales)')
    ws.cell(row=row, column=2, value=data['total_revenue']).number_format = f'"{currency}"#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value='Less: Refunds')
    ws.cell(row=row, column=2, value=-data['total_refunds']).number_format = f'"{currency}"#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value='Net Sales').font = bold_font
    ws.cell(row=row, column=2, value=data['net_revenue']).number_format = f'"{currency}"#,##0.00'
    ws.cell(row=row, column=2).font = bold_font
    row += 2

    # COGS Section
    ws.cell(row=row, column=1, value='COST OF GOODS SOLD').font = bold_font
    row += 1
    ws.cell(row=row, column=1, value='Cost of Goods Sold (COGS)')
    ws.cell(row=row, column=2, value=-data['total_cogs']).number_format = f'"{currency}"#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value='Gross Profit').font = bold_font
    ws.cell(row=row, column=2, value=data['gross_profit']).number_format = f'"{currency}"#,##0.00'
    ws.cell(row=row, column=2).font = green_font
    row += 1
    ws.cell(row=row, column=1, value='Gross Margin')
    ws.cell(row=row, column=2, value=f"{data['gross_margin']:.1f}%")
    row += 2

    # Expenses Section
    ws.cell(row=row, column=1, value='OPERATING EXPENSES').font = bold_font
    row += 1
    for cat in data['expenses_by_category']:
        ws.cell(row=row, column=1, value=cat['category_name'] or 'Uncategorized')
        ws.cell(row=row, column=2, value=float(cat['total_amount'])).number_format = f'"{currency}"#,##0.00'
        row += 1
    ws.cell(row=row, column=1, value='Total Expenses').font = bold_font
    ws.cell(row=row, column=2, value=data['total_expenses']).number_format = f'"{currency}"#,##0.00'
    ws.cell(row=row, column=2).font = red_font
    row += 2

    # Net Profit Section
    ws.cell(row=row, column=1, value='NET PROFIT').font = Font(bold=True, size=13)
    row += 1
    ws.cell(row=row, column=1, value='Gross Profit')
    ws.cell(row=row, column=2, value=data['gross_profit']).number_format = f'"{currency}"#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value='Less: Total Expenses')
    ws.cell(row=row, column=2, value=-data['total_expenses']).number_format = f'"{currency}"#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value='Net Profit').font = Font(bold=True, size=12)
    net_font = green_font if data['net_profit'] >= 0 else red_font
    ws.cell(row=row, column=2, value=data['net_profit']).number_format = f'"{currency}"#,##0.00'
    ws.cell(row=row, column=2).font = net_font
    row += 1
    ws.cell(row=row, column=1, value='Net Margin')
    ws.cell(row=row, column=2, value=f"{data['net_margin']:.1f}%")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'pnl_report_{start_date}_to_{end_date}.xlsx')


if __name__ == '__main__':
    debug = os.environ.get('FLASK_DEBUG', '0').lower() in ('1', 'true')
    port = int(os.environ.get('PORT', '5000'))
    host = os.environ.get('HOST', '0.0.0.0')
    app.run(debug=debug, host=host, port=port)
